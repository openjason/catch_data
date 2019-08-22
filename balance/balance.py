from openpyxl import load_workbook
from sys import argv
from sqlite3 import connect as sqlite3connect
import logging
from logging.handlers import RotatingFileHandler


data_dir = 'f:\\dev\\kefu\\'

def set_logging():
    global logger
    logger = logging.getLogger('balance_logger')
    handler = RotatingFileHandler('balance.log', maxBytes=5000000, backupCount=6)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)
    formatter = logging.Formatter('%(asctime)-12s %(filename)s %(lineno)d %(message)s')
    handler.setFormatter(formatter)

class Balance:
    def __init__(self):
        self.customers = ['jtyh', 'gsnx']
        self.sqlconn = sqlite3connect( "balance_jtyh.db3")
        self.sqlconn.isolation_level = None  # 这个就是事务隔离级别，默认是需要自己commit才能修改数据库，置为None则自动每次修改都提交,否则为""
        self.Holiday = ['2019-02-04', '2019-02-05', '2019-02-06', '2019-04-05',\
'2019-05-01', '2019-06-07', '2019-09-13', '2019-10-01', '2019-10-02', '2019-10-03']

    def sum_from_db(self, customer, curr_month, count_method):
        sqlselect = self.sqlconn.cursor()
        curr_month_holiday = []
        for i in self.Holiday:
            if curr_month in i:
                curr_month_holiday.append(i)

        str_sql = "SELECT wuliao,name from wuliao where kehu='" + customer + "'"
        #print(str_sql)
        sqlcursor = sqlselect.execute(str_sql)
        curr_wuliao_list = []
        for row in sqlcursor:
            #print("wuliao,name = ", row[0],row[1])
            curr_wuliao_list.append([row[0],row[1]])

        #print(curr_wuliao_list)
        list_return_from_db = []

#空白卡请求处理
        if count_method == 'kongbaika':
            for curr_wuliao in curr_wuliao_list:
                str_sql = "SELECT sum(fachu),sum(chengpin),sum(jianka),sum(feika) from days \
where kehu='" + customer +"' and wuliao = '" + str(curr_wuliao[0]) + "'"
                #print(str_sql)
                sqlcursor = sqlselect.execute(str_sql)
                for row in sqlcursor:
                    if row[0] > 0:
                        #print("fachu = ", row[0])
                        list_return_from_db.append([curr_wuliao[1],row[0]])
            return (list_return_from_db)

#个人化请求处理
        elif count_method == 'gerenhua':
            for curr_wuliao in curr_wuliao_list:
                str_sql = "SELECT sum(fachu),sum(chengpin),sum(jianka),sum(feika) from days \
where kehu='" + customer + "' and wuliao = '" + str(curr_wuliao[0]) + "'"
                sqlcursor = sqlselect.execute(str_sql)
                for row in sqlcursor:
                    if row[0] > 0:
                        #print("chengpin = ", row[1])
                        list_return_from_db.append([curr_wuliao[1],row[1],row[2]])
            return (list_return_from_db)

# 邮寄清单请求处理
        elif count_method == 'maillist':
            str_sql = "select kamiandaima, count(*) from maillist where zhikafangshi = '网络发卡' group by kamiandaima"
            sqlcursor = sqlselect.execute(str_sql)
            for row in sqlcursor:
                if row[1] > 0:
                    # print("chengpin = ", row[1])
                    list_return_from_db.append([row[0], row[1]])
            return (list_return_from_db)
        else:
            print('unknow count_methon.')

#从数据库导入价格（基础表）
    def price_list_from_db(self, customer):
        sqlselect = self.sqlconn.cursor()

        str_sql = "SELECT kamiandaima,kapianbanbenhao,wuliao,mingcheng,gerenhuajiage,kongbaikajiage from price "
        print(str_sql)
        sqlcursor = sqlselect.execute(str_sql)
        curr_price_list = []
        for row in sqlcursor:
            # print("wuliao,name = ", row[0],row[1])
            curr_price_list.append([row[0], row[1], row[2], row[3], row[4], row[5],0]) #最后一个0，为预留用于存节假日发卡数

        #self.sqlconn.close()
        #不通过人工关闭SQL
        return (curr_price_list)

#整合数据，导出生成excel文件
    def db_xls(self, customer, xlsfilename, curr_month):
        if not (customer in self.customers):
            print('不存在该客户格式资料')
            return ('不存在该客户格式资料')

        if customer == 'jtyh':
            int_first_row = 3
            day_column_start = 3  # 日数据开始位置

            list_price = self.price_list_from_db(customer)
            #print (list_price)

            #获取空白卡数量（物料，发出数，按月计算）
            list_kongbaika = self.sum_from_db(customer, curr_month, 'kongbaika')
            list_price_mingcheng = []
            for i in list_price:
                list_price_mingcheng.append(i[3])
            #print (list_price_mingcheng)
            xlsfilename = data_dir + 'jtyhkbkdzd.xlsx'
            workbook  = load_workbook(xlsfilename)  # 打开excel文件

            #worksheel = workbook.get_sheet_by_name('201904')  # 根据Sheet1这个sheet名字来获取该sheet
            worksheel = workbook.worksheets[0]
            for i in range(len(list_kongbaika)):
#                print('test',str(list_kongbaika[i][0]),list_kongbaika[i][1])
                worksheel.cell(int_first_row+i ,1).value =i+1
                list_kongbaika_mingcheng = list_kongbaika[i][0]
                worksheel.cell(int_first_row+i ,3).value =list_kongbaika_mingcheng
                worksheel.cell(int_first_row+i, 5).value = list_kongbaika[i][1]

                list_kongbaika_mingcheng_catch = False
                for j in range(len(list_price_mingcheng)):
                    #print('list_kongbaika_mingcheng= ',list_kongbaika_mingcheng)
                    #print('list_price_mingcheng[j]= ', list_price_mingcheng[j])

                    if list_kongbaika_mingcheng in list_price_mingcheng[j]:
                        worksheel.cell(int_first_row + i, 2).value = list_price[j][1]
                        worksheel.cell(int_first_row + i, 4).value = list_price[j][5]
                        list_kongbaika_mingcheng_catch = True
                        break
                if not list_kongbaika_mingcheng_catch:
                    print("基础表上找不到对应的卡片名称: ",list_kongbaika_mingcheng)
                    logger.info("基础表上找不到对应的卡片名称: " + str(list_kongbaika_mingcheng))
            worksheel.delete_rows(int_first_row+i+1,100-i-1)
            worksheel.cell(int_first_row + i+1,5).value = '=SUM(E3:E'+str(int_first_row+i)+')'
            worksheel.cell(int_first_row + i+1,6).value = '=SUM(F3:F'+str(int_first_row+i)+')'
            workbook.save(xlsfilename[:-8]+'.xlsx')  # 保存修改后的excel
# 空白卡数据文件保存


#个人化数据处理
            int_first_row = 2

            list_gerenhua = self.sum_from_db(customer, curr_month, 'gerenhua')
            list_maillist = self.sum_from_db(customer, curr_month, 'maillist')
#挑出价格表（基础表）中卡面代码与邮寄清单中卡面代码不匹配部分
            for int_maillist in range(len(list_maillist)):

                maillist_price_kamiandaima_match = False
                #print(list_maillist)
                for int_list_price in range(len(list_price)):
                    if list_maillist[int_maillist][0] == list_price[int_list_price][0]:
                        list_price[int_list_price][6] = list_maillist[int_maillist][1]
                        maillist_price_kamiandaima_match = True
                if not maillist_price_kamiandaima_match:
                    logger.info("邮件列表上的卡面代码在基础表上找不到对应的记录，请维护: " + str(list_maillist[int_maillist]))
                    print("邮件列表上的卡面代码在基础表上找不到对应的记录，请维护: " ,list_maillist[int_maillist])
            #print(list_price)
            list_price_mingcheng = []
            for i in list_price:
                list_price_mingcheng.append(i[3])
            #print (list_price_mingcheng)
            xlsfilename = data_dir + 'jtyhgrhdzd.xlsx'
            workbook  = load_workbook(xlsfilename)  # 打开excel文件
            #worksheel = workbook.get_sheet_by_name('201904')  # 根据Sheet1这个sheet名字来获取该sheet
            worksheel = workbook.worksheets[0]
            for i in range(len(list_gerenhua)):
                worksheel.cell(int_first_row+i ,1).value =i+1
                list_gerenhua_mingcheng = list_gerenhua[i][0]
                worksheel.cell(int_first_row+i ,3).value =list_gerenhua_mingcheng
                worksheel.cell(int_first_row+i, 4).value = list_gerenhua[i][1]      #个人化成品数
                worksheel.cell(int_first_row + i, 10).value = list_gerenhua[i][2]    #个人化剪卡数

                list_kongbaika_mingcheng_catch = False
                for j in range(len(list_price_mingcheng)):
                    #print('list_kongbaika_mingcheng= ',list_kongbaika_mingcheng)
                    #print('list_price_mingcheng[j]= ', list_price_mingcheng[j])
                    if list_gerenhua_mingcheng in list_price_mingcheng[j]:
                        worksheel.cell(int_first_row + i, 2).value = list_price[j][1]
                        worksheel.cell(int_first_row + i, 5).value = list_price[j][4]
                        worksheel.cell(int_first_row + i, 7).value = list_price[j][6]
                        worksheel.cell(int_first_row + i, 4).value = list_gerenhua[i][1] - list_price[j][6]  #重复赋值 个人化成品数 - 节假日数
                        worksheel.cell(int_first_row + i, 8).value = list_price[j][4] *2
                        worksheel.cell(int_first_row + i, 11).value = list_price[j][4]
                        list_gerenhua_mingcheng_catch = True
                        break
                if not list_gerenhua_mingcheng_catch:
                    logger.info("基础表上找不到对应的卡片名称: " + str(list_gerenhua_mingcheng))
                    print("基础表上找不到对应的卡片名称: " ,list_gerenhua_mingcheng)
            worksheel.delete_rows(int_first_row+i+1,100-i-1)
            worksheel.cell(int_first_row + i+1,4).value = '=SUM(D2:D'+str(int_first_row+i)+')'
            worksheel.cell(int_first_row + i+1,6).value = '=SUM(F2:F'+str(int_first_row+i)+')'
            worksheel.cell(int_first_row + i+1,7).value = '=SUM(G2:G'+str(int_first_row+i)+')'
            worksheel.cell(int_first_row + i+1,9).value = '=SUM(I2:I'+str(int_first_row+i)+')'
            worksheel.cell(int_first_row + i+1,10).value = '=SUM(J2:J'+str(int_first_row+i)+')'
            worksheel.cell(int_first_row + i + 1, 12).value = '=SUM(L2:L' + str(int_first_row + i) + ')'
            worksheel.cell(int_first_row + i + 1, 13).value = '=SUM(M2:M' + str(int_first_row + i) + ')'
            workbook.save(xlsfilename[:-8]+'.xlsx')  # 保存修改后的excel
# 保存个人化对账单excel文件



if __name__ == '__main__':
    set_logging()
    argvs = argv
    if len(argvs) < 3:
        print ('Sample: c:> d:\python37\python .\balance.py jtyh f:\\dev\\kefu\\ 2019-04')
        #print('Sample: c:> d:\python37\python .\balance.py jtyh f:\\dev\\kefu\\ 2019-04')
        exit(0)
    print('parameter: ',argvs[1],argvs[2],argvs[3])
    exporter = Balance()
    exporter.db_xls(argvs[1],argvs[2],argvs[3])
