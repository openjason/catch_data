import datetime
from configparser import ConfigParser
from tkinter import messagebox, Tk, scrolledtext,Canvas,PhotoImage,Label,StringVar,Entry, Button,END, DISABLED, Toplevel  # 导入滚动文本框的模块
import xlrd
from sqlite3 import connect as sqlite3connect
from calendar import monthrange
from openpyxl import load_workbook
import logging
from logging.handlers import RotatingFileHandler
from os.path import exists as os_path_exists
import os

def set_logging():
    global logger
    logger = logging.getLogger('balance_logger')
    handler = RotatingFileHandler('日志记录.log', maxBytes=5000000, backupCount=6)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)
    formatter = logging.Formatter('%(asctime)-12s %(filename)s %(lineno)d %(message)s')
    handler.setFormatter(formatter)


class App():
    def __init__(self, master):
        self.svar_proc_time1 = StringVar()
        self.svar_cangku_filename = StringVar()
        self.svar_kehumingcheng = StringVar()
        self.svar_proc_time2 = StringVar()
        self.svar_youjiqingdan_filename = StringVar()
        self.svar_jichu_filename = StringVar()
        self.svar_label_prompt = StringVar()

        self.master = master
        self.customer_sname = ''
        self.sqlconn = sqlite3connect("db_dz.db3")
        self.sqlconn.isolation_level = None  # 这个就是事务隔离级别，默认是需要自己commit才能修改数据库，置为None则自动每次修改都提交,否则为""
        self.Holiday = []
        self.data_dir = ''
        self.file_from_cangkujxc = ''
        self.file_from_youjiqingdan = ''
        self.file_from_jichu = ''
        self.curr_month = ''
        self.initWidgets(master)


    def sum_from_db(self, customer, curr_month, count_method):
        sqlselect = self.sqlconn.cursor()
        curr_month_holiday = []
        for i in self.Holiday:
            if curr_month in i:
                curr_month_holiday.append(i)

        str_sql = "SELECT wuliao,name from wuliao where kehu='" + customer + "'"
        # print(str_sql)
        sqlcursor = sqlselect.execute(str_sql)
        curr_wuliao_list = []
        for row in sqlcursor:
            # print("wuliao,name = ", row[0],row[1])
            curr_wuliao_list.append([row[0], row[1]])

        # print(curr_wuliao_list)
        list_return_from_db = []

        # 空白卡请求处理
        if count_method == 'kongbaika':
            for curr_wuliao in curr_wuliao_list:
                str_sql = "SELECT sum(fachu),sum(chengpin),sum(jianka),sum(feika) from days \
where kehu='" + customer + "' and wuliao = '" + str(curr_wuliao[0]) + "'"
                # print(str_sql)
                sqlcursor = sqlselect.execute(str_sql)
                for row in sqlcursor:
                    if row[0] > 0:
                        # print("fachu = ", row[0])
                        list_return_from_db.append([curr_wuliao[1], row[0]])
            return (list_return_from_db)

        # 个人化请求处理
        elif count_method == 'gerenhua':
            for curr_wuliao in curr_wuliao_list:
                str_sql = "SELECT sum(fachu),sum(chengpin),sum(jianka),sum(feika) from days \
where kehu='" + customer + "' and wuliao = '" + str(curr_wuliao[0]) + "'"
                sqlcursor = sqlselect.execute(str_sql)
                for row in sqlcursor:
                    if row[0] > 0:
                        # print("chengpin = ", row[1])
                        list_return_from_db.append([curr_wuliao[1], row[1], row[2]])
            return (list_return_from_db)

        # 邮寄清单请求处理
        elif count_method == 'maillist':
            str_sql = "select kamiandaima, count(*) from maillist where zhikafangshi = '网络发卡' group by kamiandaima"
            sqlcursor = sqlselect.execute(str_sql)
            for row in sqlcursor:
                if row[1] > 0:
                    # print("chengpin = ", row[1])
                    list_return_from_db.append([row[0], row[1]])
            return (list_return_from_db)
        else:
            print('unknow count_methon.')

        # 从数据库导入价格（基础表）
    def price_list_from_db(self, customer):
        sqlselect = self.sqlconn.cursor()

        str_sql = "SELECT kamiandaima,kapianbanbenhao,wuliao,mingcheng,grhpingrijiage1,grhjierijiage1,kongbaikajiage1,hetong2qiyongriqi,grhpingrijiage2,grhjierijiage2,kongbaikajiage2 from price "
        # print(str_sql)
        sqlcursor = sqlselect.execute(str_sql)
        curr_price_list = []
        for row in sqlcursor:
            # print("wuliao,name = ", row[0],row[1])
            curr_price_list.append([row[0], row[1], row[2], row[3], row[4], row[5],row[6],row[7],row[8],row[9],row[10], 0])  # 最后一个0，为预留用于存节假日发卡数

        # self.sqlconn.close()
        # 不通过人工关闭SQL
        return (curr_price_list)

        # 整合数据，导出生成excel文件
    def db_xls(self, customer, xlsfilename, curr_month):

        self.scr.insert(1.0, "准备生成客户对账文件: " + customer+"对账时间: " + str(curr_month) + "\n")
        self.master.update()

        int_first_row = 3
        day_column_start = 3  # 日数据开始位置

        list_price = self.price_list_from_db(customer)
        # print (list_price)

        # 获取空白卡数量（物料，发出数，按月计算）
        list_kongbaika = self.sum_from_db(customer, curr_month, 'kongbaika')
        list_price_mingcheng = []
        for i in list_price:
            list_price_mingcheng.append(i[3])
        # print (list_price_mingcheng)
        xlsfilename = self.data_dir + 'jtyhkbkdzd.xlsx'
        workbook = load_workbook(xlsfilename)  # 打开excel文件

        # worksheel = workbook.get_sheet_by_name('201904')  # 根据Sheet1这个sheet名字来获取该sheet
        worksheel = workbook.worksheets[0]
        for i in range(len(list_kongbaika)):
            #                print('test',str(list_kongbaika[i][0]),list_kongbaika[i][1])
            worksheel.cell(int_first_row + i, 1).value = i + 1
            list_kongbaika_mingcheng = list_kongbaika[i][0]
            worksheel.cell(int_first_row + i, 3).value = list_kongbaika_mingcheng
            worksheel.cell(int_first_row + i, 5).value = list_kongbaika[i][1]

            list_kongbaika_mingcheng_catch = False
            for j in range(len(list_price_mingcheng)):
                # print('list_kongbaika_mingcheng= ',list_kongbaika_mingcheng)
                # print('list_price_mingcheng[j]= ', list_price_mingcheng[j])

                if list_kongbaika_mingcheng in list_price_mingcheng[j]:
                    worksheel.cell(int_first_row + i, 2).value = list_price[j][1]
                    worksheel.cell(int_first_row + i, 4).value = list_price[j][6]
                    list_kongbaika_mingcheng_catch = True
                    break
            if not list_kongbaika_mingcheng_catch:
                print("基础表上找不到对应的卡片名称: ", list_kongbaika_mingcheng)
                logger.info("基础表上找不到对应的卡片名称: " + str(list_kongbaika_mingcheng))
                self.scr.insert(1.0, "*** " + "基础表上找不到对应的卡片名称: " + str(list_kongbaika_mingcheng) + "\n")
                self.master.update()

        # worksheel.cell(1,1).value = curr_month+'空白卡结算（东信和平）'
        worksheel.cell(1, 1).value = curr_month + '空白卡结算（东信和平）'
        worksheel.delete_rows(int_first_row + i + 1, 100 - i - 1)
        worksheel.cell(int_first_row + i + 1, 5).value = '=SUM(E3:E' + str(int_first_row + i) + ')'
        worksheel.cell(int_first_row + i + 1, 6).value = '=SUM(F3:F' + str(int_first_row + i) + ')'
        workbook.save(self.data_dir + '..\\输出对账单\\交通银行空白卡对账'+curr_month+'.xlsx')  # 保存修改后的excel
        # 空白卡数据文件保存
        self.scr.insert(1.0,"空白卡对账明细文件已保存：输出对账单\\交通银行空白卡对账"+curr_month+".xlsx" + "\n")
        self.master.update()

        # 个人化数据处理
        int_first_row = 3

        list_gerenhua = self.sum_from_db(customer, curr_month, 'gerenhua')
        list_maillist = self.sum_from_db(customer, curr_month, 'maillist')
        # 挑出价格表（基础表）中卡面代码与邮寄清单中卡面代码不匹配部分
        for int_maillist in range(len(list_maillist)):

            maillist_price_kamiandaima_match = False
            # print(list_maillist)
            for int_list_price in range(len(list_price)):
                if list_maillist[int_maillist][0] == list_price[int_list_price][0]:
                    list_price[int_list_price][6] = list_maillist[int_maillist][1]
                    maillist_price_kamiandaima_match = True
            if not maillist_price_kamiandaima_match:
                logger.info("邮件列表上的卡面代码在基础表上找不到对应的记录，请维护: " + str(list_maillist[int_maillist]))
                print("邮件列表上的卡面代码在基础表上找不到对应的记录，请维护: ", list_maillist[int_maillist])
                self.scr.insert(1.0, "*** " + "邮件列表上的卡面代码在基础表上找不到对应的记录，请维护: " + str(list_maillist[int_maillist]) + "\n")
                self.master.update()

        # print(list_price)
        list_price_mingcheng = []
        for i in list_price:
            list_price_mingcheng.append(i[3])
        # print (list_price_mingcheng)
        xlsfilename = self.data_dir + 'jtyhgrhdzd.xlsx'
        workbook = load_workbook(xlsfilename)  # 打开excel文件
        # worksheel = workbook.get_sheet_by_name('201904')  # 根据Sheet1这个sheet名字来获取该sheet
        worksheel = workbook.worksheets[0]
        for i in range(len(list_gerenhua)):
            worksheel.cell(int_first_row + i, 1).value = i + 1
            list_gerenhua_mingcheng = list_gerenhua[i][0]
            worksheel.cell(int_first_row + i, 3).value = list_gerenhua_mingcheng
            worksheel.cell(int_first_row + i, 4).value = list_gerenhua[i][1]  # 个人化成品数
            worksheel.cell(int_first_row + i, 10).value = list_gerenhua[i][2]  # 个人化剪卡数

            list_gerenhua_mingcheng_catch = False
            for j in range(len(list_price_mingcheng)):
                # print('list_kongbaika_mingcheng= ',list_kongbaika_mingcheng)
                # print('list_price_mingcheng[j]= ', list_price_mingcheng[j])
                if list_gerenhua_mingcheng in list_price_mingcheng[j]:
                    worksheel.cell(int_first_row + i, 2).value = list_price[j][1]
                    worksheel.cell(int_first_row + i, 5).value = list_price[j][4]
                    worksheel.cell(int_first_row + i, 7).value = list_price[j][11]
                    worksheel.cell(int_first_row + i, 4).value = list_gerenhua[i][1] - list_price[j][11]  # 重复赋值 个人化成品数 - 节假日数
                    worksheel.cell(int_first_row + i, 8).value = list_price[j][5]
                    worksheel.cell(int_first_row + i, 11).value = list_price[j][4]
                    list_gerenhua_mingcheng_catch = True
                    break
            if not list_gerenhua_mingcheng_catch:
                logger.info("基础表上找不到对应的卡片名称: " + str(list_gerenhua_mingcheng))
                print("基础表上找不到对应的卡片名称: ", list_gerenhua_mingcheng)
                self.scr.insert(1.0, "*** " + "基础表上找不到对应的卡片名称: " + str(list_gerenhua_mingcheng) + "\n")
                self.master.update()

        # worksheel.cell(1, 1).value = curr_month + '个人化对账（东信和平）'
        worksheel.cell(1, 1).value = curr_month + '个人化对账（东信和平）'
        worksheel.delete_rows(int_first_row + i + 1, 100 - i - 1)
        worksheel.cell(int_first_row + i + 1, 4).value = '=SUM(D2:D' + str(int_first_row + i) + ')'
        worksheel.cell(int_first_row + i + 1, 6).value = '=SUM(F2:F' + str(int_first_row + i) + ')'
        worksheel.cell(int_first_row + i + 1, 7).value = '=SUM(G2:G' + str(int_first_row + i) + ')'
        worksheel.cell(int_first_row + i + 1, 9).value = '=SUM(I2:I' + str(int_first_row + i) + ')'
        worksheel.cell(int_first_row + i + 1, 10).value = '=SUM(J2:J' + str(int_first_row + i) + ')'
        worksheel.cell(int_first_row + i + 1, 12).value = '=SUM(L2:L' + str(int_first_row + i) + ')'
        worksheel.cell(int_first_row + i + 1, 13).value = '=SUM(M2:M' + str(int_first_row + i) + ')'
        workbook.save(self.data_dir + "..\\输出对账单\\交通银行个人化对账"+curr_month+".xlsx")  # 保存修改后的excel

        self.scr.insert(1.0,"个人化对账明细文件已保存：输出对账单\\交通银行空白卡对账"+curr_month+".xlsx" + "\n")
        self.master.update()
# 保存个人化对账单excel文件

    def xls_db(self, customer, xlsfilename):
        print('xls_db: ',xlsfilename)

        int_first_row = 3
        day_column_start = 7  # 日数据开始位置

        str_sql = "delete from wuliao"
        self.sqlconn.execute(str_sql)
        self.sqlconn.commit()
        str_sql = "delete from days"
        self.sqlconn.execute(str_sql)
        self.sqlconn.commit()
        print("清空原有进销存（wuliao，days）数据...")
        self.scr.insert(1.0, "清空原有进销存（wuliao，days）数据...\n")
        self.master.update()

        workbook = xlrd.open_workbook(xlsfilename)
        sheetsname = workbook.sheet_names()  # 获取excel里的工作表sheet名称数组

        sheetsname.sort(reverse=True)

        str_curr_sheet_name = sheetsname[0]
        list_curr_sheet_name_year_month = str_curr_sheet_name.split('.')
        print('sheetsname: ', list_curr_sheet_name_year_month[0], list_curr_sheet_name_year_month[1])
        monthdaysrange = monthrange(int(list_curr_sheet_name_year_month[0]),
                                    int(list_curr_sheet_name_year_month[1]))
        int_curr_month_days = monthdaysrange[1]
        if int(list_curr_sheet_name_year_month[1]) > 9:
            str_date_y_m = str(list_curr_sheet_name_year_month[0]) + '-' + str(
                list_curr_sheet_name_year_month[1])
        else:
            str_date_y_m = str(list_curr_sheet_name_year_month[0]) + '-0' + str(
                list_curr_sheet_name_year_month[1])
        # Returns weekday of first day of the month and number of days in month, for the specified year and month.

        sheet_curr = workbook.sheet_by_name(str_curr_sheet_name)
        ##            sheet = excel.sheet_by_index(0) #根据下标获取对应的sheet表

        int_sheet_nrows = sheet_curr.nrows
        for i in range(int_first_row, int_sheet_nrows):
            cell_curr_value = sheet_curr.cell(i, 0).value

            if not isinstance(cell_curr_value, str):  # 判断数据是否最后一行
                wuliao = sheet_curr.cell(i, 1).value
                style = sheet_curr.cell(i, 2).value
                name = sheet_curr.cell(i, 3).value
                cell_curr_value = sheet_curr.cell(i, 4).value
                if isinstance(cell_curr_value, str):
                    shangyuejiecun = 0
                else:
                    shangyuejiecun = cell_curr_value

                cell_curr_value = sheet_curr.cell(i, 5).value
                if isinstance(cell_curr_value, str):
                    benyuerucang = 0
                else:
                    benyuerucang = cell_curr_value

                cell_curr_value = sheet_curr.cell(i, 6).value
                if isinstance(cell_curr_value, str):
                    benyuejiecun = 0
                else:
                    benyuejiecun = cell_curr_value
                # 表格后部分的内容
                next_path_data_pos_start = day_column_start + int_curr_month_days * 4
                cell_curr_value = sheet_curr.cell(i, next_path_data_pos_start).value
                if isinstance(cell_curr_value, str):
                    benyuefachushu = 0
                else:
                    benyuefachushu = cell_curr_value
                cell_curr_value = sheet_curr.cell(i, next_path_data_pos_start + 1).value
                if isinstance(cell_curr_value, str):
                    benyuechengpinshu = 0
                else:
                    benyuechengpinshu = cell_curr_value
                cell_curr_value = sheet_curr.cell(i, next_path_data_pos_start + 2).value
                if isinstance(cell_curr_value, str):
                    benyuejiankashu = 0
                else:
                    benyuejiankashu = cell_curr_value
                cell_curr_value = sheet_curr.cell(i, next_path_data_pos_start + 3).value
                if isinstance(cell_curr_value, str):
                    benyuefeikashu = 0
                else:
                    benyuefeikashu = cell_curr_value
                cell_curr_value = sheet_curr.cell(i, next_path_data_pos_start + 4).value
                if isinstance(cell_curr_value, str):
                    benyuefeikaleijishu = 0
                else:
                    benyuefeikaleijishu = cell_curr_value
                cell_curr_value = sheet_curr.cell(i, next_path_data_pos_start + 5).value
                if isinstance(cell_curr_value, str):
                    shangyuejiankaleijishu = 0
                else:
                    shangyuejiankaleijishu = cell_curr_value
                cell_curr_value = sheet_curr.cell(i, next_path_data_pos_start + 6).value
                if isinstance(cell_curr_value, str):
                    shangyuefeikaleijishu = 0
                else:
                    shangyuefeikaleijishu = cell_curr_value
                cell_curr_value = sheet_curr.cell(i, next_path_data_pos_start + 7).value
                if isinstance(cell_curr_value, str):
                    benyuexiaohuikongbaikashu = 0
                else:
                    benyuexiaohuikongbaikashu = cell_curr_value
                cell_curr_value = sheet_curr.cell(i, next_path_data_pos_start + 8).value
                if isinstance(cell_curr_value, str):
                    benyuexiaohuifeikashu = 0
                else:
                    benyuexiaohuifeikashu = cell_curr_value

                # 写入表格前部主要数据
                if len(wuliao) > 3:  # 物料名称字符串长度大于 3 写数据
                    # 插入数据
                    str_sql = 'insert into wuliao(kehu,wuliao,style,name,suoshuyuefen,shangyuejiecun,benyuerucang,benyuejiecun,\
benyuefachushu,benyuechengpinshu,benyuejiankashu,benyuefeikashu,benyuefeikaleijishu,shangyuejiankaleijishu,shangyuefeikaleijishu,benyuexiaohuikongbaikashu,benyuexiaohuifeikashu) \
values("' + customer + '","' + wuliao + '","' + style + '","' + name + '","' + str_date_y_m + '",' + str(
                        shangyuejiecun) + ',' + str(benyuerucang) + ',' + str(benyuejiecun) + ',' + \
                              str(benyuefachushu) + ',' + str(benyuechengpinshu) + ',' + str(
                        benyuejiankashu) + ',' + str(benyuefeikashu) + ',' + str(benyuefeikaleijishu) + \
                              ',' + str(shangyuejiankaleijishu) + ',' + str(shangyuefeikaleijishu) + ',' + str(
                        benyuexiaohuikongbaikashu) + ',' + str(benyuexiaohuifeikashu) + ')'
                    print(name)
                    self.scr.insert(1.0, "数据导入: "+name+'\n')
                    self.master.update()
                    self.sqlconn.execute(str_sql)
                    # 如果隔离级别不是自动提交就需要手动执行commit
                    self.sqlconn.commit()

                    for j in range(0, int_curr_month_days):
                        # 日期个位数加0，下面是对位置进行判断，不是对具体日期进行判断，所以从>8开始判断。
                        if j > 8:
                            str_date = str_date_y_m + '-' + str(j + 1)
                        else:
                            str_date = str_date_y_m + '-0' + str(j + 1)
                        cell_curr_value = sheet_curr.cell(i, day_column_start + 4 * j + 0).value
                        if isinstance(cell_curr_value, str):
                            fachu = 0
                        else:
                            fachu = cell_curr_value
                        cell_curr_value = sheet_curr.cell(i, day_column_start + 4 * j + 1).value
                        if isinstance(cell_curr_value, str):
                            chengpin = 0
                        else:
                            chengpin = cell_curr_value
                        cell_curr_value = sheet_curr.cell(i, day_column_start + 4 * j + 2).value
                        if isinstance(cell_curr_value, str):
                            jianka = 0
                        else:
                            jianka = cell_curr_value
                        cell_curr_value = sheet_curr.cell(i, day_column_start + 4 * j + 3).value
                        if isinstance(cell_curr_value, str):
                            feika = 0
                        else:
                            feika = cell_curr_value

                        # 仓库每日数据写入day表，包含每日发出，成品，剪卡，废卡数量。
                        str_sql = 'insert into days(kehu,wuliao,date,fachu,chengpin,jianka,feika) \
values("' + customer + '","' + wuliao + '","' + str_date + '","' + str(fachu) + '",' + str(chengpin) + ',' + str(
                            jianka) + ',' + str(feika) + ')'
                        # print (str_sql)
                        #print('.', end="")
                        self.sqlconn.execute(str_sql)
                        self.sqlconn.commit()
        print("完成仓库进销存（wuliao，days）数据导入...")
        self.scr.insert(1.0, "完成仓库进销存（wuliao，days）数据导入...\n")
        self.master.update()

    def mailxls_db(self, customer, xlsfilename):

        int_first_row = 1
        day_column_start = 7  # 日数据开始位置

        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message =messagebox.askquestion(title='提示',message='无节假日邮寄信息，继续？') #return yes/no
            return (return_message)

        str_sql = "delete from maillist"
        self.sqlconn.execute(str_sql)
        self.sqlconn.commit()
        print("清空原有邮寄清单（maillist）数据...")
        self.scr.insert(1.0, "清空原有邮寄清单（maillist）数据...\n")
        self.master.update()


        workbook = xlrd.open_workbook(xlsfilename)
        sheetsname = workbook.sheet_names()  # 获取excel里的工作表sheet名称数组

        # sheetsname.sort(reverse=True)

        ##            for str_curr_sheet_name in sheetsname:
        ##

        str_curr_sheet_name = sheetsname[0]
        # list_curr_sheet_name_year_month = str_curr_sheet_name.split('.')
        # print('sheetsname: ',list_curr_sheet_name_year_month[0],list_curr_sheet_name_year_month[1])
        # monthdaysrange = monthrange(int(list_curr_sheet_name_year_month[0]),int(list_curr_sheet_name_year_month[1]))
        # int_curr_month_days = monthdaysrange[1]
        # if int(list_curr_sheet_name_year_month[1]) > 9:
        #    str_date_y_m = str(list_curr_sheet_name_year_month[0]) + '-' + str(list_curr_sheet_name_year_month[1])
        # else:
        #    str_date_y_m = str(list_curr_sheet_name_year_month[0]) + '-0' + str(list_curr_sheet_name_year_month[1])
        # Returns weekday of first day of the month and number of days in month, for the specified year and month.

        sheet_curr = workbook.sheet_by_name(str_curr_sheet_name)
        ##            sheet = excel.sheet_by_index(0) #根据下标获取对应的sheet表
        int_sheet_nrows = sheet_curr.nrows
        print('sheetname & lines:', str_curr_sheet_name, int_sheet_nrows, '行')

        # cell_curr_value = sheet_curr.cell(2, 2).value
        # print(cell_curr_value,int_sheet_nrows)

        for i in range(int_first_row, int_sheet_nrows):
            cell_curr_value = sheet_curr.cell(i, 0).value
            if i % 100 == 0:
                print('procesed: ', i, 'rec.')
                self.scr.insert(1.0, "邮寄清单（maillist）数据导入 "+str(i) +"条.\n")
                self.master.update()

            if True:  # not isinstance(cell_curr_value,str):         #判断数据是否最后一行
                maillistid = sheet_curr.cell(i, 0).value
                picihao = sheet_curr.cell(i, 1).value
                shenqinbianhao = sheet_curr.cell(i, 2).value
                youjifangshi = sheet_curr.cell(i, 3).value
                # print(youjifangshi)
                youjidanghao = sheet_curr.cell(i, 4).value
                jichudi = sheet_curr.cell(i, 5).value
                zhikafangshi = sheet_curr.cell(i, 6).value
                chikarenxingmin = sheet_curr.cell(i, 7).value
                zhukaxingmin = sheet_curr.cell(i, 8).value
                kamiandaima = sheet_curr.cell(i, 9).value
                fakayuanyin = sheet_curr.cell(i, 10).value
                kahao = sheet_curr.cell(i, 11).value
                youjidizhi = sheet_curr.cell(i, 12).value
                youbian = sheet_curr.cell(i, 13).value
                chikarenshouji = sheet_curr.cell(i, 14).value
                youjiriqi = sheet_curr.cell(i, 15).value
                shengchengriqi = sheet_curr.cell(i, 16).value
                zhufukabiaoji = sheet_curr.cell(i, 17).value
                emsliushuihao = sheet_curr.cell(i, 18).value
                pid = sheet_curr.cell(i, 19).value
                quyu = sheet_curr.cell(i, 20).value

                if int(maillistid) > 0:  # testing
                    # 插入数据
                    str_sql = "insert into maillist(id,kehu,picihao,shenqinbianhao,youjifangshi,youjidanghao,jichudi,zhikafangshi,\
chikarenxingmin,kazhuxingmin,kamiandaima,fakayuanyin,kahao,youjidizhi,youbian,chikarenshouji,youjiriqi,shengchengriqi,zhufukabiaoji,\
emsliushuihao,pid,quyu)"
                    str_sql = str_sql + "values(" + str(
                        maillistid) + ",'" + customer + "','" + picihao + "','" + shenqinbianhao + "'," + \
                              youjifangshi + ",'" + youjidanghao + "','" + jichudi + "','" + zhikafangshi + "','" + chikarenxingmin + "','" + zhukaxingmin + "','" + kamiandaima + "','" + fakayuanyin + "','" + \
                              kahao + "','" + youjidizhi + "','" + youbian + "','" + chikarenshouji + "','" + youjiriqi + "','" + shengchengriqi + "','" + zhufukabiaoji + "','" + emsliushuihao + "','" + pid + "','" + quyu + "')"
                    # print (str_sql)
                    self.sqlconn.execute(str_sql)
                    # 如果隔离级别不是自动提交就需要手动执行commit
                    self.sqlconn.commit()
        print('procesed: ', i, 'record.')
        print("完成邮寄清单（maillist）数据导入...")
        self.scr.insert(1.0, "完成邮寄清单（maillist）数据导入.."+str(i)+".\n")
        self.master.update()

    def pricexls_db(self, customer, xlsfilename):

        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message =messagebox.askquestion(title='提示',message='无找到文件'+xlsfilename+'，继续？') #return yes/no
            return (return_message)

        int_first_row = 2
        # day_column_start = 7  # 日数据开始位置

        str_sql = "delete from price"
        self.sqlconn.execute(str_sql)
        self.sqlconn.commit()
        print("清空原有对账价格基础表（price）数据...")
        print(xlsfilename)
        self.scr.insert(1.0, "清空原有对账价格基础表（price）数据...\n")
        self.master.update()

        workbook = xlrd.open_workbook(xlsfilename)
        sheetsname = workbook.sheet_names()  # 获取excel里的工作表sheet名称数组

        # sheetsname.sort(reverse=True)
        ##            for str_curr_sheet_name in sheetsname:
        ##

        str_curr_sheet_name = sheetsname[0]

        sheet_curr = workbook.sheet_by_name(str_curr_sheet_name)
        int_sheet_nrows = sheet_curr.nrows
        print('sheetname & lines:', str_curr_sheet_name, int_sheet_nrows)

        for i in range(int_first_row, int_sheet_nrows):
            cell_curr_value = sheet_curr.cell(i, 0).value
            # print('i: ',i)
            if True:  # not isinstance(cell_curr_value,str):         #判断数据是否最后一行
                priceid = sheet_curr.cell(i, 0).value
                kamiandaima = sheet_curr.cell(i, 1).value
                kapianbanbenhao = sheet_curr.cell(i, 2).value
                wuliao = sheet_curr.cell(i, 3).value
                kapianmingcheng = sheet_curr.cell(i, 4).value

                cell_curr_value = sheet_curr.cell(i, 5).value
                if isinstance(cell_curr_value, str):
                    grhpingrijiage1 = 0
                else:
                    grhpingrijiage1 = cell_curr_value
                cell_curr_value = sheet_curr.cell(i, 6).value
                if isinstance(cell_curr_value, str):
                    grhjierijiage1 = 0
                else:
                    grhjierijiage1 = cell_curr_value
                cell_curr_value = sheet_curr.cell(i, 7).value
                if isinstance(cell_curr_value, str):
                    kongbaikajiage1 = 0
                else:
                    kongbaikajiage1 = cell_curr_value

                hetong2qiyongriqi = sheet_curr.cell(i, 8).value

                cell_curr_value = sheet_curr.cell(i, 9).value
                if isinstance(cell_curr_value, str):
                    grhpingrijiage2 = 0
                else:
                    grhpingrijiage2 = cell_curr_value
                cell_curr_value = sheet_curr.cell(i, 10).value
                if isinstance(cell_curr_value, str):
                    grhjierijiage2 = 0
                else:
                    grhjierijiage2 = cell_curr_value
                cell_curr_value = sheet_curr.cell(i, 11).value
                if isinstance(cell_curr_value, str):
                    kongbaikajiage2 = 0
                else:
                    kongbaikajiage2 = cell_curr_value

                gerenhuafuwu = sheet_curr.cell(i, 12).value
                fuwumingcheng = sheet_curr.cell(i, 13).value
                fuwuleixingbiaoshi = sheet_curr.cell(i, 14).value
                xinpianka = sheet_curr.cell(i, 15).value
                gongjiqueren = sheet_curr.cell(i, 16).value

                if int(priceid) > 0:  # testing
                    # 插入数据
                    str_sql = "insert into price(id,kamiandaima,kapianbanbenhao,wuliao,mingcheng,grhpingrijiage1,grhjierijiage1,\
kongbaikajiage1,hetong2qiyongriqi,grhpingrijiage2,grhjierijiage2,kongbaikajiage2,gerenhuafuwu,fuwumingcheng,\
fuwuleixinbiaoshi,xinpianka,gongyiqueren)"
                    str_sql = str_sql + "values(" + str(priceid) + ",'" + kamiandaima + "','" + kapianbanbenhao + "','" + wuliao + "','" + \
kapianmingcheng + "'," + str(grhpingrijiage1) +"," + str(grhjierijiage1) +","+ str(kongbaikajiage1) +",'" + str(hetong2qiyongriqi) +"'," + \
str(grhpingrijiage2) +"," + str(grhjierijiage2) +","+ str(kongbaikajiage2) + ",'"+ gerenhuafuwu + "','" + fuwumingcheng + "','" + \
fuwuleixingbiaoshi + "','" + xinpianka + "','" + gongjiqueren + "')"
                    print(kapianmingcheng)
                    self.scr.insert(1.0, "基础数据表（price）数据导入: " + str(kapianmingcheng) + ".\n")
                    self.master.update()

                    self.sqlconn.execute(str_sql)
                    # 如果隔离级别不是自动提交就需要手动执行commit
                    self.sqlconn.commit()
        print('=' * 40)
        print('共导入了 ', i - int_first_row + 1, '行数据.')
        self.scr.insert(1.0, "基础数据表（price）数据导入.."+str(i - int_first_row + 1)+"行数据..\n")
        self.master.update()

    def find_filename(self, curr_path, curr_filename_path):
        list_files = []
        for parent, dirnames, filenames in os.walk(curr_path, followlinks=True):
            for filename in filenames:
                file_path = os.path.join(parent, filename)
                if curr_filename_path in filename:
                    print('文件名：%s' % file_path)
                    list_files.append(file_path)
        if len(list_files)>0:
            return (list_files[0])
        else:
            return (None)

    def initWidgets(self,fm1):


        cp = ConfigParser()
        cp.read('配置文件.ini', encoding='gbk')
        str_kehu_name = cp.get('客户', '客户名称')
        self.customer_sname = cp.get('客户', 'sname')
        kehu_conf_jxc = cp.get(str_kehu_name,'仓库进销存')
        self.Holiday = cp.get(str_kehu_name,'节假日')
        self.file_from_cangkujxc = cp.get(str_kehu_name,'仓库进销存')
        self.file_from_youjiqingdan = cp.get(str_kehu_name,'邮寄清单')
        self.file_from_jichu = cp.get(str_kehu_name,'基础数据文件')
        # except Exception as err_message:
        #     print(err_message)

        print('host: ', str_kehu_name)
        print(self.file_from_youjiqingdan)


        self.file_from_cangkujxc = self.find_filename("..\\仓库文件\\", self.file_from_cangkujxc)
        self.file_from_youjiqingdan= self.find_filename("..\\仓库文件\\", self.file_from_youjiqingdan)

        # 设置按钮从顶部开始排列，且按钮只能在垂直（X）方向填充

        label_kehumingcheng = Label(fm1, text='客户名称：', font=('Arial', 12))
        label_kehumingcheng.place(x=20, y=30)
        self.svar_kehumingcheng.set(str_kehu_name)
        entry_kehumingcheng = Entry(fm1, textvariable=self.svar_kehumingcheng, width=30, font=('Arial', 12))
        entry_kehumingcheng.place(x=20, y=55)

        label_proc_time = Label(fm1, text='对账时间：', font=('Arial', 12))
        label_proc_time.place(x=300, y=30)

        temp_last_datetime = datetime.date.today() - datetime.timedelta(days=10)
        self.svar_proc_time1.set(temp_last_datetime.strftime('%Y%m'))

        entry_proc_time1 = Entry(fm1, textvariable=self.svar_proc_time1, width=12, font=('Arial', 12))
        entry_proc_time1.place(x=300, y=55)

        label_proc_time = Label(fm1, text='月', font=('Arial', 12))
        label_proc_time.place(x=420, y=55)

        self.svar_proc_time2.set('20190630')
#        entry_proc_time2 = Entry(fm1, textvariable=self.svar_proc_time2, width=12, font=('Arial', 12))
#        entry_proc_time2.place(x=440, y=55)

        label_cangku_filename = Label(fm1, text='仓库进销存文件名：', font=('Arial', 12))
        label_cangku_filename.place(x=620, y=30)

        self.svar_cangku_filename.set(self.file_from_cangkujxc)
        entry_cangku_filename = Entry(fm1, textvariable=self.svar_cangku_filename, width=40, font=('Arial', 12))
        entry_cangku_filename.place(x=620, y=55)

        label_youjiqingdan_filename = Label(fm1, text='邮寄清单文件名：', font=('Arial', 12))
        label_youjiqingdan_filename.place(x=620, y=80)
        self.svar_youjiqingdan_filename.set(self.file_from_youjiqingdan)
        entry_youjiqingdan_filename = Entry(fm1, textvariable=self.svar_youjiqingdan_filename, width=40, font=('Arial', 12))
        entry_youjiqingdan_filename.place(x=620, y=105)

        label_jichu_filename = Label(fm1, text='价格等基础数据文件名：', font=('Arial', 12))
        label_jichu_filename.place(x=620, y=130)
        self.svar_jichu_filename.set(self.file_from_jichu)
        entry_jichu_filename = Entry(fm1, textvariable=self.svar_jichu_filename, width=40, font=('Arial', 12))
        entry_jichu_filename.place(x=620, y=155)

        self.svar_label_prompt.set('客户名称：')

        label_author = Label(fm1, text='by流程与信息化部ITjc. August,2019', font=('Arial', 9))
        label_author.place(x=820, y=770)

        self.scr = scrolledtext.ScrolledText(fm1, width=80, height=48)
        self.scr.place(x=20, y=100)


#        btn_barcode_check = Button(fm1, text='导入仓库数据', command=self.barcode_check)
#        btn_barcode_check.place(x=650, y=160)
        btn_barcode_init = Button(fm1, text='导入数据&输出对账单', command=self.command_btn_run)
        btn_barcode_init.place(x=620, y=200)

    def command_btn_run(self):


        self.curr_month = self.svar_proc_time1.get()
        self.file_from_cangkujxc = self.svar_cangku_filename.get()
        self.file_from_youjiqingdan = self.svar_youjiqingdan_filename.get()
        self.file_from_jichu = self.svar_jichu_filename.get()


        print('curr_month',self.curr_month)
        print('file_from_cangkujxc', self.file_from_cangkujxc)
        print('customer_sname',self.customer_sname)


        if self.mailxls_db(self.customer_sname,self.file_from_youjiqingdan) == 'no':
            return (1)
        if self.pricexls_db(self.customer_sname, self.file_from_jichu) == 'no':
            return (1)

        self.xls_db(self.customer_sname, self.file_from_cangkujxc)
        # importer.xls_db(argvs[1],argvs[2])

        self.db_xls(self.customer_sname, 'c:\\', self.curr_month)
        return 0



if __name__ == '__main__':

    set_logging()
    main_window = Tk()
    main_window.title('对账单生成工具 v.19083014')

    # 设定窗口的大小(长 * 宽)，显示窗体居中，winfo_xxx获取系统屏幕分辨率。
    sw = main_window.winfo_screenwidth()
    sh = main_window.winfo_screenheight()
    ww = 1024
    wh = 800
    x = (sw - ww) / 2
    y = (sh - wh) / 2
    main_window.geometry("%dx%d+%d+%d" % (ww, wh, x, y))  # 这里的乘是小x
    display = App(main_window)
    main_window.mainloop()
