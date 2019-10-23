'''
功能：处理山西中行借记卡统计表等excel文件，输出山西中行借记卡数据对账单。
'''
from tkinter import *
from configparser import ConfigParser
from tkinter import messagebox,scrolledtext,Canvas,PhotoImage,Label,StringVar,Entry, Button,END, DISABLED, Toplevel  # 导入滚动文本框的模块
from os.path import exists as os_path_exists
import xlrd
import datetime
from sqlite3 import connect as sqlite3connect
from openpyxl import load_workbook
import os
import logging
from logging.handlers import RotatingFileHandler
from openpyxl.styles import Border, Side, Alignment, PatternFill  #设置字体和边框需要的模块
import re

xl_border = Border(left=Side(style='thin',color='FF000000'),right=Side(style='thin',color='FF000000'),top=Side(style='thin',color='FF000000'),bottom=Side(style='thin',color='FF000000'),diagonal=Side(style='thin',color='FF000000'),diagonal_direction=0,outline=Side(style='thin',color='FF000000'),vertical=Side(style='thin',color='FF000000'),horizontal=Side(style='thin',color='FF000000'))

#设置日志文件配置参数
def set_logging():
    global logger
    logger = logging.getLogger('balance_logger')
    handler = RotatingFileHandler('日志记录.log', maxBytes=5000000, backupCount=9)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)
    formatter = logging.Formatter('%(asctime)-12s %(filename)s %(lineno)d %(message)s')
    handler.setFormatter(formatter)

#定义类，脚本主要更能
class App():
    def __init__(self, master):

        self.svar_proc_time1 = StringVar()
        self.svar_proc_time2 = StringVar()
        self.svar_cangku1_filename = StringVar()
        self.svar_cangku2_filename = StringVar()
        self.svar_cangku3_filename = StringVar()
        self.svar_cangku4_filename = StringVar()
        self.text_cangku1_filename = ''
        self.text_cangku2_filename = ''
        self.text_cangku3_filename = ''
        self.text_cangku4_filename = ''
        self.text_cangku1_lable = ''
        self.text_cangku2_lable = ''
        self.text_cangku3_lable = ''
        self.text_cangku4_lable = ''

        self.svar_kehumingcheng = StringVar()

        self.svar_jichu1_filename = StringVar()
        self.svar_jichu2_filename = StringVar()
        self.svar_jichu3_filename = StringVar()
        self.svar_jichu4_filename = StringVar()
        self.text_jichu1_filename = ''
        self.text_jichu2_filename = ''
        self.text_jichu3_filename = ''
        self.text_jichu4_filename = ''
        self.text_jichu1_lable = ''
        self.text_jichu2_lable = ''
        self.text_jichu3_lable = ''
        self.text_jichu4_lable = ''
        self.svar_label_prompt = StringVar()
        self.svar_cpjs_filename = StringVar()
    # 脚本指定数据库名称sqlite3("db_dz.db3")
        self.master = master
        self.customer_sname = ''
        self.sqlconn = sqlite3connect("db_dz.db3")
        self.sqlconn.isolation_level = None  # 这个就是事务隔离级别，默认是需要自己commit才能修改数据库，置为None则自动每次修改都提交,否则为""
        self.Holiday = []
        self.data_dir = ''
        self.file_from_cangkujxc = ''
        self.file_from_fahuoqingdan = ''
        self.file_from_jichu = ''
        self.curr_month = ''
        self.initWidgets(master)

# 按文件夹统计符合条件文件列表，逐个文件导入数据库
    def proc_folder(self, customer, work_dir, filename_p):

        str_sql = "delete from fhqd"
        self.sqlconn.execute(str_sql)
        self.sqlconn.commit()
        print("清空原有汇总统计数据（fhqd）数据...")

        for parent, dirnames, filenames in os.walk(work_dir, followlinks=True):
            for filename in filenames:
                file_path = os.path.join(parent, filename)
                if filename_p in filename:
                    print('文件名：%s' % filename)
                    print('文件完整路径：%s\n' % file_path)
                    self.xls_fhqd_db('sdsb', file_path)

# 从数据库导入价格（基础表），返回含价格信息列表
    def xls_price_db(self, customer, xlsfilename):

        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message = messagebox.askquestion(title='提示',
                                                    message='无找到文件' + xlsfilename + '，继续？')  # return yes/no
            return (return_message)

        int_first_row = 2
        # day_column_start = 7  # 日数据开始位置

        str_sql = "delete from price"
        self.sqlconn.execute(str_sql)
        self.sqlconn.commit()
        print("清空原有对账价格基础表（price）数据...")
        print(xlsfilename)
        self.scr.insert(1.0, "清空原有对账价格基础表（price）数据...\n")
        self.master.update()

        workbook = xlrd.open_workbook(xlsfilename)
        sheetsname = workbook.sheet_names()  # 获取excel里的工作表sheet名称数组

        str_curr_sheet_name = sheetsname[0]

        sheet_curr = workbook.sheet_by_name(str_curr_sheet_name)
        int_sheet_nrows = sheet_curr.nrows
        print('sheetname & lines:', str_curr_sheet_name, int_sheet_nrows)
        hetong2qiyongriqi = ''
        for i in range(int_first_row, int_sheet_nrows):
            cell_curr_value = sheet_curr.cell(i, 0).value
            # print('i: ',i)
            if True:  # not isinstance(cell_curr_value,str):         #判断数据是否最后一行
                priceid = sheet_curr.cell(i, 0).value
                kamiandaima = sheet_curr.cell(i, 1).value
                kapianmingcheng = sheet_curr.cell(i, 2).value
                kapianbanbenhao = sheet_curr.cell(i, 3).value
                wuliao = sheet_curr.cell(i, 7).value

                cell_curr_value = sheet_curr.cell(i, 4).value
                if isinstance(cell_curr_value, str):
                    grhjiage1 = 0
                else:
                    grhjiage1 = cell_curr_value

                if hetong2qiyongriqi == '':
                    hetong2qiyongriqi = sheet_curr.cell(i, 5).value

                cell_curr_value = sheet_curr.cell(i, 6).value
                if isinstance(cell_curr_value, str):
                    grhjiage2 = 0
                else:
                    grhjiage2 = cell_curr_value

                if int(priceid) > 0:  # testing
                    # 插入数据
                    str_sql = "insert into price(id,kamiandaima,kapianbanbenhao,wuliao,kapianmingcheng,grhjiage1,grhjiage2,\
hetong2qiyongriqi)"
                    str_sql = str_sql + "values(" + str(priceid) + ",'" + kamiandaima + "','" + kapianbanbenhao + "','" + \
wuliao + "','" + kapianmingcheng + "'," + str(grhjiage1) + "," + str(grhjiage2) + ",'" + str(hetong2qiyongriqi) + "')"
                    #print(kapianmingcheng)
                    self.scr.insert(1.0, "数据导入 基础数据表（price）: " + str(kapianmingcheng) + ".\n")
                    self.master.update()

                    self.sqlconn.execute(str_sql)
                    # 如果隔离级别不是自动提交就需要手动执行commit
                    self.sqlconn.commit()
        print('=' * 40)
        print('共导入了 ', i - int_first_row + 1, '行数据.')
        self.scr.insert(1.0, "基础数据表（price）数据导入.." + str(i - int_first_row + 1) + "行数据..\n")
        self.master.update()

# excel数据明细表数据导入到数据库
    def xls_shujumingxibiao_db(self, customer, xlsfilename):

        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message = messagebox.askquestion(title='提示',
                                                    message='无找到文件' + xlsfilename + '，继续？')  # return yes/no
            return (return_message)

        int_first_row = 1

        str_sql = "delete from sjmxb"
        self.sqlconn.execute(str_sql)
        self.sqlconn.commit()
        print("清空原有数据明细表（sjmxb）数据...")
        print(xlsfilename)
        self.scr.insert(1.0, "清空原有数据明细表（sjmxb）数据...\n")
        self.master.update()

        workbook = xlrd.open_workbook(xlsfilename)
        sheetsname = workbook.sheet_names()  # 获取excel里的工作表sheet名称数组
    #读最左边的一个sheet
        str_curr_sheet_name = sheetsname[0]
        sheet_curr = workbook.sheet_by_name(str_curr_sheet_name)
        int_sheet_nrows = sheet_curr.nrows
        print('sheetname & lines:', str_curr_sheet_name, int_sheet_nrows)
        for i in range(int_first_row, int_sheet_nrows):
            shujuwenjianming = sheet_curr.cell(i, 0).value
            yzfyz = sheet_curr.cell(i, 1).value
            jigouhao = sheet_curr.cell(i, 2).value
            wangdianmingcheng = sheet_curr.cell(i, 3).value
            shuliang = sheet_curr.cell(i, 4).value
            chanpinma = sheet_curr.cell(i, 5).value
            chanpinmingcheng = sheet_curr.cell(i, 6).value
            qishihaoma = sheet_curr.cell(i, 7).value
            zhongzhihaoma = sheet_curr.cell(i, 8).value
            shujugeshi = sheet_curr.cell(i, 9).value
            beizhu2 = sheet_curr.cell(i, 10).value
            beizhu3 = sheet_curr.cell(i, 11).value
            beizhu4 = sheet_curr.cell(i, 12).value
            beizhu5 = sheet_curr.cell(i, 13).value
            beizhu6 = sheet_curr.cell(i, 14).value

        # 插入数据
            str_sql = "insert into sjmxb(kehu,sjwjm,yzfyz,jigouhao,wangdianmingcheng,shuliang,chanpinma,chenpinmingcheng,\
qishihaoma,zhongzhihaoma,shujugeshi,beizhu2,beizhu3,beizhu4,beizhu5,beizhu6)"
            str_sql = str_sql + " values('" + customer + "','" + shujuwenjianming + "','" + yzfyz + "','" + \
jigouhao + "','" + wangdianmingcheng + "'," + str(shuliang) + ",'" + chanpinma + "','" + chanpinmingcheng + "','" + qishihaoma+ "','" + \
zhongzhihaoma+ "','" + shujugeshi+ "','" + beizhu2+ "','" + beizhu3+ "','" + beizhu4+ "','"+beizhu5+ "','"+beizhu6+ "')"
            #print(chanpinmingcheng)
            if i%200 ==0 :
                self.scr.insert(1.0, "导入数据明细表，处理了 "+str(i)+"条.\n")
            self.master.update()
            self.sqlconn.execute(str_sql)
            # 如果隔离级别不是自动提交就需要手动执行commit
            self.sqlconn.commit()
        print('=' * 40)
        print('共导入了 ', i - int_first_row + 1, '行数据.')
        self.scr.insert(1.0, "基础数据表 数据明细（sjmxb）数据导入.." + str(i - int_first_row + 1) + "行数据..\n")
        self.master.update()

    #按字符查找符合条件文件名，返回文件列表
    def find_filename(self, curr_path, curr_filename_path):
        list_files = []
        curr_filename_path = curr_filename_path.replace('*','')
        for parent, dirnames, filenames in os.walk(curr_path, followlinks=True):
            for filename in filenames:
                file_path = os.path.join(parent, filename)
                if curr_filename_path in filename and filename[0]!='~':
                    print('文件名：%s' % file_path)
                    list_files.append(file_path)
        if len(list_files) > 0:
            return (list_files[0])
        else:
            return (None)

    # 从数据库导出价格（基础表），返回含价格信息列表
    def price_list_from_db(self, customer):
        sqlselect = self.sqlconn.cursor()

        str_sql = "SELECT rowid,id,kamiandaima,kapianbanbenhao,wuliao,kapianmingcheng,grhjiage1,grhjiage2,hetong2qiyongriqi from price "
        # print(str_sql)
        sqlcursor = sqlselect.execute(str_sql)
        curr_price_list = []
        for row in sqlcursor:
            # print("wuliao,name = ", row[0],row[1])
            curr_price_list.append([row[0], row[1], row[2], row[3], row[4], row[5],row[6],row[7],row[8], 0])  # 最后一个0，为预留用于存节假日发卡数
        # self.sqlconn.close()
        # 不通过人工关闭SQL
        return (curr_price_list)

#从数据库处理数据，导出对账文件excel
    def db_xls(self, customer, xlsfilename):

        str_temp = str(self.svar_proc_time1.get())
        time1_proc = str_temp[:4] + '-' + str_temp[4:6] + '-' + str_temp[6:8]
        str_temp = str(self.svar_proc_time2.get())
        time2_proc = str_temp[:4] + '-' + str_temp[4:6] + '-' + str_temp[6:8]


        self.scr.insert(1.0, "准备生成客户对账文件: " + customer + "对账周期: " + str(time1_proc) +' - ' +str(time2_proc)+ "\n")
        self.master.update()

        int_first_row = 3
        list_price = self.price_list_from_db(customer)
        list_price_onlycode  = []       # 现此变量暂时无用，只用于记录日志。
        for temp in list_price:
            list_price_onlycode.append(temp[0])
        logger.info ('代码列表： '+str(list_price_onlycode))

        # 获取明细表数据(制卡单位)
        zhikadanwei_list = []
        sqlselect = self.sqlconn.cursor()
        str_sql = "select xuhao,nsrsbh,dzjdh,khhzh,kpyq,zkdwmc,kpdwmc,lxr,lxfs,yjdz,yzbm from kpxxhd "
        sqlcursor = sqlselect.execute(str_sql)
        for row in sqlcursor:
            zhikadanwei_list.append(row)

#        xlsfilename = self.data_dir + 'sdsbxykdz.xlsx'
        workbook = load_workbook(xlsfilename)  # 打开excel文件
# 导出明细表begin
        logger.info('导出 ~明细表~ 表' )
        self.scr.insert(1.0, "导出 ~明细表~ 表"+ "\n")
        #worksheetj = workbook['明细表']  # 根据Sheet1这个sheet名字来获取该sheet
        worksheetj = workbook.worksheets[0]
        i = 0
        #worksheetj.cell(1,1).value = '甘肃农信信用卡对账明细表('+str(time1_proc)+'至'+str(time2_proc)+')'
        for row in zhikadanwei_list:
            row_chanpinbianma = row[0]
            #print(row)
            rowprice = 0
            for temp_bianma in list_price:
                if row_chanpinbianma == temp_bianma[0]:
                    rowprice = temp_bianma[4]
                    #break

            worksheetj.cell(int_first_row + i, 1).value = row[0]    #序号
            worksheetj.cell(int_first_row + i, 1).border = xl_border
            worksheetj.cell(int_first_row + i, 2).value = row[1]  # 纳税人识别码
            worksheetj.cell(int_first_row + i, 2).border = xl_border
            worksheetj.cell(int_first_row + i, 3).value = row[2]  # 地址及电话
            worksheetj.cell(int_first_row + i, 3).border = xl_border
            worksheetj.cell(int_first_row + i, 4).value = row[3]  # 开户行及账号
            worksheetj.cell(int_first_row + i, 4).border = xl_border
            worksheetj.cell(int_first_row + i, 5).value = row[4]  # 开票要求
            worksheetj.cell(int_first_row + i, 5).border = xl_border
            worksheetj.cell(int_first_row + i, 6).value = row[5]  # 制卡单位名称
            worksheetj.cell(int_first_row + i, 6).border = xl_border
            worksheetj.cell(int_first_row + i, 7).value = row[6]  #开票单位名称
            worksheetj.cell(int_first_row + i, 7).border = xl_border
            worksheetj.cell(int_first_row + i, 14).value = row[7]  # 联系人
            worksheetj.cell(int_first_row + i, 14).border = xl_border
            worksheetj.cell(int_first_row + i, 15).value = row[8] # 联系方式
            worksheetj.cell(int_first_row + i, 15).border = xl_border
            worksheetj.cell(int_first_row + i, 16).value = row[9] # 邮寄地址
            worksheetj.cell(int_first_row + i, 16).border = xl_border
            worksheetj.cell(int_first_row + i, 17).value = row[10] # 邮编
            worksheetj.cell(int_first_row + i, 17).border = xl_border


            i = i + 1 #下一行
        worksheetj.cell(int_first_row + i , 6).value = '=SUM(F3:F' + str(int_first_row + i-1) + ')'
        worksheetj.cell(int_first_row + i , 9).value = '=SUM(I3:I' + str(int_first_row + i-1) + ')'

        workbook.save('..\\输出文件\\山西中行卡体开票及邮寄信息核对表.xlsx')  # 保存修改后的excel
        self.scr.insert(1.0, "空白卡对账明细文件已保存：\\输出文件\\甘肃农信制卡清单对账明细表" + str(time2_proc) + ".xlsx" + "\n")
        self.master.update()

        return 0
# 导出明细表end

# 导出分联社分摊数量与金额begin
        str_sql = "select danweimingcheng,sum(shuliang),sum(shuliang*grhpingrijiage1) from jiagemingxi"
        str_sql = str_sql + " where youjiriqi>'" + time1_proc + "' and youjiriqi<'" + time2_proc + "' group by danweimingcheng"
        # print(str_sql)
        sqlcursor = sqlselect.execute(str_sql)
        worksheetj = workbook['分联社分摊数量与金额']  # 根据Sheet1这个sheet名字来获取该sheet
        logger.info('导出 ~分联社分摊数量与金额~ 表' )
        self.scr.insert(1.0, "导出 ~分联社分摊数量与金额~ 表"+ "\n")
        i = 0
        worksheetj.cell(1, 1).value = '各分联社分摊数量与金额清单(' + str(time1_proc) + '至' + str(time2_proc) + ')'
        for row in sqlcursor:
            row_chanpinbianma = row[0]
            rowprice = 0
            worksheetj.cell(int_first_row + i, 1).value = i + 1  #序号
            worksheetj.cell(int_first_row + i, 1).border = xl_border
            worksheetj.cell(int_first_row + i, 2).value = row[0]  # 分联社名称
            worksheetj.cell(int_first_row + i, 2).border = xl_border
            worksheetj.cell(int_first_row + i, 3).value = row[1]  # 数量
            worksheetj.cell(int_first_row + i, 3).border = xl_border
            worksheetj.cell(int_first_row + i, 4).value = row[2]  # 金额
            worksheetj.cell(int_first_row + i, 4).border = xl_border
            i = i + 1  # 下一行
        worksheetj.cell(int_first_row + i, 3).value = '=SUM(C3:C' + str(int_first_row + i - 1) + ')'
        worksheetj.cell(int_first_row + i, 4).value = '=SUM(D3:D' + str(int_first_row + i - 1) + ')'
# 导出分联社分摊数量与金额end

# 汇总表begin
        worksheetj = workbook['汇总表']  # 根据Sheet1这个sheet名字来获取该sheet
        logger.info('导出 ~汇总~ 表')
        self.scr.insert(1.0, "导出 ~汇总~ 表"+ "\n")
        str_sql = "select kamiandaima,mingcheng from price"
        sqlcursor = sqlselect.execute(str_sql)
        list_kamiandaima_from_price = []
        for row in sqlcursor:
            list_kamiandaima_from_price.append([row[0], row[1]])
        str_sql = "select chanpinbianma from jiagemingxi group by chanpinbianma"
        sqlcursor = sqlselect.execute(str_sql)
        list_kamiandaima_from_hztj = []
        for row in sqlcursor:
            list_kamiandaima_from_hztj.append(row[0])

        print('before:', list_kamiandaima_from_price)

        for list_kamiandaima in list_kamiandaima_from_price:
            if not (list_kamiandaima[0] in list_kamiandaima_from_hztj):
                print('list_kamiandaima', list_kamiandaima)
                list_kamiandaima_from_price.remove(list_kamiandaima)
        print('after:', list_kamiandaima_from_price)
        list_for_kamian_position = []
        for list_kamiandaima in list_kamiandaima_from_price:
            list_for_kamian_position.append(list_kamiandaima[0])
        print('list_for_kamian_position:', list_for_kamian_position)

    #汇总表表头
        worksheetj.cell(1, 1).value = '制卡清单(' + str(time1_proc) + '至' + str(time2_proc) + ')'
        row_kamian = 2  # 表头所在行.
        col_kamian_fix = 3     #表格 卡面 开始列

        col_kamian = col_kamian_fix
        for i in list_kamiandaima_from_price:
            worksheetj.cell(row_kamian, col_kamian).value = i[1] + "(" + i[0]+")"
            worksheetj.cell(row_kamian, col_kamian).border = xl_border
            col_kamian = col_kamian +1
        worksheetj.cell(row_kamian, col_kamian).value = "总数"
        worksheetj.cell(row_kamian, col_kamian).border = xl_border

        str_sql = "select youjiriqi,chanpinbianma,sum(shuliang) from jiagemingxi group by youjiriqi,chanpinbianma"
        # print(str_sql)
        sqlcursor = sqlselect.execute(str_sql)
        col_kamian = col_kamian_fix     #表格 卡面 开始列
        row_kamian = 2  # 数据首行.
        last_row_youjiriqi = ''
        for row in sqlcursor:
            code_kamian = row[1]
            shuliang_kamian = row[2]
            curr_row_youjiriqi = row[0]

            for j in range(len(list_for_kamian_position) +3):
                worksheetj.cell(row_kamian, j+1).border = xl_border

            if last_row_youjiriqi == curr_row_youjiriqi:
                int_index_kamian = list_for_kamian_position.index(code_kamian)
                worksheetj.cell(row_kamian, col_kamian+int_index_kamian).value = shuliang_kamian
            else:
                last_row_youjiriqi = curr_row_youjiriqi
    #数据下一行
                row_kamian = row_kamian +1  # 数据下一行.
                worksheetj.cell(row_kamian, 1).value = row_kamian -2
                worksheetj.cell(row_kamian, 2).value = curr_row_youjiriqi

                int_index_kamian = list_for_kamian_position.index(code_kamian)
                worksheetj.cell(row_kamian, col_kamian+int_index_kamian).value = shuliang_kamian
                worksheetj.cell(row_kamian, col_kamian+len(list_for_kamian_position)).value ='=SUM(C'+str(row_kamian)+':' + chr(66+len(list_for_kamian_position))+str(row_kamian)+')'

        row_kamian = row_kamian +1
        col_kamian = col_kamian_fix
        worksheetj.cell(row_kamian, 2).value = "总数"
        worksheetj.cell(row_kamian, 1).border = xl_border
        worksheetj.cell(row_kamian, 2).border = xl_border
        for i in range(len(list_kamiandaima_from_price)+1):
            str_cell = "=SUM("+chr(64+col_kamian)+"3:"+chr(64+col_kamian) + str(row_kamian -1) + ")"
            worksheetj.cell(row_kamian, col_kamian).value = str_cell
            worksheetj.cell(row_kamian, col_kamian).border = xl_border
            col_kamian = col_kamian +1
# 汇总表end
# 汇总表 计价 begin
        print(('list_price[0][7](hetong):' , list_price[0][7]))
        logger.info('list_price[0][7](hetong):' + str(list_price[0][7]))

        # str_temp = str(self.svar_proc_time1.get())
        # time1_proc = str_temp[:4] + '-' + str_temp[4:6] + '-' + str_temp[6:8]
        # str_temp = str(self.svar_proc_time2.get())
        # time2_proc = str_temp[:4] + '-' + str_temp[4:6] + '-' + str_temp[6:8]

        folat_zongji = 0 # 报表总计

        #判断是否存在合同2（对账周期内存在2份合同，价格不同）
        hetong2riqi = list_price[0][7]
        logger.info('合同二日期：' + str(hetong2riqi) +';time1: '+ time1_proc +';time2:'+ time2_proc)
        self.scr.insert(1.0, '合同二日期：' + str(hetong2riqi) +';time1: '+ time1_proc +';time2:'+ time2_proc+ '\n')
        print('合同二日期：' , str(hetong2riqi), ';time1: '+ time1_proc +';time2:'+ time2_proc)
        if hetong2riqi > time1_proc and hetong2riqi < time2_proc:
            print('按合同2分期统计：')
            self.scr.insert(1.0, '按合同2分期统计：' + str(hetong2riqi) +';time1: '+ time1_proc +';time2:'+ time2_proc+ '\n')
    #合同1统计
            str_sql = "select chanpinbianma,sum(shuliang) from hztj  where youjiriqi>'" + time1_proc + "' and youjiriqi<'"+hetong2riqi +"' group by chanpinbianma"
            sqlcursor = sqlselect.execute(str_sql)
            print('str_sql',str_sql)
            row_kamian = row_kamian +2
            col_kamian = col_kamian_fix
            for row in sqlcursor:
                int_index_kamian = list_for_kamian_position.index(row[0])
                worksheetj.cell(row_kamian, col_kamian + int_index_kamian).value = row[1]    #数量
                for price_from_list in list_price:
                    curr_cell_price = 0
                    if row[0] == price_from_list[0]:
                        curr_cell_price = price_from_list[4]
                        break
                if curr_cell_price ==0 :
                    logger.info('汇总按合同1查找价格，找不到对应的价格：' + str(row[0]))
                    self.scr.insert(1.0,'汇总按合同1查找价格，找不到对应的价格' + str(row[0]) + '\n')
                worksheetj.cell(row_kamian+1, col_kamian + int_index_kamian).value = curr_cell_price     #单价
                worksheetj.cell(row_kamian+2, col_kamian + int_index_kamian).value = curr_cell_price*row[1]     #小计
                folat_zongji = folat_zongji + curr_cell_price*row[1]
            worksheetj.merge_cells(start_row=row_kamian, start_column=2, end_row=row_kamian+2, end_column=2)
            worksheetj.cell(row_kamian , 2).value =  time1_proc + ' 至 ' + hetong2riqi

            cell_row_cell_to_name = self.excel_cell_rowcell_to_position(2,row_kamian)
            print('cell_row_cell_to_name',cell_row_cell_to_name)
            worksheetj[cell_row_cell_to_name].alignment = Alignment(wrapText=True)

            worksheetj.merge_cells(start_row=row_kamian, start_column=1, end_row=row_kamian+2, end_column=1)
            worksheetj.cell(row_kamian , 1).value =  '合计'

            for i in range(3):
                for j in range(len(list_for_kamian_position)+3):
                    worksheetj.cell(row_kamian + i, j+1).border = xl_border


    #合同2统计
            str_sql = "select chanpinbianma,sum(shuliang) from hztj  where youjiriqi>'" + hetong2riqi + "' and youjiriqi<'"+time2_proc +"' group by chanpinbianma"
            sqlcursor = sqlselect.execute(str_sql)
            print('str_sql',str_sql)
            row_kamian = row_kamian +4
            col_kamian = col_kamian_fix
            for row in sqlcursor:
                int_index_kamian = list_for_kamian_position.index(row[0])
                worksheetj.cell(row_kamian, col_kamian + int_index_kamian).value = row[1]    #数量
                for price_from_list in list_price:
                    curr_cell_price = 0
                    if row[0] == price_from_list[0]:
                        curr_cell_price = price_from_list[4]
                        break
                if curr_cell_price ==0 :
                    logger.info('汇总按合同1查找价格，找不到对应的价格：' + str(row[0]))
                    self.scr.insert(1.0,'汇总按合同1查找价格，找不到对应的价格' + str(row[0]) + '\n')
                worksheetj.cell(row_kamian+1, col_kamian + int_index_kamian).value = curr_cell_price     #单价
                worksheetj.cell(row_kamian+2, col_kamian + int_index_kamian).value = curr_cell_price*row[1]     #小计
                folat_zongji = folat_zongji + curr_cell_price * row[1]

            worksheetj.merge_cells(start_row=row_kamian, start_column=2, end_row=row_kamian+2, end_column=2)

            worksheetj.cell(row_kamian , 2).value = hetong2riqi + ' 至 ' + time2_proc
            cell_row_cell_to_name = self.excel_cell_rowcell_to_position(2,row_kamian )
            print('cell_row_cell_to_name',cell_row_cell_to_name)
            worksheetj[cell_row_cell_to_name].alignment = Alignment(wrapText=True)

            worksheetj.merge_cells(start_row=row_kamian, start_column=1, end_row=row_kamian+2, end_column=1)
            worksheetj.cell(row_kamian , 1).value =  '合计'

            worksheetj.merge_cells(start_row=row_kamian+3, start_column=2, end_row=row_kamian+3, end_column=len(list_for_kamian_position)+3)
            worksheetj.cell(row_kamian +3, 1).value = '总计'
            worksheetj.cell(row_kamian +3, 1).border = xl_border
            worksheetj.cell(row_kamian + 3, 2).value = folat_zongji
            #worksheetj.cell(row_kamian + 3, 2).border = xl_border
            for i in range(4):
                for j in range(len(list_for_kamian_position)+3):
                    worksheetj.cell(row_kamian + i, j+1).border = xl_border

    #按一份合同统计
        else:
            folat_zongji = 0
            str_sql = "select chanpinbianma,sum(shuliang) from hztj  where youjiriqi>'" + time1_proc + "' and youjiriqi<'"+time2_proc +"' group by chanpinbianma"
            sqlcursor = sqlselect.execute(str_sql)
            print('str_sql',str_sql)
            row_kamian = row_kamian +2
            col_kamian = col_kamian_fix
            for row in sqlcursor:
                int_index_kamian = list_for_kamian_position.index(row[0])
                worksheetj.cell(row_kamian, col_kamian + int_index_kamian).value = row[1]    #数量
                for price_from_list in list_price:
                    curr_cell_price = 0
                    if row[0] == price_from_list[0]:
                        curr_cell_price = price_from_list[4]
                        break
                if curr_cell_price ==0 :
                    logger.info('汇总按合同1查找价格，找不到对应的价格：' + str(row[0]))
                    self.scr.insert(1.0,'汇总按合同1查找价格，找不到对应的价格' + str(row[0]) + '\n')
                worksheetj.cell(row_kamian+1, col_kamian + int_index_kamian).value = curr_cell_price     #单价
                worksheetj.cell(row_kamian+2, col_kamian + int_index_kamian).value = curr_cell_price*row[1]     #小计
                folat_zongji = folat_zongji + curr_cell_price * row[1]

            worksheetj.merge_cells(start_row=row_kamian, start_column=2, end_row=row_kamian+2, end_column=2)

            worksheetj.cell(row_kamian , 2).value = time1_proc + ' 至 ' + time2_proc
            cell_row_cell_to_name = self.excel_cell_rowcell_to_position(2,row_kamian )
            print('cell_row_cell_to_name',cell_row_cell_to_name)
            worksheetj[cell_row_cell_to_name].alignment = Alignment(wrapText=True)

            worksheetj.merge_cells(start_row=row_kamian, start_column=1, end_row=row_kamian+2, end_column=1)
            worksheetj.cell(row_kamian , 1).value =  '合计'

            worksheetj.merge_cells(start_row=row_kamian+3, start_column=2, end_row=row_kamian+3, end_column=len(list_for_kamian_position)+3)
            worksheetj.cell(row_kamian +3, 1).value = '总计'
            worksheetj.cell(row_kamian +3, 1).border = xl_border
            worksheetj.cell(row_kamian + 3, 2).value = folat_zongji
            #worksheetj.cell(row_kamian + 3, 2).border = xl_border
            for i in range(4):
                for j in range(len(list_for_kamian_position)+3):
                    worksheetj.cell(row_kamian + i, j+1).border = xl_border

# 汇总表 计价 end

# 格式begin
        aligmentCenter = Alignment(horizontal='center', vertical ='center',wrapText=True)
        #commonBackgroundColorHex = "AACF91"
        #commonFill = PatternFill(start_color=commonBackgroundColorHex, end_color=commonBackgroundColorHex,fill_type="solid")
        for eachCommonRow in worksheetj.iter_rows(min_row=1, max_col=len(list_for_kamian_position)+3, max_row=row_kamian + 3):
            for eachCellInRow in eachCommonRow:
                eachCellInRow.alignment = aligmentCenter
                #eachCellInRow.fill = commonFill
# 格式end

        workbook.save(self.data_dir + '..\\输出文件\\甘肃农信制卡清单对账明细表' + str(time2_proc) + '.xlsx')  # 保存修改后的excel
        self.scr.insert(1.0, "空白卡对账明细文件已保存：\\输出文件\\甘肃农信制卡清单对账明细表" + str(time2_proc) + ".xlsx" + "\n")
        self.master.update()
# 保存excel文件

    def import_heduibiao(self,custom,xlsfilename):
        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message = messagebox.askquestion(title='提示',
                                                    message='无找到文件' + xlsfilename + '，继续？')  # return yes/no
            return (return_message)

        int_first_row = 2

        str_sql = "delete from kpxxhd"
        self.sqlconn.execute(str_sql)
        self.sqlconn.commit()
        print("清空原有数据表（kpxxhd）数据...")
        print(xlsfilename)
        self.scr.insert(1.0, "清空原有数据表（kpxxhd）数据...\n")
        self.master.update()

        workbook = xlrd.open_workbook(xlsfilename)
        sheetsname = workbook.sheet_names()  # 获取excel里的工作表sheet名称数组
    #读最左边的一个sheet
        str_curr_sheet_name = sheetsname[0]
        sheet_curr = workbook.sheet_by_name(str_curr_sheet_name)
        int_sheet_nrows = sheet_curr.nrows
        print('sheetname & lines:', str_curr_sheet_name, int_sheet_nrows)
        for i in range(int_first_row, int_sheet_nrows):
            cell_value = sheet_curr.cell(i, 0).value
            if type(cell_value) == float:
                xuhao =  int(cell_value)

                cell_value = sheet_curr.cell(i, 1).value
                if type(cell_value) == float:
                    nsrsbh = int(cell_value)
                else:
                    nsrsbh = sheet_curr.cell(i, 1).value

                dzjdh = sheet_curr.cell(i, 2).value
                khhzh = sheet_curr.cell(i, 3).value
                kpyq = sheet_curr.cell(i, 4).value
                zkdwmc = sheet_curr.cell(i, 5).value
                kpdwmc = sheet_curr.cell(i, 6).value
                lxr = sheet_curr.cell(i, 13).value #skit 6 cell

                cell_value = sheet_curr.cell(i, 14).value
                if type(cell_value) == float:
                    lxfs = int(cell_value)
                else:
                    lxfs = sheet_curr.cell(i, 14).value

                yjdz = sheet_curr.cell(i, 15).value
                cell_value = sheet_curr.cell(i, 16).value
                if type(cell_value) == float:
                    yzbm = int(cell_value)
                else:
                    yzbm = sheet_curr.cell(i, 16).value
                print('yzbm',yzbm,yjdz)
            # 插入数据
                str_sql = "insert into kpxxhd(xuhao,nsrsbh,dzjdh,khhzh,kpyq,zkdwmc,kpdwmc,lxr,\
lxfs,yjdz,yzbm)"
                str_sql = str_sql + " values(" + str(xuhao) + ",'" + nsrsbh + "','" + dzjdh + "','" + \
khhzh + "','" + kpyq + "','" + str(zkdwmc) + "','" + \
kpdwmc + "','" + lxr + "','" + str(lxfs)+ "','" + str(yjdz)+ "','" + \
str(yzbm) + "')"
                print(str_sql)
                if i%200 ==0 :
                    self.scr.insert(1.0, "导入数据明细表，处理了 "+str(i)+"条.\n")
                self.master.update()
                self.sqlconn.execute(str_sql)
                # 如果隔离级别不是自动提交就需要手动执行commit
                self.sqlconn.commit()
        print('=' * 40)
        print('共导入了 ', i - int_first_row + 1, '行数据.')
        self.scr.insert(1.0, "基础数据表 数据明细（sjmxb）数据导入.." + str(i - int_first_row + 1) + "行数据..\n")
        self.master.update()


    def excel_cell_rowcell_to_position(self,int_row,int_column):
        if int_row < 26:
            str_excel_cell_pos = chr(64+int_row)
            str_excel_cell_pos = str_excel_cell_pos + str(int_column)
        return str_excel_cell_pos

# 导入基础文件1，成品接收excel文件（机构、接收人信息）
    def xls_yzfyz_db(self, customer, xlsfilename):
        logger.info('xls_yzfyz_db: ' + xlsfilename)
        int_first_row = 2

        workbook = xlrd.open_workbook(xlsfilename)
        sheetsname = workbook.sheet_names()  # 获取excel里的工作表sheet名称数组

        str_curr_sheet_name = '预制卡'
        sheet_curr = workbook.sheet_by_name(str_curr_sheet_name)
        ##            sheet = excel.sheet_by_index(0) #根据下标获取对应的sheet表

        int_sheet_nrows = sheet_curr.nrows

        print('lines: ', int_first_row, int_sheet_nrows)
        print('CELL（A2）: ', sheet_curr.cell(1, 0).value)
        if sheet_curr.cell(1, 0).value != '上级机构名称':
            print('单元格位置有误，A2应为文字“上级机构名称” ', sheet_curr.cell(1, 0).value)
            logger.info('单元格位置有误，A2应为文字“上级机构名称” '+ str(sheet_curr.cell(1, 0).value))
            self.scr.insert(1.0, "** 文件格式或内容有误: " + xlsfilename + '\n')
            return 2
        else:
            self.scr.insert(1.0, "数据导入 xls_db: " + xlsfilename + '\n')
            self.master.update()
        #导入sheet（预计卡）begin
            yzfyz = 'Y'
            for i in range(int_first_row, int_sheet_nrows):
                sjjgmc = sheet_curr.cell(i, 0).value
                sjjgmc = sjjgmc.strip()
                jgmc = sheet_curr.cell(i, 1).value
                jgmc = jgmc.strip()

                ctype = sheet_curr.cell(i, 2).ctype
                cell_value = sheet_curr.cell(i, 2).value
                if ctype == 2 and cell_value % 1 == 0.0:
                    jgdm = str(int(cell_value))
                else:
                    jgdm = cell_value.strip()

                jsr = sheet_curr.cell(i, 3).value
                jsr = jsr.strip()
                jsrdh = str(sheet_curr.cell(i, 4).value)
                jsrdh =jsrdh.strip()

                ctype = sheet_curr.cell(i, 5).ctype
                cell_value = sheet_curr.cell(i, 5).value
                if ctype == 2 and cell_value % 1 == 0.0:
                    jsrsj = str(int(cell_value))
                else:
                    jsrsj = cell_value.strip()

                jsrdz = sheet_curr.cell(i, 6).value
                jsrdz = jsrdz.strip()

                ctype = sheet_curr.cell(i, 7).ctype
                cell_value = sheet_curr.cell(i, 7).value
                if ctype == 2 and cell_value % 1 == 0.0:
                    jsryb = str(int(cell_value))
                else:
                    jsryb = cell_value.strip()

                logger.info('机构名称:' + str(jgmc))
                string_have_old_jgmc = self.search_old_yzfyz_data(yzfyz,str(jgmc))
                logger.info(string_have_old_jgmc)
                if not string_have_old_jgmc:
                    str_sql = "insert into yzfyz(kehu,yzfyz,sjjgmc,jgmc,jgdm,jsr,jsrdh,jsrsj,jsrdz,jsryb) \
                    values('" + customer + "','" + yzfyz + "','" + sjjgmc + "','" + jgmc + "','" + jgdm + "','" + jsr + "','" + jsrdh + "','" + jsrsj + \
                              "','" + jsrdz + "','" + jsryb + "')"
                    # print('数据导入 xls_db: ',kehumingcheng)
                    logger.info(str_sql)
                    self.sqlconn.execute(str_sql)
                    # 如果隔离级别不是自动提交就需要手动执行commit
                    self.sqlconn.commit()
                    self.scr.insert(1.0, "数据导入 机构名称: " + str(jgmc) + "\n")
                    self.master.update()
                else:
                    str_compare_curr = yzfyz + sjjgmc + jgmc + jgdm + jsr + jsrdh + jsrsj + jsrdz + jsryb
                    logger.info(str_compare_curr)
                    if string_have_old_jgmc == str_compare_curr:
                        self.scr.insert(1.0, "数据导入 机构名称: " + str(jgmc) + "  已存在\n")
                    else:
            # 更新数据
                        str_sql = "update yzfyz set jsr2= jsr where yzfyz = '" + yzfyz + "' and jgmc ='" + jgmc + "'"
                        logger.info(str_sql)
                        self.sqlconn.execute(str_sql)
                        self.sqlconn.commit()

                        str_sql = "update yzfyz set kehu='" + customer + "',sjjgmc='"+sjjgmc+ \
"',jgdm='" + jgdm+"',jsr='"+jsr+"',jsrdh='"+jsrdh+"',jsrsj='"+jsrsj+"',jsrdz='"+jsrdz+"',jsryb='" + jsryb + \
"' where yzfyz='" + yzfyz + "' and jgmc='" + jgmc + "'"
                        #print('数据导入 xls_db: ',kehumingcheng)
                        logger.info(str_sql)
                        self.sqlconn.execute(str_sql)
                        # 如果隔离级别不是自动提交就需要手动执行commit
                        self.sqlconn.commit()
                        self.scr.insert(1.0, "数据导入 机构名称: " + str(jgmc) + "\n")
                        self.master.update()
        # 导入sheet（预计卡）end

        # 导入sheet（非预计卡）begin
            str_curr_sheet_name = '非预制卡'
            sheet_curr = workbook.sheet_by_name(str_curr_sheet_name)
            ##            sheet = excel.sheet_by_index(0) #根据下标获取对应的sheet表
            int_sheet_nrows = sheet_curr.nrows

            last_sjjgmc = ''
            last_jsr = ''
            last_jsrdh = ''
            last_jsrsj =''
            last_jsrdz = ''
            last_jsryb = ''

            yzfyz = 'F'
            for i in range(int_first_row, int_sheet_nrows):
                sjjgmc = sheet_curr.cell(i, 0).value
                sjjgmc = sjjgmc.strip()
                jgmc = sheet_curr.cell(i, 1).value
                jgmc = jgmc.strip()

                ctype = sheet_curr.cell(i, 2).ctype
                cell_value = sheet_curr.cell(i, 2).value
                if ctype == 2 and cell_value % 1 == 0.0:
                    jgdm = str(int(cell_value))
                else:
                    jgdm = cell_value.strip()

                jsr = sheet_curr.cell(i, 3).value
                jsr = jsr.strip()
                jsrdh = str(sheet_curr.cell(i, 4).value)
                jsrdh = jsrdh.strip()

                ctype = sheet_curr.cell(i, 5).ctype
                cell_value = sheet_curr.cell(i, 5).value
                if ctype == 2 and cell_value % 1 == 0.0:
                    jsrsj = str(int(cell_value))
                else:
                    jsrsj = cell_value.strip()

                jsrdz = sheet_curr.cell(i, 6).value
                jsrdz = jsrdz.strip()

                ctype = sheet_curr.cell(i, 7).ctype
                cell_value = sheet_curr.cell(i, 7).value
                if ctype == 2 and cell_value % 1 == 0.0:
                    jsryb = str(int(cell_value))
                else:
                    jsryb = cell_value.strip()
                if sjjgmc == '':
                    sjjgmc = last_sjjgmc
                    jsr = last_jsr
                    jsrdh = last_jsrdh
                    jsrsj = last_jsrsj
                    jsrdz = last_jsrdz
                    jsryb = last_jsryb
                else:
                    last_sjjgmc = sjjgmc
                    last_jsr = jsr
                    last_jsrdh = jsrdh
                    last_jsrsj = jsrsj
                    last_jsrdz = jsrdz
                    last_jsryb = jsryb

                logger.info('机构名称:' + str(jgmc))
                string_have_old_jgmc = self.search_old_yzfyz_data(yzfyz,str(jgmc))
                logger.info(string_have_old_jgmc)
                if not string_have_old_jgmc:
                # 是否插入数据
                    str_sql = "insert into yzfyz(kehu,yzfyz,sjjgmc,jgmc,jgdm,jsr,jsrdh,jsrsj,jsrdz,jsryb) \
                    values('" + customer + "','" + yzfyz + "','" + sjjgmc + "','" + jgmc + "','" + jgdm + "','" + jsr + "','" + jsrdh + "','" + jsrsj + \
                              "','" + jsrdz + "','" + jsryb + "')"
                    # print('数据导入 xls_db: ',kehumingcheng)
                    logger.info(str_sql)
                    self.sqlconn.execute(str_sql)
                    # 如果隔离级别不是自动提交就需要手动执行commit
                    self.sqlconn.commit()
                    self.scr.insert(1.0, "数据导入 机构名称: " + str(jgmc) + "\n")
                    self.master.update()
                else:
                    str_compare_curr = yzfyz + sjjgmc + jgmc + jgdm + jsr + jsrdh + jsrsj + jsrdz + jsryb
                    logger.info(str_compare_curr)
                    if string_have_old_jgmc == str_compare_curr:
                        self.scr.insert(1.0, "数据导入 机构名称: " + str(jgmc) + "  已存在\n")
                    else:
                # 更新数据
                        str_sql = "update yzfyz set jsr2= jsr where yzfyz = '" + yzfyz + "' and jgmc ='" + jgmc + "'"
                        logger.info(str_sql)
                        self.sqlconn.execute(str_sql)
                        self.sqlconn.commit()

                        str_sql = "update yzfyz set kehu='" + customer + "',sjjgmc='" + sjjgmc + \
"',jgdm='" + jgdm + "',jsr='" + jsr + "',jsrdh='" + jsrdh + "',jsrsj='" + jsrsj + "',jsrdz='" + jsrdz + "',jsryb='" + jsryb + \
"' where yzfyz ='" + yzfyz + "' and jgmc='" + jgmc + "'"
                        logger.info(str_sql)
                        self.sqlconn.execute(str_sql)
                        self.sqlconn.commit()
                        self.scr.insert(1.0, "更新 数据导入 机构名称: " + str(jgmc) + "\n")
                        self.master.update()
        self.scr.insert(1.0, "完成：  数据导入 机构信息/接收人清单 \n")
        # 导入sheet（预计卡）end

# 导入预制/非预制数据时搜索是否已存在相同机构名称的情况
    def search_old_yzfyz_data(self,yzfyz,jigoumingcheng):
        sqlselect = self.sqlconn.cursor()
        str_sql = "select yzfyz,sjjgmc,jgmc,jgdm,jsr,jsrdh,jsrsj,jsrdz,jsryb from yzfyz where yzfyz = '" + yzfyz + "' and jgmc ='" + jigoumingcheng + "'"
        sqlcursor = sqlselect.execute(str_sql)
        logger.info(str_sql)
        string_return = ""
        if sqlcursor:
            logger.info('search_old_yzfyz_data 查询结果数：' )
            for row in sqlcursor:
                logger.info(row)
                string_return = ''
                for i in range(len(row)):
                    string_return = string_return + str(row[i])
                break
            return  string_return
        return ""

    # 程序主gui界面。
    def initWidgets(self, fm1):

        cp = ConfigParser()
        cp.read('配置文件.ini', encoding='gbk')
        str_kehu_name = cp.get('客户', '客户名称')
        try:
            self.customer_sname = cp.get('客户', 'sname')

            self.text_jichu1_lable = cp.get(str_kehu_name, '基础文件标题1')
            self.text_jichu2_lable = cp.get(str_kehu_name, '基础文件标题2')
            self.text_jichu3_lable = cp.get(str_kehu_name, '基础文件标题3')
            self.text_jichu4_lable = cp.get(str_kehu_name, '基础文件标题4')

            self.text_jichu1_filename = cp.get(str_kehu_name, '基础文件名1')
            self.text_jichu2_filename = cp.get(str_kehu_name, '基础文件名2')
            self.text_jichu3_filename = cp.get(str_kehu_name, '基础文件名3')
            self.text_jichu4_filename = cp.get(str_kehu_name, '基础文件名4')

            self.text_cangku1_lable = cp.get(str_kehu_name, '仓库或数据标题1')
            self.text_cangku2_lable = cp.get(str_kehu_name, '仓库或数据标题2')
            self.text_cangku3_lable = cp.get(str_kehu_name, '仓库或数据标题3')
            self.text_cangku4_lable = cp.get(str_kehu_name, '仓库或数据标题4')

            self.text_cangku1_filename = cp.get(str_kehu_name, '仓库或数据文件名1')
            self.text_cangku2_filename = cp.get(str_kehu_name, '仓库或数据文件名2')
            self.text_cangku3_filename = cp.get(str_kehu_name, '仓库或数据文件名3')
            self.text_cangku4_filename = cp.get(str_kehu_name, '仓库或数据文件名4')

        except Exception as err_message:
            print(err_message)
            logger.error(err_message.__str__())
            logger.exception(sys.exc_info())
            return_message = messagebox.showinfo(title='提示',message='无法打开配置文件.ini或配置有误!' )
            exit(2)

        print('host: ', str_kehu_name)

        label_kehumingcheng = Label(fm1, text='客户名称：', font=('Arial', 12))
        label_kehumingcheng.place(x=20, y=30)
        self.svar_kehumingcheng.set(str_kehu_name)
        entry_kehumingcheng = Entry(fm1, textvariable=self.svar_kehumingcheng, width=30, font=('Arial', 12))
        entry_kehumingcheng.place(x=20, y=55)

        label_proc_time = Label(fm1, text='对账时间：', font=('Arial', 12))
        label_proc_time.place(x=300, y=30)

        temp_last_datetime = datetime.date.today() - datetime.timedelta(days=10)

        self.svar_proc_time1.set('20180201')
        entry_proc_time1 = Entry(fm1, textvariable=self.svar_proc_time1, width=12, font=('Arial', 12))
        entry_proc_time1.place(x=300, y=55)

        label_proc_time = Label(fm1, text='-', font=('Arial', 12))
        label_proc_time.place(x=420, y=55)


        self.svar_proc_time2.set('20190630')
        entry_proc_time2 = Entry(fm1, textvariable=self.svar_proc_time2, width=12, font=('Arial', 12))
        entry_proc_time2.place(x=440, y=55)
        svar_label_prompt = StringVar()
        svar_label_prompt.set('客户名称：')
        btn_barcode_init = Button(fm1, text=' 退  出 ', command=self.command_btn_exit)
        btn_barcode_init.place(x=620, y=500)

        btn_xxb_init = Button(fm1, text='开票及邮寄信息表导入', command=self.command_btn_import_hedui)
        btn_xxb_init.place(x=620, y=450)

        btn_tongji_xxb_xls_creat = Button(fm1, text='开票及邮寄信息表统计输出', command=self.command_btn_xxb_output)
        btn_tongji_xxb_xls_creat.place(x=620, y=600)

        label_author = Label(fm1, text='by流程与信息化部IT. oct,2019', font=('Arial', 9))
        label_author.place(x=820, y=770)

        label_file_position = 70
        label_file_position_space = 50
    #两标题/文件间隔离高度

        label_jichu1_filename = Label(fm1, text=self.text_jichu1_lable, font=('Arial', 12))
        label_jichu1_filename.place(x=620, y=label_file_position)
        logger.info('call find_filename :' + self.text_jichu1_filename)
        text_jichu1_filename_result = self.find_filename('..\\价格等基础数据',self.text_jichu1_filename)
        self.svar_jichu1_filename.set(text_jichu1_filename_result)
        entry_jichu1_filename = Entry(fm1, textvariable=self.svar_jichu1_filename, width=40, font=('Arial', 12))
        entry_jichu1_filename.place(x=620, y=label_file_position+25)
        label_file_position = label_file_position + label_file_position_space
    #基础文件2
        if self.text_jichu2_lable !='':
            label_jichu2_filename = Label(fm1, text=self.text_jichu2_lable, font=('Arial', 12))
            label_jichu2_filename.place(x=620, y=label_file_position)
            logger.info('call find_filename :' + self.text_jichu2_filename)
            text_jichu2_filename_result = self.find_filename('..\\价格等基础数据', self.text_jichu2_filename)
            self.svar_jichu2_filename.set(text_jichu2_filename_result)
            entry_jichu2_filename = Entry(fm1, textvariable=self.svar_jichu2_filename, width=40, font=('Arial', 12))
            entry_jichu2_filename.place(x=620, y=label_file_position+25)
            label_file_position = label_file_position + label_file_position_space
    #基础文件3
        if self.text_jichu3_lable !='':
            label_jichu3_filename = Label(fm1, text=self.text_jichu3_lable, font=('Arial', 12))
            label_jichu3_filename.place(x=620, y=label_file_position)
            logger.info('call find_filename :' + self.text_jichu3_filename)
            text_jichu3_filename_result = self.find_filename('..\\价格等基础数据', self.text_jichu3_filename)
            self.svar_jichu3_filename.set(text_jichu3_filename_result)
            entry_jichu3_filename = Entry(fm1, textvariable=self.svar_jichu3_filename, width=40, font=('Arial', 12))
            entry_jichu3_filename.place(x=620, y=label_file_position+25)
            label_file_position = label_file_position + label_file_position_space
    # 基础文件4
        if self.text_jichu4_lable !='':
            label_jichu4_filename = Label(fm1, text=self.text_jichu4_lable, font=('Arial', 12))
            label_jichu4_filename.place(x=620, y=label_file_position)
            logger.info('call find_filename :' + self.text_jichu4_filename)
            text_jichu4_filename_result = self.find_filename('..\\价格等基础数据', self.text_jichu4_filename)
            self.svar_jichu4_filename.set(text_jichu4_filename_result)
            entry_jichu4_filename = Entry(fm1, textvariable=self.svar_jichu4_filename, width=40, font=('Arial', 12))
            entry_jichu4_filename.place(x=620, y=label_file_position+25)
            label_file_position = label_file_position + label_file_position_space

    # 仓库文件1
        if self.text_cangku1_lable != '':
            label_cangku1_filename = Label(fm1, text=self.text_cangku1_lable, font=('Arial', 12))
            label_cangku1_filename.place(x=620, y=label_file_position)
            logger.info('call find_filename :' + self.text_cangku1_filename)
            text_cangku1_filename_result = self.find_filename('..\\仓库文件', self.text_cangku1_filename)
            self.svar_cangku1_filename.set(text_cangku1_filename_result)
            entry_cangku1_filename = Entry(fm1, textvariable=self.svar_cangku1_filename, width=40, font=('Arial', 12))
            entry_cangku1_filename.place(x=620, y=label_file_position + 25)
            label_file_position = label_file_position + label_file_position_space

        self.scr = scrolledtext.ScrolledText(fm1, width=80, height=48)
        self.scr.place(x=20, y=100)
        #
        btn_banlance_output = Button(fm1, text='导入数据&输出对账单', command=self.command_btn_run)
        btn_banlance_output.place(x=620, y=700)


    # 退出键
    def command_btn_exit(self):
        self.master.destroy()

    def command_btn_import_hedui(self):
        try:
            self.import_heduibiao(self.customer_sname, '..\\客户核对后文件\\山西中行卡体开票及邮寄信息核对1903-1907.xlsx')
        except Exception as err_message:
            print(err_message)
            logger.error(err_message.__str__())
            logger.exception(sys.exc_info())
    def command_btn_xxb_output(self):
        try:
            self.db_xls(self.customer_sname, '..\\程序\\山西中行卡体开票及邮寄信息核对表.xlsx')
        except Exception as err_message:
            print(err_message)
            logger.error(err_message.__str__())
            logger.exception(sys.exc_info())

    # 主功能键
    def command_btn_run(self):

        self.scr.delete(1.0,END)

        label_tips1_filename = Label(self.master, text='正在导入数据... ', font=('Arial', 12))
        label_tips1_filename.place(x=620, y=430)

        self.file_from_jiechuwenjian1 = self.svar_cpjs_filename.get()
        self.text_jichu1_filename = self.svar_jichu1_filename.get()
        if not self.text_jichu2_lable == '':
            self.text_jichu2_filename = self.svar_jichu2_filename.get()
        if not self.text_jichu3_lable == '':
            self.text_jichu3_filename = self.svar_jichu3_filename.get()
        if not self.text_jichu4_lable == '':
            self.text_jichu4_filename = self.svar_jichu4_filename.get()

        if not self.text_cangku1_lable == '':
            self.text_cangku1_filename = self.svar_cangku1_filename.get()
        if not self.text_cangku2_lable == '':
            self.text_cangku2_filename = self.svar_cangku2_filename.get()
        if not self.text_cangku3_lable == '':
            self.text_cangku3_filename = self.svar_cangku3_filename.get()
        if not self.text_cangku4_lable == '':
            self.text_cangku4_filename = self.svar_cangku4_filename.get()

        print('curr_month', self.curr_month)
        print('text_jichu1_filename', self.text_jichu1_filename)
        print('customer_sname', self.customer_sname)

        try:


#            self.xls_shujumingxibiao_db(self.customer_sname, self.text_cangku1_filename)
            self.xls_price_db(self.customer_sname, self.text_jichu2_filename)
            self.xls_yzfyz_db(self.customer_sname, self.text_jichu1_filename)
        except Exception as err_message:
            print(err_message)
            logger.error(err_message.__str__())
            logger.exception(sys.exc_info())

        label_tips1_filename = Label(self.master, text='完成...                     ', font=('Arial', 12))
        label_tips1_filename.place(x=620, y=430)

        return 0


if __name__ == '__main__':

    set_logging()

    main_window = Tk()
    main_window.title('对账单生成工具 v.19092410')

    # 设定窗口的大小(长 * 宽)，显示窗体居中，winfo_xxx获取系统屏幕分辨率。
    sw = main_window.winfo_screenwidth()
    sh = main_window.winfo_screenheight()
    ww = 1024
    wh = 800
    x = (sw - ww) / 2
    y = (sh - wh) / 2
    main_window.geometry("%dx%d+%d+%d" % (ww, wh, x, y))  # 这里的乘是小x
    display = App(main_window)
    main_window.mainloop()
