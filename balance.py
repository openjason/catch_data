from openpyxl import load_workbook
from sys import argv
from sqlite3 import connect as sqlite3connect


class Balance:
    def __init__(self):
        self.customers = ['jtyh', 'gsnx']
        self.sqlconn = sqlite3connect("F:\\dev\\kefu\\jinxiaocun.db3")
        self.sqlconn.isolation_level = None  # 这个就是事务隔离级别，默认是需要自己commit才能修改数据库，置为None则自动每次修改都提交,否则为""
        self.Holiday = ['2019-02-04', '2019-02-05', '2019-02-06', '2019-02-07', '2019-02-08', '2019-02-09', '2019-02-10', '2019-04-05',\
'2019-07-07', '2019-05-01', '2019-05-02', '2019-05-03', '2019-05-04', '2019-06-07', '2019-06-08', '2019-06-09',\
'2019-09-13', '2019-09-14','2019-09-15', '2019-10-01', '2019-10-02', '2019-10-03', '2019-10-04', '2019-10-05', '2019-10-06', '2019-10-07']

    def sum_from_db(self, customer, curr_month, count_method):
        sqlselect = self.sqlconn.cursor()
        curr_month_holiday = []
        for i in self.Holiday:
            if curr_month in i:
                curr_month_holiday.append(i)

        str_sql = "SELECT wuliao,name from wuliao where kehu='" + customer + "'"
        print(str_sql)
        sqlcursor = sqlselect.execute(str_sql)
        curr_wuliao_list = []
        for row in sqlcursor:
            #print("wuliao,name = ", row[0],row[1])
            curr_wuliao_list.append([row[0],row[1]])

        print(curr_wuliao_list)
        list_return_from_db = []
        if count_method == 'kongbaika':

            for curr_wuliao in curr_wuliao_list:
                str_sql = "SELECT sum(fachu),sum(chengpin),sum(jianka),sum(feika) from days \
where kehu='" + customer +"' and wuliao = '" + str(curr_wuliao[0]) + "'"

                print(str_sql)
                sqlcursor = sqlselect.execute(str_sql)
                for row in sqlcursor:
                    if row[0] > 0:
                        print("fachu = ", row[0])
                        list_return_from_db.append([curr_wuliao[1],row[0]])
            return (list_return_from_db)

        elif count_method == 'gerenhua':
            if len(curr_month_holiday) > 0:
                print("Holiday: ", curr_month_holiday)
                str_sql = "SELECT sum(fachu),sum(chengpin),sum(jianka),sum(feika) from days "
                str_sql_holiday = ''
                for i in curr_month_holiday:
                    if str_sql_holiday == '':
                        str_sql_holiday = "where date='" + i + "'"
                    else:
                        str_sql_holiday = str_sql_holiday + " or date='" + i + "'"
                str_sql_holiday = str_sql + str_sql_holiday
                print(str_sql_holiday)
                sqlcursor = sqlselect.execute(str_sql_holiday)
                for row in sqlcursor:
                    print("chengpin = ", row[1])
                # 编辑排除节假日sql
                str_sql_holiday = ''
                for i in curr_month_holiday:
                    str_sql_holiday = str_sql_holiday + " and date <> '" + i + "'"
            else:
                print('this month has not holiday.')
                str_sql_holiday = ''
            # 计算节假日以外的成品数量
            str_sql = "SELECT sum(fachu),sum(chengpin),sum(jianka),sum(feika) from days \
            where date>='" + curr_month + "-01' and date <= '" + curr_month + "-31'" + str_sql_holiday
            print(str_sql)
            sqlcursor = sqlselect.execute(str_sql)

            for row in sqlcursor:
                print("chengpin = ", row[1])
        else:
            print('unknow count_methon.')

    def db_xls(self, customer, xlsfilename, curr_month):
        if not (customer in self.customers):
            print('不存在该客户格式资料')
            return ('不存在该客户格式资料')

        if customer == 'jtyh':
            int_first_row = 3
            day_column_start = 3  # 日数据开始位置


            #获取空白卡数量（物料，发出数，按月计算）
            list_kongbaika = self.sum_from_db(customer, curr_month, 'kongbaika')
            print("list_kongbaika:",list_kongbaika)

#            print("chengpin = ", sqlcursor[0][1])

            workbook  = load_workbook(xlsfilename)  # 打开excel文件

            #worksheel = workbook.get_sheet_by_name('201904')  # 根据Sheet1这个sheet名字来获取该sheet
            worksheel = workbook.worksheets[0]
            # 添加一栏为地区，并且给上数据
            for i in range(len(list_kongbaika)):
#                print('test',str(list_kongbaika[i][0]),list_kongbaika[i][1])
                worksheel.cell(int_first_row+i ,1).value =i+1
                worksheel.cell(int_first_row+i ,day_column_start).value =list_kongbaika[i][0]
                worksheel.cell(int_first_row+i, day_column_start+2).value = list_kongbaika[i][1]

            workbook.save(xlsfilename[:-8]+'.xlsx')  # 保存修改后的excel
            # 保存

            self.sqlconn.close()

if __name__ == '__main__':
    argvs = argv
    #d:\python37\python .\balance.py jtyh f:\\dev\\kefu\\jtyhdzd.xlsx 2019-04
    if len(argvs) < 3:
        print ('Sample: c:> d:\python37\python .\balance.py jtyh f:\\dev\\kefu\\jtyhdzd.xlsx 2019-04')
        exit(0)
    print('parameter: ',argvs[1],argvs[2],argvs[3])
    exporter = Balance()
    exporter.db_xls(argvs[1],argvs[2],argvs[3])
