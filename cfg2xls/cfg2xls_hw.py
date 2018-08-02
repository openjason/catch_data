# -*- coding: UTF-8 -*-
'''
根据华为防火墙配置文件配置内容，提取分类内容，保存到excel文件不同的sheet相应栏目中
author:jason chan
2018-08-01
'''
# import os
import openpyxl
import copy
import re

def judge_legal_ip(one_str):
    '''''
    正则匹配方法
    判断一个字符串是否是合法IP地址
    '''
    compile_ip = re.compile('^((25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(25[0-5]|2[0-4]\d|[01]?\d\d?)$')
    if compile_ip.match(one_str):
        return True
    else:
        return False

def fix_address_string(add_str):
    #停用    return add_str
    fix_addr_str = "地址有误"
    tmp_list = add_str.split()
    aname = tmp_list[0]
    if len(tmp_list)< 3:
        return add_str
    if 'range' == tmp_list[2]:
        fix_addr_str = tmp_list[3] + '-' + tmp_list[4]

    if 'mask' == tmp_list[3]:
        fix_addr_str = tmp_list[2] + '/' + tmp_list[4]
    return fix_addr_str


def save_block_file(blocked_list, block_child_list, fn):
    #配置文件一块为单位，进行保存
    with open(fn, 'w') as fopen:
        for i in range(len(blocked_list)):
            if 'ipv6' in blocked_list[i]:
                continue
            fopen.write('\n===========\n')
            fopen.writelines(blocked_list[i])
#            if block_list[i] != 'null':
            fopen.writelines(block_child_list[i])

def getServiceList(blocked_list, block_child_list,sname):
    #获取具体服务对应的服务名和端口号
    standard_service = ['ftp','ssh','ntp','http','icmp']
    return_service_list = []
    if sname in standard_service:
        return_service_list.append(sname)
        return return_service_list
    target_string = 'ip service-set ' + sname + ' '
    target_str_len = len(target_string)

    for i in range(len(blocked_list)):
        tempStr_raw = blocked_list[i]
        tempStr1 = tempStr_raw
        if target_string == tempStr1[:target_str_len]:
            tempStr2 = tempStr1[-11:len(tempStr1)]
            if 'type object' == tempStr1[-11:]:
                i = i + 1
                tempStr1 = blocked_list[i]
                while tempStr1[0] == ' ':
                    if ' service' == tempStr1[:8]:
                        tempList3 = tempStr1.split()
                        if tempList3[3] == 'icmp':
                            return_service_list.append('ICMP')
                        elif tempList3[3] == 'tcp':
                            if len(tempList3) == 10:
                                return_service_list.append('tcp' + tempList3[9])
                            elif len(tempList3) == 12:
                                return_service_list.append('tcp' + tempList3[9]+'-' + tempList3[11])
                        elif tempList3[3] == 'udp':
                            if len(tempList3) == 10:
                                return_service_list.append('udp' + tempList3[9])
                            elif len(tempList3) == 12:
                                return_service_list.append('udp' + tempList3[9]+'-' + tempList3[11])

                    i = i + 1
                    tempStr1 = blocked_list[i]
                return return_service_list
            if 'type group' == tempStr1[-10:]:
                i = i + 1
                tempStr1 = blocked_list[i]
                while tempStr1[0] == ' ':
                    if ' service' == tempStr1[:8]:
                        tempList = tempStr1.split()
                        sub_group_string = tempList[3]
                        sub_group_r_l = getServiceList(blocked_list, block_child_list, sub_group_string)
                        for j in range(len(sub_group_r_l)):
                            return_service_list.append(sub_group_r_l[j])
                    i = i + 1
                    tempStr1 = blocked_list[i]
                return return_service_list
    return []


def getAddressList(blocked_list, block_child_list,aname):
    #获取具体地址对应的地址名和地址
    address_detail = ''
    return_address_list = []
    target_string = 'ip address-set ' + aname + ' '
    target_str_len = len(target_string)
    for i in range(len(blocked_list)):
        tempStr_raw = blocked_list[i]
        tempStr1 = tempStr_raw
        if target_string == tempStr1[:target_str_len]:
            tempStr2 = tempStr1[-11:len(tempStr1)]
            if 'type object' == tempStr1[-11:]:
                i = i + 1
                tempStr1 = blocked_list[i]
                while tempStr1[0] == ' ':
                    if ' address' == tempStr1[:8]:
                        tempList = tempStr1.split()
                        try:
                            return_address_list.append(fix_address_string(tempStr1))
                        except:
                            print("err_return_address_list.append(fix_address_string(tempStr1))" + tempStr1)
                    i = i + 1
                    tempStr1 = blocked_list[i]
                return return_address_list
            if 'type group' == tempStr1[-10:]:
                i = i + 1
                tempStr1 = blocked_list[i]
                while tempStr1[0] == ' ':
                    if ' address' == tempStr1[:8]:
                        tempList = tempStr1.split()
                        sub_group_string = tempList[3]
                        sub_group_r_l = getAddressList(blocked_list, block_child_list, sub_group_string)
                        for j in range(len(sub_group_r_l)):
                            return_address_list.append(sub_group_r_l[j])
                    i = i + 1
                    tempStr1 = blocked_list[i]
                return return_address_list


                return address_detail
    return []


def save_xls_file(blocked_list, block_child_list):
    #保存配置文件，保存配置文件文件名
    xlsfile = 'HuaWeiUSG.xlsx'

    workbook = openpyxl.load_workbook(xlsfile)
    vaild_counter = 1
    print(xlsfile)
    wb = workbook
    writecell = []

#Edit sheet "interface" begin
    sheet = wb.get_sheet_by_name('interface')
    cellrow = 2
    for i in range(len(blocked_list)):
        tempStr1 = blocked_list[i]
        if 'interface' == tempStr1[:9] :
            interface_name = tempStr1[10:len(tempStr1)]
            tempList1 = interface_name.split()
            if len(tempList1) == 3:                     #interface name like 'X3 vlan 2',fix it
                interface_name = tempList1[0]+':V'+tempList1[2]
            tempList1 = block_child_list[i]
            tempList2 = tempList1[0].split()            #catch alias name
            if len(tempList2) == 3:
                tempStr2 = tempList2[1].strip()
                ipassignment = tempList2[2].strip()
            else:
                tempStr2 = '-'
                ipassignment = '-'
            interface_alias = tempStr2
            if interface_alias != '-' :                 #if have ip address
                tempList2 = tempList1[1].split()            #ip / netmask
                if len(tempList2) == 4:
                    interface_ip = tempList2[1].strip()
                    interface_netmask = tempList2[3].strip()
                if len(tempList2) == 2 and tempList2[0]=='ip':
                    interface_ip = tempList2[1].strip()
                    tempList2 = tempList1[2].split()
                    interface_netmask = tempList2[1].strip()
            else:
                interface_ip = '-'
                interface_netmask = '-'
            for i in range(1,len(tempList1)):
                tempStr1 = tempList1[i].strip()
                if len(tempStr1) > 7 :
                    if tempStr1[:7] == 'comment':
                        interface_comment = tempStr1[8:len(tempStr1)]
                        break
                else:
                    interface_comment = ' '
            cellcolumn = 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_name
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_alias
            cellcolumn += 1
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_ip
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_netmask
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = ipassignment
            cellcolumn += 1
            cellcolumn += 1
#            sheet.cell(row=cellrow, column=cellcolumn).value = interface_comment
            cellrow += 1
# Edit sheet "interface" end

# Edit sheet "route" begin
    sheet = wb.get_sheet_by_name('route')
    cellrow = 3
    for i in range(len(blocked_list)):

        tempStr1 = blocked_list[i]
        if 'ip route-static' == tempStr1[:15]:
            child_List = block_child_list[i]
            for j in range(len(child_List)):
                tempStr2 = child_List[j].strip()
                if len(tempStr2) > 16 :
                    if  'policy interface' == tempStr2[:16] :
                        tempStr3 = child_List[j+1].strip()
                        tempList3 = tempStr3.split()
                        route_id = tempList3[1]

                        tempStr3 = child_List[j+2].strip()
                        tempList3 = tempStr3.split()
                        route_interface = tempList3[1]

                        tempStr3 = child_List[j+3].strip()
                        tempList3 = tempStr3.split()
                        route_metric = tempList3[1]

                        tempStr3 = child_List[j+4].strip()
                        tempList3 = tempStr3.split()
                        route_source = tempList3[1]

                        tempStr3 = child_List[j+5].strip()
                        tempList3 = tempStr3.split()
                        if len(tempList3) > 2:
#                            route_distination = tempList3[1] + tempList3[2]
                            route_distination =  tempList3[2]
                        else:
                            route_distination = tempList3[1]

                        tempStr3 = child_List[j+6].strip()
                        tempList3 = tempStr3.split()
                        route_service = tempList3[1]

                        tempStr3 = child_List[j+7].strip()
                        tempList3 = tempStr3.split()
                        if len(tempList3) > 2:
                            route_gateway = tempList3[2]
                            if route_gateway[0] == '"':
                                tempList3 = tempStr3.split('"')
                                route_gateway = '"'+tempList3[1]+'"'
                        else:
                            route_gateway = tempList3[1]

                        tempStr3 = child_List[j+8].strip()
                        tempList3 = tempStr3.split()
                        route_comment = tempList3[1]

#                        print(route_id,route_interface,route_metric,route_source,route_distination,route_service,route_gateway,route_comment)
                        cellcolumn = 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_id
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_source
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_distination
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_service
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_gateway
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_interface
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_metric
                        cellcolumn += 1
                        cellrow += 1
# Edit sheet "route" end

# Edit sheet "rule" begin

    rule_name = ''
    rule_source_zone = 'any'
    rule_destination_zone = 'any'
    rule_source_address = 'any'
    rule_destination_address = 'any'
    rule_source_address_d = 'any'
    rule_destination_address_d = 'any'
    rule_service_d = 'any'
    rule_service = 'any'
    rule_action = ''
    rule_source_address_l = []
    rule_destination_address_l = []
    rule_service_l = []


    sheet = wb.get_sheet_by_name('rule')
    cellrow = 2
    for i in range(len(blocked_list)):
        if 'security-policy' in blocked_list[i]:   #remove include "ipv6" string
            in_security_polich = True
            while in_security_polich:
                i = i + 1
                tempStr1 = blocked_list[i]
                if tempStr1[0] != " ":
                    in_security_polich = False
                if ' rule name' == tempStr1[:10] :
#                    print (type(rule_source_address_l))

                    for tempInt4 in range(len(rule_source_address_l)):
                        rule_source_address_d = rule_source_address_d + rule_source_address_l[tempInt4]
                        if tempInt4 < len(rule_source_address_l) - 1:
                            rule_source_address_d = rule_source_address_d + '\n'
                    for tempInt4 in range(len(rule_destination_address_l)):
                        rule_destination_address_d = rule_destination_address_d + rule_destination_address_l[tempInt4]
                        if tempInt4 < len(rule_destination_address_l) - 1:
                            rule_destination_address_d = rule_destination_address_d + '\n'

                    for tempInt4 in range(len(rule_service_l)):
                        rule_service_d = rule_service_d + rule_service_l[tempInt4]
                        if tempInt4 < len(rule_service_l) - 1:
                            rule_service_d = rule_service_d + '、'

                    if cellrow > 2 :
                        if rule_source_zone == '':
                            rule_source_zone = 'any'
                        if rule_destination_zone == '':
                            rule_destination_zone = 'any'
                        if rule_source_address == '':
                            rule_source_address = 'any'
                        if rule_destination_address == '':
                            rule_destination_address = 'any'
                        if rule_service == '':
                            rule_service = 'any'

                        cellcolumn = 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = cellrow - 2
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = rule_name
                        cellcolumn += 1

                        sheet.cell(row=cellrow, column=cellcolumn).value = rule_source_zone
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = rule_destination_zone
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = rule_source_address
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = rule_source_address_d
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = rule_destination_address
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = rule_destination_address_d
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = rule_service
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = rule_service_d
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = rule_action

                    rule_source_zone = ''
                    rule_destination_zone = ''
                    rule_source_address = ''
                    rule_destination_address = ''
                    rule_source_address_d = ''
                    rule_destination_address_d = ''
                    rule_service_d = ''
                    rule_service = ''
                    rule_action = ''
                    rule_source_address_l = []
                    rule_destination_address_l = []
                    rule_service_l = []

                    rule_name = tempStr1[10:]
                    rule_name = rule_name.strip()
                    cellrow = cellrow + 1


                elif '  source-zone' == tempStr1[:13]:
                    rule_source_zone = rule_source_zone + tempStr1[13:]
                    rule_source_zone = rule_source_zone.strip()

                elif '  destination-zone' == tempStr1[:18]:
                    rule_destination_zone = rule_destination_zone + tempStr1[18:]
                    rule_destination_zone = rule_destination_zone.strip()

                elif '  source-address' == tempStr1[:16]:
                    tmp_rule_str = tempStr1[16+13:]
                    tmp_rule_str = tmp_rule_str.strip()
                    if len(rule_source_address) > 3:
                        rule_source_address = rule_source_address + '\n'
                    rule_source_address = rule_source_address + tmp_rule_str
                    tmp_rule_list = getAddressList(blocked_list, block_child_list, tmp_rule_str)
                    for temp_rule_int in range(len(tmp_rule_list)):
                        rule_source_address_l.append(tmp_rule_list[temp_rule_int])

                elif '  destination-address' == tempStr1[:21]:
                    tmp_rule_str = tempStr1[21+13:]
                    tmp_rule_str = tmp_rule_str.strip()
                    if len(rule_destination_address) > 3:
                        rule_destination_address = rule_destination_address + '\n'
                    rule_destination_address = rule_destination_address + tmp_rule_str
                    tmp_rule_list = getAddressList(blocked_list, block_child_list, tmp_rule_str)
                    for temp_rule_int in range(len(tmp_rule_list)):
                        rule_destination_address_l.append(tmp_rule_list[temp_rule_int])
                    # rule_destination_address = rule_destination_address + tempStr1[21+13:]
                    # rule_destination_address = rule_destination_address.strip()
                    # rule_temp_l = getAddressList(blocked_list, block_child_list, rule_destination_address)
                    # rule_destination_address_l = rule_temp_l

                elif '  service' == tempStr1[:9]:
                    tmp_rule_str = tempStr1[9:]
                    tmp_rule_str = tmp_rule_str.strip()
                    rule_service = rule_service + tmp_rule_str + '\n'
                    tmp_rule_list = getServiceList(blocked_list, block_child_list, tmp_rule_str)
                    for temp_rule_int in range(len(tmp_rule_list)):
                        rule_service_l.append(tmp_rule_list[temp_rule_int])
                    # rule_service =  rule_service + tempStr1[9:]
                    # rule_service = rule_service.strip()
                    # rule_service_l = getServiceList(blocked_list, block_child_list, rule_service)

                elif '  action' == tempStr1[:8]:
                    rule_action = tempStr1[8:]
                    rule_action = rule_action.strip()

                else:
                    print('Unknow keywork.'+ tempStr1)

            #process last line begin
            for tempInt4 in range(len(rule_source_address_l)):
                rule_source_address_d = rule_source_address_d + rule_source_address_l[tempInt4]
                if tempInt4 < len(rule_source_address_l) - 1:
                    rule_source_address_d = rule_source_address_d + '\n'
            for tempInt4 in range(len(rule_destination_address_l)):
                rule_destination_address_d = rule_destination_address_d + rule_destination_address_l[tempInt4]
                if tempInt4 < len(rule_destination_address_l) - 1:
                    rule_destination_address_d = rule_destination_address_d + '\n'

            for tempInt4 in range(len(rule_service_l)):
                rule_service_d = rule_service_d + rule_service_l[tempInt4]
                if tempInt4 < len(rule_service_l) - 1:
                    rule_service_d = rule_service_d + '、'

            cellcolumn = 1
            sheet.cell(row=cellrow, column=cellcolumn).value = cellrow - 2
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = rule_name
            cellcolumn += 1

            sheet.cell(row=cellrow, column=cellcolumn).value = rule_source_zone
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = rule_destination_zone
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = rule_source_address
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = rule_source_address_d
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = rule_destination_address
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = rule_destination_address_d
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = rule_service
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = rule_service_d
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = rule_action
            # process last line end

    # Edit sheet "rule" end
    try:
        workbook.save('hw_new.xlsx')
    except:
        print("xlsx文件被锁定，无法保存。。。")
    finally:
        workbook.close()

def format_service_list(s_list):
    #格式化服务列表，将重名的内容去除。
    if s_list == "":
        return ""
    f_service_str = ''
    last_str1 = ""
    tmpList1 = s_list.split('\n')
    for tempInt1 in range(len(tmpList1)):
            tmp_str2 = tmpList1[tempInt1]
            tmpList2 = tmp_str2.split()
            if len(tmpList2) == 3:
                if last_str1 != tmpList2[0]:
                    if last_str1 == "":
                        if f_service_str == '':
                            f_service_str = tmpList2[0]
                        else:
                            f_service_str = f_service_str + "、" + tmpList2[0]
                    else:
                        f_service_str = f_service_str + "\n" +tmpList2[0]
                    last_str1 = tmpList2[0]
                    if tmpList2[1] == tmpList2[2]:
                        f_service_str = f_service_str + " " + tmpList2[1]
                    else:
                        f_service_str = f_service_str + " " + tmpList2[1]+"-" + tmpList2[2]
                else:
                    if tmpList2[1] == tmpList2[2]:
                        f_service_str = f_service_str + "、" + tmpList2[1]
                    else:
                        f_service_str = f_service_str + "、" + tmpList2[1]+"-" + tmpList2[2]
            else:
                if len(tmpList2) == 1:
                    if f_service_str == '':
                        f_service_str = tmp_str2
                    else:
                        f_service_str = f_service_str + "、" + tmp_str2

    return f_service_str

def get_first_word(str):
    tstr1 = str.lstrip(' ')
    tlist = tstr1.split(' ')
    return tlist[0]

def list_to_string(rawlist):
    tstr = ''
    for tlist1 in rawlist:
        tlist2 = "\n".join(str(elm) for elm in tlist1)
        tstr =tstr + tlist2 + '\n'
    return tstr


def get_block(filename, confile_blocked, wfilename):
    confile_block = []
    confile_raw = []
    confile_temp = []
    fpline = ''
    with open(filename, 'r', encoding='gbk') as fp:
        for fpline in fp:
            if len(fpline) > 1:
                fpline = fpline.rstrip('\n')
                if len(fpline.strip(" ")) < 1:
                    continue
                confile_raw.append(fpline)
    confile_block = copy.deepcopy(confile_raw)
    confile_blocked = copy.deepcopy(confile_raw)

    save_xls_file(confile_blocked, confile_block)


def get_xls_keys(workbook, keyslist):
    wb = workbook
    sheet = wb.get_sheet_by_name('route')
    cellrow = 1
    cellcolumn = 1
    sheetcell = sheet.cell(row=cellrow, column=cellcolumn).value
    while sheetcell != None:
        celllist = sheetcell.split()
        keyslist.append(celllist)
        cellrow += 1
        #    print (sheetcell)
        sheetcell = sheet.cell(row=cellrow, column=cellcolumn).value
    wb.close()

def cfgxlsproc():
    xlsfile = 'HuaWeiUSG.xlsx'
    keyslist = []

    confile_blocked = []
    swcfgfile = 'vrpcfg.cfg'

    get_block(swcfgfile, confile_blocked, 'hw_block.txt')

if __name__ == '__main__':
    cfgxlsproc()
