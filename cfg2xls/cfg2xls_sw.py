# -*- coding: UTF-8 -*-
'''
根据防火墙配置文件配置内容，提取分类内容，保存到excel文件不同的sheet相应栏目中
author:jason chan
2020-03-11 16:28
'''
# import os
import openpyxl
import datetime
import re

var_global_MGMT_IP = 'none'
var_global_host_dict = {'1.1.1.101':'GSM第一道sonicwall防火墙策略',
    '1.1.1.102':'GSM第二道sonicwall防火墙墙策略',
    '192.168.168.168':'SM外部sonicwall防火墙策略'}

def judge_legal_ip(one_str):
    '''''
    正则匹配方法
    判断一个字符串是否是合法IP地址
    '''
    compile_ip = re.compile(r'^((25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(25[0-5]|2[0-4]\d|[01]?\d\d?)$')
    if compile_ip.match(one_str):
        return True
    else:
        return False

def fix_address_string(add_str):
    #停用    return add_str
    fix_addr_str = "无IP地址"
    if add_str[0]=='"':
        tmp_pos = add_str[1:].find('"')
        aname = add_str[0:tmp_pos+2]
        tmp_list = add_str[tmp_pos:].split()
    else:
        tmp_list = add_str.split()
        aname = tmp_list[0]
    if len(tmp_list)< 3:
        return add_str
    if 'host' == tmp_list[1]:
        fix_addr_str = tmp_list[2] + '/32'

    if 'range' == tmp_list[1]:
        fix_addr_str = tmp_list[2] + '-' + tmp_list[3]

    if 'network' == tmp_list[1]:
        if '255.255.255.0' == tmp_list [3]:
            netmaskstr = '24'
        elif '255.255.255.248' == tmp_list [3]:
            netmaskstr = '29'
        elif '255.255.255.128' == tmp_list [3]:
            netmaskstr = '25'
        else:
            netmaskstr = tmp_list [3]
        fix_addr_str = tmp_list[2] +'/'+ netmaskstr
    if 'RemoteAccess Networks' in add_str:
        fix_addr_str = 'RemoteAccessNetworks'
    if fix_addr_str == '无IP地址':
        print('无IP地址:' ,add_str)
    return fix_addr_str


def save_block_file(blocked_list, block_child_list, fn):
    #配置文件一块为单位，进行保存
    with open(fn, 'w') as fopen:
        for i in range(len(blocked_list)):
            if 'ipv6' in blocked_list[i]:
                continue
            fopen.write('\n===========\n')
            fopen.writelines(blocked_list[i])
#            if block_list[i] != 'null':
            fopen.writelines(block_child_list[i])

def getAddressGroupList(blocked_list, block_child_list,Aname,AddressGroupList):
    #对地址明细进行解释，解析address-group明细，获取具体地址列表
    address_g_list = []
    if '"' in Aname:
        Aname = Aname[1:len(Aname) - 1]

    for i in range(len(blocked_list)):
        if 'ipv6' in blocked_list[i]:  # remove include "ipv6" string
            continue
        tempStr1 = blocked_list[i]
        if 'address-group' == tempStr1[:13]:

            if '"' in tempStr1:
                tempList1 = tempStr1.split('"')
#                tempGname = '"'+tempList1[1]+'"'
                tempAname = tempList1[1]
            else:
                tempList1 = tempStr1.split()
#                print (tempList1)
                tempAname = tempList1[2]
            if Aname == tempAname:
                tempList2 = block_child_list[i]
                for j in range(len(tempList2)):
                    tempStr2 = tempList2[j].strip()
                    if 'address-object' == tempStr2[:14]:
                        tempGetSObject = tempStr2[15+5:len(tempStr2)]
                        AddressGroupList.append(getAddressList(blocked_list, block_child_list,tempGetSObject))
                    if 'address-group' == tempStr2[:13]:
                        tempGetSObject = tempStr2[14+5:len(tempStr2)]
                        tempGetSObject = tempGetSObject.strip()
                        tempList3 = getServiceGroupList(blocked_list, block_child_list,tempGetSObject,AddressGroupList)
#                        for tempInt1 in range(len(tempList3)):
#                            AddressGroupList.append(tempList3[tempInt1])
                break
    return AddressGroupList



def getServiceGroupList(blocked_list, block_child_list,Gname,ServicePortList):
    # 对服务（端口）明细进行解释，解析service-group明细，获取具体地址列表
    #    ServicePortList = []
    if Gname == 'ICMP':
        ServicePortList.append('ICMP')
        return ServicePortList
    if Gname == 'Ping':
        ServicePortList.append('Ping')
        return ServicePortList

    if Gname == '5000&ping':
        Gname = Gname

    for i in range(len(blocked_list)):
        if 'ipv6' in blocked_list[i]:  # remove include "ipv6" string
            continue
        tempStr1 = blocked_list[i]
        if 'service-group' == tempStr1[:13]:
            if '"' in Gname:
                Gname = Gname[1:len(Gname)-1]

            if '"' in tempStr1:
                tempList1 = tempStr1.split('"')
#                tempGname = '"'+tempList1[1]+'"'
                tempGname = tempList1[1]
            else:
                tempList1 = tempStr1.split()
#                print (tempList1)
                tempGname = tempList1[1]
            if Gname == tempGname:
                tempList2 = block_child_list[i]
                for j in range(len(tempList2)):
                    tempStr2 = tempList2[j].strip()
                    if 'service-object' == tempStr2[:14]:
                        tempGetSObject = tempStr2[15:len(tempStr2)]
                        ServicePortList.append(getServiceList(blocked_list, block_child_list,tempGetSObject))
                    if 'service-group' == tempStr2[:13]:
                        tempGetSObject = tempStr2[14:len(tempStr2)]
                        tempGetSObject = tempGetSObject.strip()

                        tempList3 = getServiceGroupList(blocked_list, block_child_list,tempGetSObject,ServicePortList)
#                        ServicePortList = ServicePortList + tempList3

    return ServicePortList

def getServiceList(blocked_list, block_child_list,sname):
    #获取具体服务对应的服务名和端口号
    ServicePort = ''
    for i in range(len(blocked_list)):
        if 'ipv6' in blocked_list[i]:  # remove include "ipv6" string
            continue
        tempStr1 = blocked_list[i]
        if 'service-object' == tempStr1[:14]:
            if sname == tempStr1[15:15+len(sname)] and tempStr1[15+len(sname):15+len(sname)+1]== ' ':
                ServicePort = tempStr1[15 + len(sname)+1:len(tempStr1)]
                break
    return ServicePort

def getAddressList(blocked_list, block_child_list,aname):
    #获取具体地址对应的地址名和地址
    address_detail = ''
    aname = aname.strip()
    if aname =='any':
        return '0.0.0.0'

    if re.search(r'[XU]\d+ IP',aname):
        return('systeminterface')
    #if '新加坡分公司' in aname :
    #    print('aname')

    for i in range(len(blocked_list)):
        tempStr_raw = blocked_list[i]
        tempStr1 = tempStr_raw
        if 'address-object ipv4' == tempStr1[:19]:
            if aname == tempStr1[20:20+len(aname)] and tempStr1[20+len(aname):21+len(aname)]== ' ':
                address_detail = tempStr_raw[len("address-object ipv4 ") :len(tempStr_raw)]
                address_detail = fix_address_string(address_detail)
                break
            #不处理 'ipv6' 
        #处理偶尔出现在address-object前多一个空格情况，另外，如果前面有4个空格以上将是group下的，所以只处理多一个空格
        #上面提及问题处理已在ver652_conv.py中处理
    #print(address_detail)
    if address_detail == '':
        print('ip address not found...',aname)
    return address_detail


def save_xls_file(blocked_list, block_child_list):
    #保存配置文件，保存配置文件文件名
    global var_global_MGMT_IP
    var_system_time = 'none'
    xlsfile = 'sonicwall.xlsx'

    workbook = openpyxl.load_workbook(xlsfile)
    vaild_counter = 1
    print(xlsfile)
    wb = workbook
    writecell = []

    #Edit sheet "interface" begin
    sheet = wb['interface']
    cellrow = 2
    for i in range(len(blocked_list)):
        if 'ipv6' in blocked_list[i]:   #remove include "ipv6" string
            continue

        tempStr1 = blocked_list[i]

        if 'system-time' ==tempStr1[:11]:
            var_system_time = tempStr1[13:-1]

        if 'interface' == tempStr1[:9] :
            interface_name = tempStr1[10:len(tempStr1)]
            tempList1 = interface_name.split()
            if len(tempList1) == 3:                     #interface name like 'X3 vlan 2',fix it
                interface_name = tempList1[0]+':V'+tempList1[2]
            tempList1 = block_child_list[i]
            tempList2 = tempList1[0].split()            #catch alias name
            if len(tempList2) == 3:
                tempStr2 = tempList2[1].strip()
                ipassignment = tempList2[2].strip()
            else:
                tempStr2 = '-'
                ipassignment = '-'
            interface_alias = tempStr2
            if interface_alias != '-' :                 #if have ip address
                tempList2 = tempList1[1].split()            #ip / netmask
                if len(tempList2) == 4:
                    interface_ip = tempList2[1].strip()
                    interface_netmask = tempList2[3].strip()
                if len(tempList2) == 2 and tempList2[0]=='ip':
                    interface_ip = tempList2[1].strip()
                    tempList2 = tempList1[2].split()
                    interface_netmask = tempList2[1].strip()
            else:
                interface_ip = '-'
                interface_netmask = '-'
            for i in range(1,len(tempList1)):
                tempStr1 = tempList1[i].strip()
                if len(tempStr1) > 7 :
                    if tempStr1[:7] == 'comment':
                        interface_comment = tempStr1[8:len(tempStr1)]
                        break
                else:
                    interface_comment = ' '
            cellcolumn = 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_name
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_alias
            cellcolumn += 1
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_ip
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_netmask
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = ipassignment
            cellcolumn += 1
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_comment
            cellrow += 1
            if interface_name == 'MGMT':
                var_global_MGMT_IP = interface_ip
    # Edit sheet "interface" end

    print('cfg file create time:',var_system_time)
    # Edit sheet "route" begin
    sheet = wb['route']
    cellrow = 2
    for i in range(len(blocked_list)):
        if 'ipv6' in blocked_list[i]:  # remove include "ipv6" string
            continue

        tempStr1 = blocked_list[i]
        if 'routing' == tempStr1[:7]:
            child_List = block_child_list[i]
            for j in range(len(child_List)):
                tempStr2 = child_List[j].strip()
                if len(tempStr2) > 16 :
                    if  'policy interface' == tempStr2[:16] :
                        tempStr3 = child_List[j+1].strip()
                        tempList3 = tempStr3.split()
                        route_id = tempList3[1]

                        tempStr3 = child_List[j+2].strip()
                        tempList3 = tempStr3.split()
                        route_interface = tempList3[1]

                        tempStr3 = child_List[j+3].strip()
                        tempList3 = tempStr3.split()
                        route_metric = tempList3[1]

                        tempStr3 = child_List[j+4].strip()
                        tempList3 = tempStr3.split()
                        route_source = tempList3[1]

                        tempStr3 = child_List[j+5].strip()
                        tempList3 = tempStr3.split()
                        if len(tempList3) > 2:
#                            route_distination = tempList3[1] + tempList3[2]
                            route_distination =  tempList3[2]
                        else:
                            route_distination = tempList3[1]

                        tempStr3 = child_List[j+6].strip()
                        tempList3 = tempStr3.split()
                        route_service = tempList3[1]

                        tempStr3 = child_List[j+7].strip()
                        tempList3 = tempStr3.split()
                        if len(tempList3) > 2:
                            route_gateway = tempList3[2]
                            if route_gateway[0] == '"':
                                tempList3 = tempStr3.split('"')
                                route_gateway = '"'+tempList3[1]+'"'
                        else:
                            route_gateway = tempList3[1]

                        tempStr3 = child_List[j+8].strip()
                        tempList3 = tempStr3.split()
                        route_comment = tempList3[1]

#                        print(route_id,route_interface,route_metric,route_source,route_distination,route_service,route_gateway,route_comment)
                        cellcolumn = 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_id
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_source
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_distination
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_service
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_gateway
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_interface
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_metric
                        cellcolumn += 1
#配置文件没找到优先级                        sheet.cell(row=cellrow, column=cellcolumn).value = route_priority
                        cellrow += 1
# Edit sheet "route" end

#        print(str(cellrow) + interface_name+':'+interface_alias + interface_ip+';'+ interface_netmask + interface_comment)

# Edit sheet "rule" begin
    print('MGMT: ',var_global_MGMT_IP)
    sheet = wb['rule']
    cellrow = 3
    for i in range(len(blocked_list)):
        if 'ipv6' in blocked_list[i]:   #remove include "ipv6" string
            continue

        tempStr1 = blocked_list[i]
        if 'access-rule' == tempStr1[:11] :
#            print(tempStr1)

            rule_service_port = ''
            rule_address_detail= ''
            rule_destination_detail = ''
            rule_destination_list = []
            rule_name = tempStr1[10:30]
            tempList1 = block_child_list[i]
            for rule_child in tempList1:
                tempStr1 = rule_child.strip()
                tempList2 = tempStr1.split()
                if len(tempList2) < 2:
                    continue
                if len(tempList2) > 1:
#                    if 'id' == tempList2[0]:
                    if 'id' == tempList2[0] or 'uuid' == tempList2[0]:
                        rule_id = tempList2[1]
                    if 'from' == tempList2[0]:
                        rule_from = tempList2[1]
                    if 'to' == tempList2[0]:
                        rule_to = tempList2[1]
                    if 'action' == tempList2[0]:
                        rule_action = tempList2[1]
                    if 'source' == tempList2[0]:
                        if 'address' == tempList2[1]:
                            if len(tempList2)>3:
                                rule_source_address = ''
                                for k in range(3,len(tempList2)):
                                    rule_source_address = rule_source_address + tempList2[k] + ' '
                                rule_source_address = rule_source_address.strip()
####
                                if '&' in rule_source_address:
                                    aa = 0
####
                                if tempList2[2] == 'name':
                                    rule_address_detail = getAddressList(blocked_list, block_child_list, rule_source_address)
                                if tempList2[2] == 'group':
                                    gList_temp = []
                                    gList_temp = getAddressGroupList(blocked_list, block_child_list, rule_source_address,gList_temp)

                                    for tempInt1 in range(len(gList_temp)):
                                        rule_address_detail = rule_address_detail + gList_temp[tempInt1] + '\n'

                            else:
                                rule_source_address = tempList2[2]
                        if 'port' == tempList2[1]:
                            rule_source_port = tempList2[2]
                        rule_source_address = rule_source_address.strip()
                    if 'destination' == tempList2[0]:
                        if 'address' == tempList2[1]:
                            if len(tempList2) > 3:
                                rule_destination_address = ''
                                for k in range(3, len(tempList2)):
                                    rule_destination_address = rule_destination_address + tempList2[k] + ' '
                                rule_destination_address = rule_destination_address.strip()
                                if tempList2[2] == 'name':
                                    rule_destination_detail = getAddressList(blocked_list, block_child_list, rule_destination_address)
                                if tempList2[2] == 'group':
                                    gList_temp = []
                                    gList_temp = getAddressGroupList(blocked_list, block_child_list, rule_destination_address,gList_temp)

                                    for tempInt1 in range(len(gList_temp)):
                                        rule_destination_detail = rule_destination_detail + gList_temp[tempInt1] + '\n'

                            else:
                                rule_destination_address = tempList2[2]
                        rule_destination_address = rule_destination_address.strip()

                    if 'service' == tempList2[0]:
                        if (len(tempList2)) <3:
                            rule_service = tempList2[1]
                            rule_service = rule_service.strip()
                        else:
                            rule_service = ''
                            for k in range(2, len(tempList2)):
                                rule_service = rule_service + tempList2[k] + ' '
                            rule_service = rule_service.strip()
                            if tempList2[1] == 'name':
                                rule_service_port = rule_service_port + getServiceList(blocked_list, block_child_list, rule_service)
                            if tempList2[1] == 'group':
                                ServicePortList = []
                                rule_service_port_list = getServiceGroupList(blocked_list, block_child_list, rule_service,ServicePortList)
#add 2018
                                rule_service_port_list.sort()
                                for tempInt1 in range(len(rule_service_port_list)):
#                                    print(rule_service_port_list)
                                    rule_service_port = rule_service_port + rule_service_port_list[tempInt1] + '\n'
            format_rule_str = format_service_list(rule_service_port)
#            format_rule_str = rule_service_port

            if rule_from == rule_to or rule_from == 'VPN' or rule_from == 'SSLVPN' or rule_to == 'SSLVPN' or\
                    (rule_source_address == 'any' and rule_destination_address == 'any' and rule_service == 'any') or \
                    rule_source_address == '"WLAN RemoteAccess Networks"' or rule_source_address == '"WAN RemoteAccess Networks"':
                pass
            else:

                cellcolumn = 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_id
                sheet.cell(row=cellrow, column=cellcolumn).value = vaild_counter
                vaild_counter +=1
                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_from
                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = '>'
                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_to
                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_source_address
                cellcolumn += 1

                row_diaplay = rule_address_detail.split('\n')
                rule_source_address_row = len(row_diaplay)
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_address_detail

                cellcolumn += 1

                sheet.cell(row=cellrow, column=cellcolumn).value = rule_destination_address
                cellcolumn += 1


                row_diaplay = rule_destination_detail.split('\n')
                rule_destination_address_row = len(row_diaplay)
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_destination_detail

                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_service
                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_service_port
                sheet.cell(row=cellrow, column=cellcolumn).value = format_rule_str
                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_action

                if rule_source_address_row > rule_destination_address_row :
                    sheet.row_dimensions[cellrow].height = 16 * rule_source_address_row
                else :
                    sheet.row_dimensions[cellrow].height = 16 * rule_destination_address_row

                cellrow +=1

    sheet.row_dimensions[cellrow].height = 16 * 4
    sheet.merge_cells(start_row=cellrow, start_column=1, end_row=cellrow, end_column=11)
    ending_string='''                                                              检查人：                                     检查日期：

                                                              审核人：                                     审核日期：'''
    sheet.cell(cellrow,1).value=ending_string
    var_system_time = var_system_time[:10]
    date_p_ = datetime.datetime.strptime(var_system_time, '%m/%d/%Y').date()
    sheet.cell(1,1).value = var_system_time[6:]+'年'+var_system_time[:2]+'月  ' +var_global_host_dict[var_global_MGMT_IP]

    for i in range(cellrow+1,300):
        sheet.delete_rows(cellrow+1)

# Edit sheet "rule" end
    try:
        workbook.save('cfg_new.xlsx')
    except:
        print("xlsx文件被锁定，无法保存。。。")
    finally:
        workbook.close()

def format_service_list(s_list):
    #格式化服务列表，将重名的内容去除。
    if s_list == "":
        return ""
    f_service_str = ''
    last_str1 = ""
    tmpList1 = s_list.split('\n')
    for tempInt1 in range(len(tmpList1)):
            tmp_str2 = tmpList1[tempInt1]
            tmpList2 = tmp_str2.split()
            if len(tmpList2) == 3:
                if last_str1 != tmpList2[0]:
                    if last_str1 == "":
                        if f_service_str == '':
                            f_service_str = tmpList2[0]
                        else:
                            f_service_str = f_service_str + "、" + tmpList2[0]
                    else:
                        f_service_str = f_service_str + "\n" +tmpList2[0]
                    last_str1 = tmpList2[0]
                    if tmpList2[1] == tmpList2[2]:
                        f_service_str = f_service_str + " " + tmpList2[1]
                    else:
                        f_service_str = f_service_str + " " + tmpList2[1]+"-" + tmpList2[2]
                else:
                    if tmpList2[1] == tmpList2[2]:
                        f_service_str = f_service_str + "、" + tmpList2[1]
                    else:
                        f_service_str = f_service_str + "、" + tmpList2[1]+"-" + tmpList2[2]
            else:
                if len(tmpList2) == 1:
                    if f_service_str == '':
                        f_service_str = tmp_str2
                    else:
                        f_service_str = f_service_str + "、" + tmp_str2

    return f_service_str

def get_first_word(str):
    tstr1 = str.lstrip(' ')
    tlist = tstr1.split(' ')
    return tlist[0]

def list_to_string(rawlist):
    tstr = ''
    for tlist1 in rawlist:
        tlist2 = "\n".join(str(elm) for elm in tlist1)
        tstr =tstr + tlist2 + '\n'
    return tstr


def get_block(filename, confile_blocked, wfilename):
    confile_block = []
    confile_raw = []
    confile_temp = []
    fpline = ''
    with open(filename, 'r', encoding='gbk') as fp:
        for fpline in fp:
            #print(len(fpline),fpline)
            #        if len(fpline) > 3 and fpline[0] != '#': 井号也是标志，不可清理
            if len(fpline) > 1:
                fpline = fpline.rstrip('\n')
                if len(fpline.strip(" ")) < 1:
                    continue
                confile_raw.append(fpline)
    for line_num in range(len(confile_raw) - 1):
        # = confile_raw.index(fpline)
        lineStr = confile_raw[line_num]
        space1 = len(confile_raw[line_num]) - len(confile_raw[line_num].lstrip(' '))
        space2 = len(confile_raw[line_num + 1]) - len(confile_raw[line_num + 1].lstrip(' '))
        step1 = 0
 #       step2 = 1
        if space1 == 0 and space2 == 0 :
            confile_blocked.append (lineStr)
            confile_block.append ("null")
        if space1 == 0 and space2 > 0:
            confile_blocked.append(lineStr)

        if space1 > 0 and space2 > 0:
            confile_temp.append('\n'+ lineStr)
#            confile_temp.append('\n')
        if space1 > 0 and space2 == 0:
            confile_temp.append('\n'+ lineStr)
            confile_block.append(confile_temp)
            confile_temp = []
    save_block_file(confile_blocked, confile_block, wfilename)
    save_xls_file(confile_blocked, confile_block)


def get_xls_keys(workbook, keyslist):
    wb = workbook
    sheet = wb.get_sheet_by_name('route')
    cellrow = 1
    cellcolumn = 1
    sheetcell = sheet.cell(row=cellrow, column=cellcolumn).value
    while sheetcell != None:
        celllist = sheetcell.split()
        keyslist.append(celllist)
        cellrow += 1
        #    print (sheetcell)
        sheetcell = sheet.cell(row=cellrow, column=cellcolumn).value
    wb.close()

def cfgxlsproc():
    xlsfile = 'sonicwall.xlsx'
    keyslist = []

    confile_blocked = []
    swcfgfile = 'sw.log'

    get_block(swcfgfile, confile_blocked, 'sn_block.txt')

if __name__ == '__main__':
    cfgxlsproc()
