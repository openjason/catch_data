# -*- coding: UTF-8 -*-
'''
根据防火墙配置文件配置内容，提取分类内容，保存到excel文件不同的sheet相应栏目中
author:jason chan
2017-11-29
'''
# import os
import openpyxl

import re

def judge_legal_ip(one_str):
    '''''
    正则匹配方法
    判断一个字符串是否是合法IP地址
    '''
    compile_ip = re.compile('^((25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(25[0-5]|2[0-4]\d|[01]?\d\d?)$')
    if compile_ip.match(one_str):
        return True
    else:
        return False

def fix_address_string(add_str):
#    return add_str
    tmp_list = add_str.split()
    if len(tmp_list)< 3:
        return add_str
    if 'host' == tmp_list[0]:
        fix_addr_str = tmp_list[0] + ' ' + tmp_list[1] + '/32'
        return fix_addr_str

    if 'range' == tmp_list[0]:
        fix_addr_str = tmp_list[0] + ' ' + tmp_list[1] + '-' + tmp_list[2]
        return fix_addr_str

    if 'network' == tmp_list[0]:
        if '255.255.255.0' == tmp_list [2]:
            netmaskstr = '/24'
        else:
            netmaskstr = tmp_list [2]
        fix_addr_str = tmp_list[0] +' '+ tmp_list[1] + netmaskstr
    else:
        fix_addr_str = add_str
    return fix_addr_str


def save_block_file(blocked_list, block_child_list, fn):
    with open(fn, 'w') as fopen:
        for i in range(len(blocked_list)):
            if 'ipv6' in blocked_list[i]:
                continue
            fopen.write('\n===========\n')
            fopen.writelines(blocked_list[i])
#            if block_list[i] != 'null':
            fopen.writelines(block_child_list[i])

def getAddressGroupList(blocked_list, block_child_list,Aname,AddressGroupList):
    address_g_list = []
    if '&' in Aname :
        a= 0
    if '"' in Aname:
        Aname = Aname[1:len(Aname) - 1]

    for i in range(len(blocked_list)):
        if 'ipv6' in blocked_list[i]:  # remove include "ipv6" string
            continue
        tempStr1 = blocked_list[i]
        if 'address-group' == tempStr1[:13]:

            if '200.40' in tempStr1:
                a = 0

            if '"' in tempStr1:
                tempList1 = tempStr1.split('"')
#                tempGname = '"'+tempList1[1]+'"'
                tempAname = tempList1[1]
            else:
                tempList1 = tempStr1.split()
#                print (tempList1)
                tempAname = tempList1[2]
            if Aname == tempAname:
                tempList2 = block_child_list[i]
                for j in range(len(tempList2)):
                    tempStr2 = tempList2[j].strip()
                    if 'address-object' == tempStr2[:14]:
                        tempGetSObject = tempStr2[15+5:len(tempStr2)]
                        AddressGroupList.append(getAddressList(blocked_list, block_child_list,tempGetSObject))
                    if 'address-group' == tempStr2[:13]:
                        tempGetSObject = tempStr2[14+5:len(tempStr2)]
                        tempGetSObject = tempGetSObject.strip()
                        tempList3 = getServiceGroupList(blocked_list, block_child_list,tempGetSObject,AddressGroupList)
#                        for tempInt1 in range(len(tempList3)):
#                            AddressGroupList.append(tempList3[tempInt1])
                break
    return AddressGroupList



def getServiceGroupList(blocked_list, block_child_list,Gname,ServicePortList):
#    ServicePortList = []
    if Gname == 'ICMP' or Gname == 'Ping':
        return ['ICMP']



    for i in range(len(blocked_list)):
        if 'ipv6' in blocked_list[i]:  # remove include "ipv6" string
            continue
        tempStr1 = blocked_list[i]
        if 'service-group' == tempStr1[:13]:
            if '"' in Gname:
                Gname = Gname[1:len(Gname)-1]

            if '"' in tempStr1:
                tempList1 = tempStr1.split('"')
#                tempGname = '"'+tempList1[1]+'"'
                tempGname = tempList1[1]
            else:
                tempList1 = tempStr1.split()
#                print (tempList1)
                tempGname = tempList1[1]
            if Gname == tempGname:
                tempList2 = block_child_list[i]
                for j in range(len(tempList2)):
                    tempStr2 = tempList2[j].strip()
                    if 'service-object' == tempStr2[:14]:
                        tempGetSObject = tempStr2[15:len(tempStr2)]
                        ServicePortList.append(getServiceList(blocked_list, block_child_list,tempGetSObject))
                    if 'service-group' == tempStr2[:13]:
                        tempGetSObject = tempStr2[14:len(tempStr2)]
                        tempGetSObject = tempGetSObject.strip()

                        tempList3 = getServiceGroupList(blocked_list, block_child_list,tempGetSObject,ServicePortList)
#                        for tempInt1 in range(len(tempList3)):
#                            ServicePortList.append(tempList3[tempInt1])
    return ServicePortList

def getServiceList(blocked_list, block_child_list,sname):
    ServicePort = ''
    for i in range(len(blocked_list)):
        if 'ipv6' in blocked_list[i]:  # remove include "ipv6" string
            continue
        tempStr1 = blocked_list[i]
        if 'service-object' == tempStr1[:14]:
            if sname == tempStr1[15:15+len(sname)] and tempStr1[15+len(sname):15+len(sname)+1]== ' ':
                ServicePort = tempStr1[15 + len(sname)+1:len(tempStr1)]
                break
    return ServicePort

def getAddressList(blocked_list, block_child_list,aname):
    address_detail = ''
    if aname =='any':
        return ' '

    for i in range(len(blocked_list)):
        if 'ipv6' in blocked_list[i]:  # remove include "ipv6" string
            continue
        tempStr1 = blocked_list[i]
        if 'address-object' == tempStr1[:14]:
            if aname == tempStr1[20:20+len(aname)] and tempStr1[20+len(aname):21+len(aname)]== ' ':
                try:
                    first_space = tempStr1[21 + len(aname):len(tempStr1)].index(' ')
                    c_name = tempStr1[21 + len(aname):21 + len(aname)+first_space]
                except:
                    first_space = 0
                address_detail = tempStr1[21 + len(aname)+first_space:len(tempStr1)]
                break
    if cname == 'host':
        if judge_legal_ip(aname):
            address_string = 'host'+address_detail
        else:
            address_string = aname + address_detail

    if cname == 'host':
        if judge_legal_ip(aname):
            address_string = 'host'+address_detail
        else:
            address_string = aname + address_detail

    return address_string


def save_xls_file(blocked_list, block_child_list):
    workbook = openpyxl.load_workbook(xlsfile)

    print(xlsfile)
    wb = workbook
    writecell = []

#Edit sheet "interface" begin
    sheet = wb.get_sheet_by_name('interface')
    cellrow = 2
    for i in range(len(blocked_list)):
        if 'ipv6' in blocked_list[i]:   #remove include "ipv6" string
            continue

        tempStr1 = blocked_list[i]
        if 'interface' == tempStr1[:9] :
            interface_name = tempStr1[10:len(tempStr1)]
            tempList1 = interface_name.split()
            if len(tempList1) == 3:                     #interface name like 'X3 vlan 2',fix it
                interface_name = tempList1[0]+':V'+tempList1[2]
            tempList1 = block_child_list[i]
            tempList2 = tempList1[0].split()            #catch alias name
            if len(tempList2) == 3:
                tempStr2 = tempList2[1].strip()
            else:
                tempStr2 = '-'
            interface_alias = tempStr2
            if interface_alias != '-' :                 #if have ip address
                tempList2 = tempList1[1].split()            #ip / netmask
                if len(tempList2) == 4:
                    interface_ip = tempList2[1].strip()
                    interface_netmask = tempList2[3].strip()
            else:
                interface_ip = '-'
                interface_netmask = '-'
            for i in range(1,len(tempList1)):
                tempStr1 = tempList1[i].strip()
                if len(tempStr1) > 7 :
                    if tempStr1[:7] == 'comment':
                        interface_comment = tempStr1[8:len(tempStr1)]
                        break
                else:
                    interface_comment = ' '
            cellcolumn = 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_name
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_alias
            cellcolumn += 1
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_ip
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_netmask
            cellcolumn += 1
            cellcolumn += 1
            cellcolumn += 1
            sheet.cell(row=cellrow, column=cellcolumn).value = interface_comment
            cellrow += 1
# Edit sheet "interface" end

# Edit sheet "route" begin
    sheet = wb.get_sheet_by_name('route')
    cellrow = 2
    for i in range(len(blocked_list)):
        if 'ipv6' in blocked_list[i]:  # remove include "ipv6" string
            continue

        tempStr1 = blocked_list[i]
        if 'routing' == tempStr1[:7]:
            child_List = block_child_list[i]
            for j in range(len(child_List)):
                tempStr2 = child_List[j].strip()
                if len(tempStr2) > 16 :
                    if  'policy interface' == tempStr2[:16] :
                        tempStr3 = child_List[j+1].strip()
                        tempList3 = tempStr3.split()
                        route_id = tempList3[1]

                        tempStr3 = child_List[j+2].strip()
                        tempList3 = tempStr3.split()
                        route_interface = tempList3[1]

                        tempStr3 = child_List[j+3].strip()
                        tempList3 = tempStr3.split()
                        route_metric = tempList3[1]

                        tempStr3 = child_List[j+4].strip()
                        tempList3 = tempStr3.split()
                        route_source = tempList3[1]

                        tempStr3 = child_List[j+5].strip()
                        tempList3 = tempStr3.split()
                        if len(tempList3) > 2:
#                            route_distination = tempList3[1] + tempList3[2]
                            route_distination =  tempList3[2]
                        else:
                            route_distination = tempList3[1]

                        tempStr3 = child_List[j+6].strip()
                        tempList3 = tempStr3.split()
                        route_service = tempList3[1]

                        tempStr3 = child_List[j+7].strip()
                        tempList3 = tempStr3.split()
                        if len(tempList3) > 2:
                            route_gateway = tempList3[2]
                        else:
                            route_gateway = tempList3[1]

                        tempStr3 = child_List[j+8].strip()
                        tempList3 = tempStr3.split()
                        route_comment = tempList3[1]

#                        print(route_id,route_interface,route_metric,route_source,route_distination,route_service,route_gateway,route_comment)
                        cellcolumn = 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_id
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_source
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_distination
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_service
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_gateway
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_interface
                        cellcolumn += 1
                        sheet.cell(row=cellrow, column=cellcolumn).value = route_metric
                        cellcolumn += 1
#配置文件没找到优先级                        sheet.cell(row=cellrow, column=cellcolumn).value = route_priority
                        cellrow += 1
# Edit sheet "route" end

#        print(str(cellrow) + interface_name+':'+interface_alias + interface_ip+';'+ interface_netmask + interface_comment)

# Edit sheet "rule" begin
    sheet = wb.get_sheet_by_name('rule')
    cellrow = 2
    for i in range(len(blocked_list)):
        if 'ipv6' in blocked_list[i]:   #remove include "ipv6" string
            continue

        tempStr1 = blocked_list[i]
        if 'access-rule' == tempStr1[:11] :
#            print(tempStr1)

            rule_service_port = ''
            rule_address_detail= ''
            rule_destination_detail = ''
            rule_destination_list = []
            rule_name = tempStr1[10:30]
            tempList1 = block_child_list[i]
            for rule_child in tempList1:
                tempStr1 = rule_child.strip()
                tempList2 = tempStr1.split()
                if len(tempList2) < 2:
                    continue
                if len(tempList2) > 1:
                    if 'id' == tempList2[0]:
                        rule_id = tempList2[1]
                    if 'from' == tempList2[0]:
                        rule_from = tempList2[1]
                    if 'to' == tempList2[0]:
                        rule_to = tempList2[1]
                    if 'action' == tempList2[0]:
                        rule_action = tempList2[1]
                    if 'source' == tempList2[0]:
                        if 'address' == tempList2[1]:
                            if len(tempList2)>3:
                                rule_source_address = ''
                                for k in range(3,len(tempList2)):
                                    rule_source_address = rule_source_address + tempList2[k] + ' '
                                rule_source_address = rule_source_address.strip()
####
                                if '&' in rule_source_address:
                                    aa = 0
####
                                if tempList2[2] == 'name':
                                    rule_address_detail = getAddressList(blocked_list, block_child_list, rule_source_address)
                                    rule_address_detail = fix_address_string(rule_address_detail)
                                if tempList2[2] == 'group':
                                    gList_temp = []
                                    gList_temp = getAddressGroupList(blocked_list, block_child_list, rule_source_address,gList_temp)

                                    for tempInt1 in range(len(gList_temp)):
                                        rule_address_detail = rule_address_detail + fix_address_string(gList_temp[tempInt1]) + '\n'

                            else:
                                rule_source_address = tempList2[2]
                        if 'port' == tempList2[1]:
                            rule_source_port = tempList2[2]
                        rule_source_address = rule_source_address.strip()
                    if 'destination' == tempList2[0]:
                        if 'address' == tempList2[1]:
                            if len(tempList2) > 3:
                                rule_destination_address = ''
                                for k in range(3, len(tempList2)):
                                    rule_destination_address = rule_destination_address + tempList2[k] + ' '
                                rule_destination_address = rule_destination_address.strip()
                                if tempList2[2] == 'name':
                                    rule_destination_detail = getAddressList(blocked_list, block_child_list, rule_destination_address)
                                    rule_destination_detail = fix_address_string(rule_destination_detail)
                                if tempList2[2] == 'group':
                                    gList_temp = []
                                    gList_temp = getAddressGroupList(blocked_list, block_child_list, rule_destination_address,gList_temp)

                                    for tempInt1 in range(len(gList_temp)):
                                        rule_destination_detail = rule_destination_detail + fix_address_string(gList_temp[tempInt1]) + '\n'

                            else:
                                rule_destination_address = tempList2[2]
                        rule_destination_address = rule_destination_address.strip()

                    if 'service' == tempList2[0]:
                        if (len(tempList2)) <3:
                            rule_service = tempList2[1]
                            rule_service = rule_service.strip()
                        else:
                            rule_service = ''
                            for k in range(2, len(tempList2)):
                                rule_service = rule_service + tempList2[k] + ' '
                            rule_service = rule_service.strip()
                            if tempList2[1] == 'name':
                                rule_service_port = getServiceList(blocked_list, block_child_list, rule_service)
                            if tempList2[1] == 'group':
                                ServicePortList = []
                                rule_service_port_list = getServiceGroupList(blocked_list, block_child_list, rule_service,ServicePortList)
                                for tempInt1 in range(len(rule_service_port_list)):
#                                    print(rule_service_port_list)
                                    rule_service_port = rule_service_port + rule_service_port_list[tempInt1] + '\n'


            if rule_from == rule_to or rule_from == 'VPN' or \
                    (rule_source_address == 'any' and rule_destination_address == 'any' and rule_service == 'any') or \
                    rule_source_address == '"WLAN RemoteAccess Networks"' or rule_source_address == '"WAN RemoteAccess Networks"':
                pass
            else:

                cellcolumn = 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_id
                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_from
                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = '>'
                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_to
                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_source_address
                cellcolumn += 1

                row_diaplay = rule_address_detail.split('\n')
                rule_source_address_row = len(row_diaplay)
                # if len(row_diaplay) >1:
                #     for i in range(len(row_diaplay)):
                #         sheet.cell(row=cellrow + rule_source_address_row, column=cellcolumn).value = row_diaplay[i]
                #         rule_source_address_row +=1
                # else:
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_address_detail

                cellcolumn += 1

                sheet.cell(row=cellrow, column=cellcolumn).value = rule_destination_address
                cellcolumn += 1


                row_diaplay = rule_destination_detail.split('\n')
                rule_destination_address_row = len(row_diaplay)
                # if len(row_diaplay) > 1:
                #     for i in range(len(row_diaplay)):
                #         sheet.cell(row=cellrow + rule_destination_address_row, column=cellcolumn).value = row_diaplay[i]
                #         rule_destination_address_row += 1
                # else:
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_destination_detail


                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_service
                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_service_port
                cellcolumn += 1
                sheet.cell(row=cellrow, column=cellcolumn).value = rule_action

                if rule_source_address_row > rule_destination_address_row :
                    sheet.row_dimensions[cellrow].height = 16 * rule_source_address_row
                else :
                    sheet.row_dimensions[cellrow].height = 16 * rule_destination_address_row

                cellrow +=1


# Edit sheet "rule" end
    try:
        workbook.save(WorkDir+'cfg_new.xlsx')
    except:
        print("xlsx文件被锁定，无法保存。。。")
    finally:
        workbook.close()


def get_first_word(str):
    tstr1 = str.lstrip(' ')
    tlist = tstr1.split(' ')
    return tlist[0]


def list_to_string(rawlist):
    tstr = ''
    for tlist1 in rawlist:
        tlist2 = "\n".join(str(elm) for elm in tlist1)
        tstr =tstr + tlist2 + '\n'
    return tstr


def get_block(filename, confile_blocked, wfilename):
    confile_raw = []
    confile_temp = []
    fpline = ''
    with open(filename, 'r', encoding='gbk') as fp:
        for fpline in fp:
            #print(len(fpline),fpline)
            #        if len(fpline) > 3 and fpline[0] != '#': 井号也是标志，不可清理
            if len(fpline) > 1:
                fpline = fpline.rstrip('\n')
                if len(fpline.strip(" ")) < 1:
                    continue
                confile_raw.append(fpline)
    for line_num in range(len(confile_raw) - 1):
        # = confile_raw.index(fpline)
        lineStr = confile_raw[line_num]
        space1 = len(confile_raw[line_num]) - len(confile_raw[line_num].lstrip(' '))
        space2 = len(confile_raw[line_num + 1]) - len(confile_raw[line_num + 1].lstrip(' '))
        step1 = 0
 #       step2 = 1
        if space1 == 0 and space2 == 0 :
            confile_blocked.append (lineStr)
            confile_block.append ("null")
        if space1 == 0 and space2 > 0:
            confile_blocked.append(lineStr)

        if space1 > 0 and space2 > 0:
            confile_temp.append('\n'+ lineStr)
#            confile_temp.append('\n')
        if space1 > 0 and space2 == 0:
            confile_temp.append('\n'+ lineStr)
            confile_block.append(confile_temp)
            confile_temp = []
    save_block_file(confile_blocked, confile_block, wfilename)
    save_xls_file(confile_blocked, confile_block)


def get_xls_keys(workbook, keyslist):
    wb = workbook
    sheet = wb.get_sheet_by_name('route')
    cellrow = 1
    cellcolumn = 1
    sheetcell = sheet.cell(row=cellrow, column=cellcolumn).value
    while sheetcell != None:
        celllist = sheetcell.split()
        keyslist.append(celllist)
        cellrow += 1
        #    print (sheetcell)
        sheetcell = sheet.cell(row=cellrow, column=cellcolumn).value
    wb.close()


if __name__ == '__main__':
    WorkDir = 'E:\\test\\'
    xlsfile = WorkDir + 'sonicwall.xlsx'
    filename = 'hw254.cfg'
    wfilename = 'hwblock.txt'
    keyslist = []

    confile_blocked = []
    confile_block = []
    swcfgfile = WorkDir + 'sw.log'
    get_block(swcfgfile, confile_blocked, WorkDir+'sn_block.txt')
