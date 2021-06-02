"""
功能：考勤数据转换脚本。
"""
from tkinter import Tk
from tkinter import (
    messagebox,
    scrolledtext,
    Canvas,
    PhotoImage,
    Label,
    StringVar,
    Entry,
    Button,
    END,
    DISABLED,
    Toplevel,
)  # 导入滚动文本框的模块
from os.path import exists as os_path_exists
from openpyxl import load_workbook
import datetime, time
from openpyxl.cell import cell
import os
import sys
import logging
from logging.handlers import RotatingFileHandler
from copy import copy

#import chardet

# 设置日志文件配置参数
def set_logging():
    global logger
    logger = logging.getLogger("balance_logger")
    handler = RotatingFileHandler("日志记录.log", maxBytes=5000000, backupCount=9)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)
    formatter = logging.Formatter("%(asctime)-12s %(filename)s %(lineno)d %(message)s")
    handler.setFormatter(formatter)


def add_member_patten(ws1,paste_start_row,firstrow,lastrow):
    for rows in range(firstrow,lastrow):
        ws1.row_dimensions[paste_start_row+rows].height = ws1.row_dimensions[rows].height 
        for col in range(1,40):
            #wbsheet_new.column_dimensions[get_column_letter(col)].width = wbsheet.column_dimensions[get_column_letter(col)].width
            #ws1.row_dimensions[get_row_letter(col)].height = ws1.row_dimensions[get_row_letter(col)].height
            if col <5 or col >35:
                ws1.cell(row=paste_start_row+rows,column=col,value=ws1.cell(rows,col).value)

            #if ws1.cell(rows,col).has_style:	#拷贝格式
            ws1.cell(row=paste_start_row+rows,column=col).font = copy(ws1.cell(rows,col).font)
            ws1.cell(row=paste_start_row+rows,column=col).border = copy(ws1.cell(rows,col).border)
            ws1.cell(row=paste_start_row+rows,column=col).fill = copy(ws1.cell(rows,col).fill)
            ws1.cell(row=paste_start_row+rows,column=col).number_format = copy(ws1.cell(rows,col).number_format)
            ws1.cell(row=paste_start_row+rows,column=col).protection = copy(ws1.cell(rows,col).protection)
            ws1.cell(row=paste_start_row+rows,column=col).alignment = copy(ws1.cell(rows,col).alignment)

            ws1.cell(row=paste_start_row+rows,column=col).hyperlink = copy(ws1.cell(rows,col).hyperlink)
            ws1.cell(row=paste_start_row+rows,column=col).comment = copy(ws1.cell(rows,col).comment)


    mergedcell = 'A'+str(paste_start_row + 5)+':A'+str(paste_start_row + 5 + 13)
    ws1.merge_cells(mergedcell)
    mergedcell = 'B'+str(paste_start_row + 5)+':B'+str(paste_start_row + 5 + 13)
    ws1.merge_cells(mergedcell)
    mergedcell = 'C'+str(paste_start_row + 5)+':C'+str(paste_start_row + 5 + 13)
    ws1.merge_cells(mergedcell)
    for tempi in range(5,14,2):
        mergedcell = 'AJ'+str(paste_start_row + tempi)+':AJ'+str(paste_start_row + tempi+1)
        ws1.merge_cells(mergedcell)
        mergedcell = 'AK'+str(paste_start_row + tempi)+':AK'+str(paste_start_row + tempi+1)
        ws1.merge_cells(mergedcell)
    



# 定义类，脚本主要更能
class App:
    def __init__(self, master):

        self.svar_proc_time1 = StringVar()
        self.svar_cangku_filename = StringVar()
        self.svar_kehumingcheng = StringVar()
        self.svar_kehumingcheng2 = StringVar()
        
        self.svar_proc_time2 = StringVar()
        self.svar_youjiqingdan_filename = StringVar()
        self.svar_jichu_filename = StringVar()
        self.svar_label_prompt = StringVar()
        self.master = master
        self.customer_sname = ""
        self.userid_list = []
        self.data_dir = ""
        self.file_from_cangkujxc = ""
        self.file_from_youjiqingdan = ""
        self.file_from_jichu = ""
        self.curr_month = ""
        self.curr_year_month = "2020-09-"
        self.next_year_month = "2020-09-"
        self.initWidgets(master)
        self.is_workday=[0,32]
        #是否工作日，列表内日子计算工时时按星期几的相反操作，即是否扣 120 分钟工作时

    #处理字符串存在非标准内容
    def fix_string_to_standard(self, massstring):
        txtfile_line = massstring
        txtfile_line = txtfile_line.replace("£º", ":")
        txtfile_line = txtfile_line.replace("：", ":")
        txtfile_line = txtfile_line.replace("##", "#")
        txtfile_line = txtfile_line.replace("时间", "")
        txtfile_line = txtfile_line.replace("上班", "")
        txtfile_line = txtfile_line.replace("下班", "")
        txtfile_line = txtfile_line.replace("（", "")
        txtfile_line = txtfile_line.replace("）", "")
        txtfile_line = txtfile_line.replace("(", "")
        txtfile_line = txtfile_line.replace(")", "")
        txtfile_line = txtfile_line.replace("忘记", "")
        txtfile_line = txtfile_line.replace("打卡", "")
        txtfile_line = txtfile_line.replace("迟到", "")
        txtfile_line = txtfile_line.replace("缺卡", "")
        txtfile_line = txtfile_line.replace("班车", "")
        txtfile_line = txtfile_line.replace("补卡", "")
        if massstring != txtfile_line:
            print(massstring,' --> ',txtfile_line)
            logger.info(massstring + ' --> ' + txtfile_line)
        return (txtfile_line)


    def get_buqian_data(self, xlsfilename):

        self.scr.delete(1.0,END)
        temp_curr_datetime = datetime.datetime.now()
        self.scr.insert(END, "数据处理时间" + str(temp_curr_datetime) + "...\n\n")

        
        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return(None)

        int_first_row = 2
        # 日数据开始位置
        print(xlsfilename)
        self.scr.insert(END, "补签数据" + xlsfilename + "...\n")

        str_kaoqin_date_begin = self.svar_kehumingcheng2.get()
        str_kaoqin_date_end = self.svar_kehumingcheng.get()

        self.scr.insert(END,'处理考勤时间段 开始:'+str_kaoqin_date_begin+'')
        #try:
        workbook = load_workbook(xlsfilename)
        sheet_curr = workbook.worksheets[0]  # 获取excel里的工作表sheet名称数组
        # except Exception as err_message:
        #     print('ERR:',err_message)
        #     err_message_str = err_message.__str__()
        #     logger.error(err_message_str)
        #     logger.exception(sys.exc_info())

        int_sheet_nrows = sheet_curr.max_row
        int_sheet_ncols = sheet_curr.max_column        

        logger.info("sheet size:")
        logger.info(int_sheet_nrows)
        logger.info(int_sheet_ncols)

        #print("sheetname & lines:", str_curr_sheet_name, int_sheet_nrows)

        int_first_col = 6

        self.curr_year_month = str_kaoqin_date_begin[:8]
        # self.scr.insert(END, "\n\n\n注意： " + str(self.curr_year_month) + "...\n\n\n")
        # self.master.update()

        userid_count_attend = 0
        userid_count = 0
        userid_attend = 0
        sheet_title = sheet_curr.cell(1, 1).value
        self.scr.insert(
            END, "\nExcel File Title: " + sheet_title + "...\n"
        )
        sheet_title_lastday = sheet_title[-10:]

        buqian_list = []
        for i in range(int_first_row, int_sheet_nrows):
            # print('i: ',i)
            xuhao = sheet_curr.cell(i, 3).value
            username = sheet_curr.cell(i, 3).value
            logger.info(xuhao)
            logger.info(username)
            userid = sheet_curr.cell(i, 4).value
            # 匹配工号
            buqian_onerow = []
            if userid:
                for j in range(1,11):
                    buqian_onerow.append(sheet_curr.cell(i, j).value)
                buqian_list.append(buqian_onerow)
                self.scr.insert(END, xuhao + " " + username +"    \n")
                self.master.update()
            else:
                break
        logger.info('补签表：')
        logger.info(buqian_list)

        workbook.close()
        return(buqian_list)




    def buqian_data_ech(self, xlsfilename):

        self.scr.delete(1.0,END)
        temp_curr_datetime = datetime.datetime.now()
        self.scr.insert(END, "数据处理时间" + str(temp_curr_datetime) + "...\n\n")

        
        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message = messagebox.askquestion(
                title="提示", message="无找到文件: " + xlsfilename + "，继续？"
            )  # return yes/no
            return return_message
        except_attend_number = ["0022"]

        int_first_row = 2
        # 日数据开始位置
        print(xlsfilename)
        self.scr.insert(END, "补签数据" + xlsfilename + "...\n")

        str_kaoqin_date_begin = self.svar_kehumingcheng2.get()
        str_kaoqin_date_end = self.svar_kehumingcheng.get()

        self.scr.insert(END,'处理考勤时间段 开始:'+str_kaoqin_date_begin+'')
        #try:
        workbook = load_workbook(xlsfilename)
        sheet_curr = workbook.worksheets[0]  # 获取excel里的工作表sheet名称数组
        # except Exception as err_message:
        #     print('ERR:',err_message)
        #     err_message_str = err_message.__str__()
        #     logger.error(err_message_str)
        #     logger.exception(sys.exc_info())

        int_sheet_nrows = sheet_curr.max_row
        int_sheet_ncols = sheet_curr.max_column        

        logger.info("sheet size:")
        logger.info(int_sheet_nrows)
        logger.info(int_sheet_ncols)

        #print("sheetname & lines:", str_curr_sheet_name, int_sheet_nrows)

        int_first_col = 6

        self.curr_year_month = str_kaoqin_date_begin[:8]
        # self.scr.insert(END, "\n\n\n注意： " + str(self.curr_year_month) + "...\n\n\n")
        # self.master.update()

        userid_count_attend = 0
        userid_count = 0
        userid_attend = 0
        sheet_title = sheet_curr.cell(1, 1).value
        self.scr.insert(
            END, "\nExcel File Title: " + sheet_title + "...\n"
        )
        sheet_title_lastday = sheet_title[-10:]

        logger.info('xuhao')
        logger.info('username')
        buqian_list = []
        for i in range(int_first_row, int_sheet_nrows):
            # print('i: ',i)
            xuhao = sheet_curr.cell(i, 3).value
            username = sheet_curr.cell(i, 3).value
            logger.info(xuhao)
            logger.info(username)
            userid = sheet_curr.cell(i, 4).value
            # 匹配工号
            buqian_onerow = []
            if userid:
                for j in range(1,11):
                    buqian_onerow.append(sheet_curr.cell(i, j).value)
                buqian_list.append(buqian_onerow)
                self.scr.insert(END, xuhao + " " + username +"    \n")
                self.master.update()
            else:
                break
            logger.info(buqian_list)

            workbook.close()

            for buqian_onerow in buqian_list:
                workbook = load_workbook('tess.xlsx')
                sheet_finduserid = workbook.worksheets[0]  # 获取excel里的工作表sheet名称数组

                first_member_row = 5
                member_step = 14
                catch_userid = False
                while(True):
                    getuserid = sheet_finduserid.cell(first_member_row,2).value
                    if getuserid == buqian_onerow[2]:
                        catch_userid = True
                        break
                    first_member_row = first_member_row + member_step

                    if catch_userid : 
                        buka_date = buqian_onerow[4]    #日期
                        buka_banci = buqian_onerow[8]    #班次 白班晚班
                        buka_gongshi = buqian_onerow[9]    #班次 白班晚班
                        buka_date_day = buka_date.day
                        if buka_banci == '白班':
                            sheet_finduserid.cell(first_member_row,4+buka_date_day).value = buka_gongshi
                        elif buka_banci == '晚班':
                            sheet_finduserid.cell(first_member_row + 3 ,4+buka_date_day).value = buka_gongshi
                        else:
                            logger.info('找不到对应班次')
                            logger.info(buka_banci)
                    else:
                        logger.info('补卡资料有工号，考勤表找不到工号：')
                        logger.info(buqian_onerow)


        self.scr.insert(END, "导出.." + str(i - int_first_row + 1) + "行数据..\n")
        self.scr.insert(END, "有工号 .." + str(userid_count) + "行数据..\n")
        self.scr.insert(END, "有工号正常打卡 .." + str(userid_count_attend) + "行数据..\n")

        self.scr.insert(END, "..\n")

    # 生产车间员工就餐时间加班表查找返回
    def load_lunch_dinner_proc(self, xlsfilename,staffname):
        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message = messagebox.askquestion(
                title="提示", message="无找到文件" + xlsfilename + "，继续？"
            )  # return yes/no
            return return_message
        int_first_row = 5
        # 日数据开始位置
        #print(xlsfilename)
        #self.scr.insert(END, "钉钉数据" + xlsfilename + "...\n")
        #self.master.update()

        workbook = load_workbook(xlsfilename)  # 打开excel文件
        #worksheetj = workbook["员工信息表1"]  # 根据Sheet1这个sheet名字来获取该sheet
        worksheetj = workbook.worksheets[0]
        # worksheetj.cell(1, 1).value = str(xls_date.year)+'年广发银行'+str(xls_date.month)+' 月物料收发进销存日报表'

        i = int_first_row
        member_step = 14

        logger.info('查找姓名:'+staffname)
        
        while True:
            if not worksheetj.cell(i, 1).value:
                break
            sheet_staffname = worksheetj.cell(i, 3).value
            logger.info('表格姓名:'+sheet_staffname)
            lunch_supper_list = []
            if staffname in sheet_staffname:
                temp_list = []
                lunch_supper_list = []
                for j in range(5,35):
                    temp_list.append ( worksheetj.cell(i+1, j).value)
                lunch_supper_list.append(temp_list)
                temp_list = []
                for j in range(5,35):
                    temp_list.append ( worksheetj.cell(i+2, j).value)
                lunch_supper_list.append(temp_list)
                temp_list = []
                for j in range(5,35):
                    temp_list.append ( worksheetj.cell(i+4, j).value)
                lunch_supper_list.append(temp_list)
                temp_list = []
                for j in range(5,35):
                    temp_list.append ( worksheetj.cell(i+5, j).value)
                lunch_supper_list.append(temp_list)
                break                
            i = i + member_step
        print("=" * 40)
        self.scr.insert(END, "查找" + str(staffname)+ " 找到.."+ str(staffname)+"\n")
        self.master.update()
        workbook.close()
        return (lunch_supper_list)

    def bool_need_deduction_2hour(self,checkin_time):
        #是否需要扣减2小时
        
        int_weekday_click = checkin_time.weekday()
        check_work_day = checkin_time.day

        if int_weekday_click < 5:
            b_need_deduction = True
        else:
            b_need_deduction = False
        if check_work_day in self.is_workday:
                b_need_deduction = not b_need_deduction

        return(b_need_deduction)

    #工时计算-白班
    def workhours(self,fistclick,lastclick):
        curr_day = 10
        firstclicktime = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day),int(fistclick[:2]),int(fistclick[-2:]),0,0)

        lastclicktime = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day),int(lastclick[:2]),int(lastclick[-2:]),0,0,)

        workdaybefore = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day),6,0,0,0,)
        workday1b = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day),8,20,0,0,)
        workday1f = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day),11,20,0,0,)
        workday3b = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day),12,5,0,0,)
        workday3f = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day),17,25,0,0,)
        workday5b = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day),18,10,0,0,)
        workday5f = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day),20,20,0,0,)

        workhosts_sum = 0
        if firstclicktime > workdaybefore and firstclicktime < workday1b :
            if lastclicktime > workday1f:
                workhosts_sum = (workday1f - firstclicktime).seconds / 60   #第一班工时 = 第一班下班 - 上班打卡
                if lastclicktime > workday3f:
                    workhosts_sum = workhosts_sum + (60*5 + 20)
                    if lastclicktime > workday5b:   #早班5上班后 算加班了多少分钟
                        workhosts_sum = workhosts_sum + ((lastclicktime - workday5b).seconds / 60)
                else:
                        workhosts_sum = workhosts_sum + ((lastclicktime - workday3b).seconds / 60)
            else:
                workhosts_sum = workhosts_sum + ((lastclicktime - workday1b).seconds / 60)
        else:
            if firstclicktime < workday1f:
                workhosts_sum = workhosts_sum + ((workday1f -firstclicktime).seconds / 60)
                if lastclicktime > workday3f:
                    workhosts_sum = workhosts_sum + (60*5 + 20)
                    if lastclicktime > workday5b:   #早班5上班后 算加班了多少分钟
                        workhosts_sum = workhosts_sum + ((lastclicktime - workday5b).seconds / 60)
                else:
                        workhosts_sum = workhosts_sum + ((lastclicktime - workday3b).seconds / 60)
            else:
                if firstclicktime < workday3f:
                    workhosts_sum = workhosts_sum + ((workday3f -firstclicktime).seconds / 60)
                    if lastclicktime > workday5b:   #早班5上班后 算加班了多少分钟
                        workhosts_sum = workhosts_sum + ((lastclicktime - workday5b).seconds / 60)
                else:
                    if firstclicktime < workday5f:
                        workhosts_sum = workhosts_sum + ((workday5f -firstclicktime).seconds / 60)
                        if lastclicktime > workday5b:   #早班5上班后 算加班了多少分钟
                            workhosts_sum = workhosts_sum + ((lastclicktime - workday5b).seconds / 60)
        return(workhosts_sum)

    #工时计算-晚班
    def workhours_night(self,fistclick,lastclick):
        curr_day = 10   #随机选一天作为基准，用于计算工时
        #firstclicktime = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        #int(curr_day),int(fistclick[:2]),int(fistclick[-2:]),0,0)
        if fistclick > '17:00':
            firstclicktime = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
                int(curr_day),int(fistclick[:2]),int(fistclick[-2:]),0,0,)
            #计算当日、次日
        else:
            firstclicktime = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
                int(curr_day)+1,int(fistclick[:2]),int(fistclick[-2:]),0,0,)

        if lastclick > '18:30':
            lastclicktime = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
                int(curr_day),int(lastclick[:2]),int(lastclick[-2:]),0,0,)
        else:
            lastclicktime = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
                int(curr_day)+1,int(lastclick[:2]),int(lastclick[-2:]),0,0,)

        worknightbefore = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day),18,0,0,0,)
        worknight1b = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day),20,20,0,0,)
        worknight1f = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day),23,50,0,0,)
        worknight3b = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day)+1,0,35,0,0,)
        worknight3f = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day)+1,3,30,0,0,)
        worknight5b = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day)+1,4,0,0,0,)
        worknight5f = datetime.datetime(int(self.curr_year_month[:4]),int(self.curr_year_month[5:7]),
        int(curr_day)+1,8,20,0,0,)
        workhosts_sum = 0
        if firstclicktime > worknightbefore and firstclicktime < worknight1b :
            if lastclicktime > worknight1f:
                workhosts_sum = (worknight1f - firstclicktime).seconds / 60   #晚班第一班工时 = 第一班下班 - 上班打卡
                if lastclicktime > worknight3f:
                    workhosts_sum = workhosts_sum + (60*2 + 55)
                    if lastclicktime > worknight5b:   #早班5上班后 算加班了多少分钟
                        workhosts_sum = workhosts_sum + ((lastclicktime - worknight5b).seconds / 60)
                else:
                        workhosts_sum = workhosts_sum + ((lastclicktime - worknight3b).seconds / 60)
            else:
                workhosts_sum = workhosts_sum + ((lastclicktime - worknight1b).seconds / 60)
        else:
            if firstclicktime < worknight1f:
                workhosts_sum = workhosts_sum + ((worknight1f -firstclicktime).seconds / 60)
                if lastclicktime > worknight3f:
                    workhosts_sum = workhosts_sum + (60*2 + 55)
                    if lastclicktime > worknight5b:   #早班5上班后 算加班了多少分钟
                        workhosts_sum = workhosts_sum + ((lastclicktime - worknight5b).seconds / 60)
                else:
                        workhosts_sum = workhosts_sum + ((lastclicktime - worknight3b).seconds / 60)
            else:
                if firstclicktime < worknight3f:
                    workhosts_sum = workhosts_sum + ((worknight3f -firstclicktime).seconds / 60)
                    if lastclicktime > worknight5b:   #早班5上班后 算加班了多少分钟
                        workhosts_sum = workhosts_sum + ((lastclicktime - worknight5b).seconds / 60)
                else:
                    if firstclicktime < worknight5f:
                        workhosts_sum = workhosts_sum + ((worknight5f -firstclicktime).seconds / 60)
                        if lastclicktime > worknight5b:   #早班5上班后 算加班了多少分钟
                            workhosts_sum = workhosts_sum + ((lastclicktime - worknight5b).seconds / 60)
        return(workhosts_sum)

    
    # 生产车间计算工时 处理
    def proc_shengchan_proc(self, xlsfilename):
        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message = messagebox.askquestion(
                title="提示", message="无找到文件：" + xlsfilename + "，继续？"
            )  # return yes/no
            return return_message

        int_first_row = 4
        # 日数据开始位置
        print(xlsfilename)
        self.scr.insert(END, "钉钉数据" + xlsfilename + "...\n")

        self.scr.insert(
            END, "Getatime " + str(time.ctime(os.path.getatime(xlsfilename))) + "...\n")
        self.scr.insert(
            END, "Getmtime " + str(time.ctime(os.path.getmtime(xlsfilename))) + "...\n")
        self.scr.insert(
            END, "Getctime " + str(time.ctime(os.path.getctime(xlsfilename))) + "...\n")
        self.master.update()

        #################################################计算工时表格       
        worktime_wb = load_workbook('test.xlsx')
        worktime_ws1 = worktime_wb.worksheets[0]
        #计算工时表格
        one_member_rows = 14
        member_num = 0
        
        workbook = load_workbook(xlsfilename)
        # 钉钉打卡数据，只读取Excel文件第一个Sheet
        sheet_curr = workbook.worksheets[0]

        int_sheet_nrows = sheet_curr.max_row
        int_sheet_ncols = sheet_curr.max_column

        logger.info("sheet size:")
        logger.info(int_sheet_nrows)
        logger.info(int_sheet_ncols)

        jie_jia_ri = []
        for j in range(7,int_sheet_ncols):
            tempvalue = sheet_curr.cell(3, j).value
            if not tempvalue.isdigit():
                if len(str(tempvalue)) > 1:
                    logger.info('jie_jia_ri：')
                    logger.info(tempvalue)
                    jie_jia_ri.append(j-6)
                    logger.info(jie_jia_ri)
                
        int_first_col = 6

        kaoqinbiao_year_month_value = sheet_curr.cell(1, 1).value
        kaoqinbiao_year_month_value = kaoqinbiao_year_month_value[-10:]
        print('kaoqinbiao_year_month_value',kaoqinbiao_year_month_value)
        self.curr_month = kaoqinbiao_year_month_value[5:-3]
        self.curr_year_month = kaoqinbiao_year_month_value[:-2]

        self.scr.insert(END, "\n\n\n注意： " + str(self.curr_year_month) + "...\n\n\n")
        self.master.update()

        userid_count_attend = 0
        userid_count = 0
        userid_attend = 0

        self.scr.insert(
            END, "\nExcel File Title: " + sheet_curr.cell(1, 1).value + "...\n")

        buqian_list = self.get_buqian_data('补卡统计表.xlsx')

        jie_jia_ri_biaoti = []
        for i in range(int_first_col + 1, int_sheet_ncols + 1):
            cell_curr_value = sheet_curr.cell(int_first_row -1, i).value
            jie_jia_ri_biaoti.append(cell_curr_value)
            #标题行，用于查找法定节假日

        for i in range(int_first_row, int_sheet_nrows + 1):
            cell_curr_value = sheet_curr.cell(i, 1).value
            print("i: ", i)
            username = sheet_curr.cell(i, 1).value
            logger.info(username)

            userid = sheet_curr.cell(i, 4).value
            print("username id", username, userid)

            # 匹配工号
            work_time_minute_int = 0
            work_time_str = ""
            exist_attent = False
            for j in range(int_first_col + 1, int_sheet_ncols + 1):
                cell_value = sheet_curr.cell(i, j).value
                if cell_value != None:
                    exist_attent = True

            if exist_attent:
                #计算工时表
                member_num = member_num + 1
                paste_start_row = one_member_rows*(member_num-1)    #在当前位置新增表格
                if member_num > 1:
                    add_member_patten(worktime_ws1,paste_start_row,5,19) #5-19行
                paste_start_row = paste_start_row + 5
                #logger.info('member_num')
                #logger.info(member_num)
                logger.info('paste_start_row')
                logger.info(paste_start_row)

                worktime_ws1.cell(paste_start_row,1).value = member_num #序号
                cell_value = str(sheet_curr.cell(i, 1).value)
                staffname = cell_value
                worktime_ws1.cell(paste_start_row,3).value = cell_value #姓名
                cell_value = str(sheet_curr.cell(i, 4).value)
                worktime_ws1.cell(paste_start_row,2).value = cell_value #工号

                #加载用餐时间补
                lunch_supper_xlsfile = '员工连续不停机统计表.xlsx'
                lunch_supper_time = self.load_lunch_dinner_proc(lunch_supper_xlsfile,staffname)
                logger.info(lunch_supper_time)
                if len(lunch_supper_time ) < 1:
                    logger.info('没有找到员工用餐时间补')
                    logger.info(staffname)
                else:
                    for lunch_supper_day in range(0,30):
                        worktime_ws1.cell(paste_start_row + 1 ,lunch_supper_day+5).value = lunch_supper_time[0][lunch_supper_day]
                        worktime_ws1.cell(paste_start_row + 2 ,lunch_supper_day+5).value = lunch_supper_time[1][lunch_supper_day]
                        #空一行
                        worktime_ws1.cell(paste_start_row + 4 ,lunch_supper_day+5).value = lunch_supper_time[2][lunch_supper_day]
                        worktime_ws1.cell(paste_start_row + 5 ,lunch_supper_day+5).value = lunch_supper_time[3][lunch_supper_day]


                userid_count = userid_count + 1
                userid_attend = 0
                daka_line = ""
                on_duty_night = 0
                can_bu_days = 0
                onefive_times = 0
                two_times =0
                three_times = 0
                for j in range(int_first_col, int_sheet_ncols + 1):
                    # 插入数据
                    cell_value = str(sheet_curr.cell(i, j).value)
                    # print(cell_value)
                    if cell_value == None:
                        continue
                    cell_value = cell_value.replace("\n", "*")
                    cell_value_cut = cell_value.replace("\n", "*")
                    cell_value = cell_value_cut.replace(" ", "")
                    # logger.info(cell_value)
                    if len(cell_value) > 0:
                        daka_line = daka_line + str(cell_value) + "@" + str(j - 5) + "#"
                        click_one_times = cell_value.split("*")
                        #click_one_times.sort() #20200615 不能排序，凌晨下班打卡是次日，排序则变成当天
                        print(click_one_times)
                        # 打卡时间列表排序
                        if len(click_one_times) < 2:
                            print("一天打卡小于2次，工号： ", userid, "col: ",j,"记录：", click_one_times)
                            continue
                        for click_one_time in click_one_times:
                            if "外勤" in click_one_time:
                                print("含外勤打卡： ", click_one_times)
                                continue
                        if (j - int_first_col) > 9:
                            curr_day = str(j - int_first_col)
                        else:
                            curr_day = "0" + str(j - int_first_col)

                        first_click = click_one_times[0]
                        last_click = click_one_times[-1]
                        #上班打卡时间 5：00 -- 18：00 算白班
                        if first_click < '18:00' and first_click > '05:00':
                            gongshijisuan_jieguo = self.workhours(first_click,last_click)
                            logger.info('day-first_click,last_click:' + str(first_click) +'-'+ str(last_click))
                            logger.info(gongshijisuan_jieguo)
                            #写工时到工时表
                            worktime_ws1.cell(paste_start_row ,j-2).value = round(gongshijisuan_jieguo / 60, 2)
                            logger.info('write to worktimesheet写入工时表')
                            logger.info(str(paste_start_row) + ':' +str(j-2))
                        #上班打卡时间 5：00 -- 18：00 之外算夜晚班
                        else:
                            on_duty_night = on_duty_night +1
                            gongshijisuan_jieguo = self.workhours_night(first_click,last_click)
                            logger.info('night-first_click,last_click:' + str(first_click) +'-'+ str(last_click))
                            logger.info(gongshijisuan_jieguo)
                            #写工时到工时表
                            worktime_ws1.cell(paste_start_row+3 ,j-2).value = round(gongshijisuan_jieguo / 60, 2)
                            logger.info('write to worktimesheet写入工时表')
                            logger.info(str(paste_start_row+3) + ':' +str(j-2))
                        #计算餐补，工作时间3.5小时以上
                        if gongshijisuan_jieguo /60 > 3.5:
                            can_bu_days = can_bu_days +1

                        userid_attend = 1


                for find_buka_one in buqian_list:
                    find_userid_one = find_buka_one[3]
                    if find_userid_one == str(userid):
                        logger.info('开始补签 '+str(find_userid_one))
                        find_buka_date = find_buka_one[4]   #日期
                        find_buka_banci = find_buka_one[8]  #班次 白班晚班
                        find_buka_gongshi = find_buka_one[9]  #工时

                        find_buka_date_day = find_buka_date.day
                        if find_buka_banci == '早班':
                            worktime_ws1.cell(paste_start_row,4+find_buka_date_day).value = find_buka_gongshi
                        elif find_buka_banci == '晚班':
                            worktime_ws1.cell(paste_start_row + 3 ,4+find_buka_date_day).value = find_buka_gongshi
                        else:
                            logger.info('error找不到对应班次')
                            logger.info(find_buka_one)

                if int(self.curr_month)%2 ==0:     #单双月
                    logger.info('双月')
                    ot_days = 0 #加班天数
                    for j in range(34 ,4 ,-1): #AI=34, E=4
                        sum_gongshi_oneday =0 
                        for k in range(0,5):
                            temp_workt = worktime_ws1.cell(paste_start_row+k ,j).value
                            if temp_workt != None:
                                sum_gongshi_oneday = sum_gongshi_oneday + temp_workt
                        #worktime_ws1.cell(paste_start_row+9 ,j).value = sum_gongshi_oneday
                        #每天工时合计（临时显示）

                        if sum_gongshi_oneday >= 8:
                            worktime_ws1.cell(paste_start_row+10 ,j).value = 1
                            ot_days = ot_days +1
                            sum_gongshi_oneday = sum_gongshi_oneday - (sum_gongshi_oneday % 0.5)
                            if ot_days > 22:
                                if (j -4) in jie_jia_ri:    #-4,日期对应位置左移 4 位
                                    three_times = three_times + sum_gongshi_oneday
                                    worktime_ws1.cell(paste_start_row+13 ,j).value = sum_gongshi_oneday
                                    #三倍工资计算
                                else:                        
                                    #非节假日计算
                                    worktime_ws1.cell(paste_start_row+12 ,j).value = sum_gongshi_oneday
                                    two_times = two_times + sum_gongshi_oneday
                            else:
                                if (j -4) in jie_jia_ri:
                                    three_times = three_times + sum_gongshi_oneday
                                    worktime_ws1.cell(paste_start_row+13 ,j).value = sum_gongshi_oneday
                                    #三倍工资计算
                                else:                        
                                    #非节假日计算
                                    worktime_ws1.cell(paste_start_row+11 ,j).value = sum_gongshi_oneday - 8
                                    onefive_times = onefive_times + sum_gongshi_oneday -8
                        else:
                            worktime_ws1.cell(paste_start_row+10 ,j).value = round(sum_gongshi_oneday/8,2)
                            ot_days = ot_days + round(sum_gongshi_oneday/8,2)
                            if ot_days > 22:
                                worktime_ws1.cell(paste_start_row+12 ,j).value = sum_gongshi_oneday
                                two_times = two_times + sum_gongshi_oneday

                else:
                    #单月计算加班
                    logger.info('单月')
                    ot_days = 0 #加班天数
                    for j in range(4 ,34): #AI=34, E=4
                        sum_gongshi_oneday =0 
                        for k in range(0,5):
                            temp_workt = worktime_ws1.cell(paste_start_row+k ,j).value
                            if temp_workt != None:
                                sum_gongshi_oneday = sum_gongshi_oneday + temp_workt
                        worktime_ws1.cell(paste_start_row+9 ,j).value = sum_gongshi_oneday
                        #每天工时合计（临时显示）

                        if sum_gongshi_oneday >= 8:
                            worktime_ws1.cell(paste_start_row+10 ,j).value = 1
                            ot_days = ot_days +1
                            sum_gongshi_oneday = sum_gongshi_oneday - (sum_gongshi_oneday % 0.5)
                            if ot_days > 22:
                                if (j -4) in jie_jia_ri:    #-4,日期对应位置左移 4 位
                                    three_times = three_times + sum_gongshi_oneday
                                    worktime_ws1.cell(paste_start_row+13 ,j).value = sum_gongshi_oneday
                                    #三倍工资计算
                                else:                        
                                    #非节假日计算
                                    worktime_ws1.cell(paste_start_row+12 ,j).value = sum_gongshi_oneday
                                    two_times = two_times + sum_gongshi_oneday
                            else:
                                if (j -4) in jie_jia_ri:
                                    three_times = three_times + sum_gongshi_oneday
                                    worktime_ws1.cell(paste_start_row+13 ,j).value = sum_gongshi_oneday
                                    #三倍工资计算
                                else:                        
                                    #非节假日计算
                                    worktime_ws1.cell(paste_start_row+11 ,j).value = sum_gongshi_oneday - 8
                                    onefive_times = onefive_times + sum_gongshi_oneday -8
                        else:
                            worktime_ws1.cell(paste_start_row+10 ,j).value = round(sum_gongshi_oneday/8,2)
                            ot_days = ot_days + round(sum_gongshi_oneday/8,2)
                            if ot_days > 22:
                                worktime_ws1.cell(paste_start_row+12 ,j).value = sum_gongshi_oneday
                                two_times = two_times + sum_gongshi_oneday

                #汇总
                yingchuqin_days_str = worktime_ws1.cell(2 ,27).value
                logger.info('yingchuqin_days_str')
                logger.info(yingchuqin_days_str)
                yingchuqin_days = yingchuqin_days_str[-2:]
                worktime_ws1.cell(paste_start_row+2 ,37).value = yingchuqin_days

                worktime_ws1.cell(paste_start_row+2 ,36).value = on_duty_night
                worktime_ws1.cell(paste_start_row+6 ,36).value = can_bu_days
                worktime_ws1.cell(paste_start_row+6 ,37).value = three_times
                worktime_ws1.cell(paste_start_row+10 ,36).value = ot_days
                worktime_ws1.cell(paste_start_row+10 ,37).value = '0' #onefive_times
                worktime_ws1.cell(paste_start_row+11 ,36).value = onefive_times
                worktime_ws1.cell(paste_start_row+12 ,36).value = two_times
                worktime_ws1.cell(paste_start_row+12 ,37).value = two_times + onefive_times + three_times

                print("工作时长 ", str(round(work_time_minute_int / 60, 2)))
                sheet_curr.cell(i, 38).value = round(work_time_minute_int / 60, 2)
                sheet_curr.cell(i, 40).value = work_time_str
                if userid_attend == 1:
                    userid_count_attend = userid_count_attend + 1
                logger.info(daka_line)

            else:
                logger.info('warning无打卡记录')

        #workbook.save("东信和平科技股份有限公司_打卡时间表-研发-工时.xlsx")
        print("=" * 40)
        print("共导入了 ", i - int_first_row + 1, "行数据.")
        self.scr.insert(END, "共导入了 .." + str(i - int_first_row + 1) + "行数据..\n")
        self.scr.insert(END, "有工号 .." + str(userid_count) + "行数据..\n")
        self.scr.insert(END, "有工号正常打卡 .." + str(userid_count_attend) + "行数据..\n")

        self.master.update()

        worktime_wb.save('tess.xlsx')

    # 

    # 程序主gui界面。
    def initWidgets(self, fm1):


        self.customer_sname = "ep"
        kehu_conf_jxc = "仓库进销存"
        self.Holiday = "节假日"
        self.file_from_cangkujxc = "仓库进销存"
        self.file_from_youjiqingdan = "邮寄清单"
        self.file_from_jichu = "基础数据文件"

        # print('host: ', str_kehu_name)
        # print(self.file_from_youjiqingdan)

        temp_curr_datetime = datetime.datetime.now()
        if temp_curr_datetime.day < 5:
            temp_curr_datetime = temp_curr_datetime - datetime.timedelta(days=6)
        str_kehu_name = temp_curr_datetime.strftime('%Y-%m-%d')
        str_kehu_name2 = (temp_curr_datetime- datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        str_kehu_name = '钉钉考勤表最新日期'

        # label_kehumingcheng = Label(fm1, text='客户名称：', font=('Arial', 12))
        # label_kehumingcheng.place(x=20, y=30)
        self.svar_kehumingcheng.set(str_kehu_name)
        #entry_kehumingcheng = Entry(fm1, textvariable=self.svar_kehumingcheng, width=20, font=('Arial', 12))
        #entry_kehumingcheng.place(x=620, y=125)

        self.svar_kehumingcheng2.set(str_kehu_name2)
        entry_kehumingcheng2 = Entry(fm1, textvariable=self.svar_kehumingcheng2, width=20, font=('Arial', 12))
        entry_kehumingcheng2.place(x=620, y=100)

        label_proc_time = Label(fm1, text='请输入处理的考勤开始日期，例子：2020-08-09', font=('Arial', 12))
        label_proc_time.place(x=620, y=70)

        temp_last_datetime = datetime.date.today() - datetime.timedelta(days=10)

        label_author = Label(fm1, text="by流程与信息化部IT. May,2020", font=("Arial", 9))
        label_author.place(x=820, y=770)

        self.scr = scrolledtext.ScrolledText(fm1, width=80, height=54)
        self.scr.place(x=20, y=30)

        # btn_id_import_init = Button(
        #     fm1, text="更新员工信息", command=self.command_id_import_run
        # )
        # btn_id_import_init.place(x=620, y=200)

        #btn_dingding_exchage_run = Button(fm1, text="补签工时导入", command=self.command_buqian_run)
        #btn_dingding_exchage_run.place(x=620, y=270)

        # btn_fix_rec_run = Button(
        #     fm1, text="补 签 卡", command=self.command_fix_recorder_run
        # )
        # btn_fix_rec_run.place(x=620, y=340)

        # btn_fix_userid_run = Button(fm1, text='匹配通信录工号', command=self.command_fix_txl_id_run)
        # btn_fix_userid_run.place(x=620, y=410)

        btn_fix_userid_run = Button(fm1, text="生产车间算工时", command=self.command_fix_yf_run)
        btn_fix_userid_run.place(x=620, y=410)

        # btn_fix_userid_run = Button(
        #     fm1, text="研发补卡>xls", command=self.command_fix_yf_buka_run
        # )
        # btn_fix_userid_run.place(x=620, y=510)

        btn_barcode_init = Button(fm1, text=" 退  出 ", command=self.command_btn_exit)
        btn_barcode_init.place(x=620, y=690)

    # 退出键
    def command_btn_exit(self):
        self.master.destroy()

    # 导入员工工号
    def command_id_import_run(self):
        # 功能停用
        # return(0)

        label_tips1_filename = Label(
            self.master, text="正在导入员工工号数据... ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

        userid_filename = "在职人员信息表.xls"
        self.user_id_import_list(userid_filename)

        label_tips1_filename = Label(
            self.master, text="完成...                     ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

    # 补签卡
    def command_fix_recorder_run(self):
        label_tips1_filename = Label(
            self.master, text="正在导入补签卡数据... ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

        work_dir = "补签卡\\"
        self.fix_recorder_proc(work_dir)

        label_tips1_filename = Label(
            self.master, text="完成...                     ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

    # 匹配员工工号
    def command_fix_yf_run(self):

        dingding_userid_filename = "东信和平科技股份有限公司_打卡时间表_20210401-20210430.xlsx"

        self.proc_shengchan_proc(dingding_userid_filename)

        label_tips1_filename = Label(
            self.master, text="完成...                     ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

    # 研发中心 补卡
    def command_fix_yf_buka_run(self):

        dingding_userid_filename = "东信和平科技股份有限公司_打卡时间表-yf.xlsx"

        work_dir = "补签卡\\"

        self.proc_yf_buka_proc(work_dir, dingding_userid_filename)

        label_tips1_filename = Label(
            self.master, text="完成...                     ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

    # 匹配员工工号
    def command_fix_txl_id_run(self):
        label_tips1_filename = Label(
            self.master, text="正在匹配员工工号数据... ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=530)

        dingding_userid_filename = "东信和平科技股份有限公司-通讯录.xlsx"

        self.download_txl_id_run_proc(dingding_userid_filename)

        label_tips1_filename = Label(
            self.master, text="完成...                     ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

    # 钉钉数据转NC 文本文件
    def command_buqian_run(self):

        label_tips1_filename = Label(
            self.master, text="补卡统计表... ", font=("Arial", 12))
        label_tips1_filename.place(x=620, y=590)

        file_from_dingding = "补卡统计表.xlsx"

        #try:
        self.buqian_data_ech(file_from_dingding)
        # except Exception as err_message:
        #     print(err_message)
        #     self.scr.insert(END, err_message)
        #     self.scr.update()
        #     logger.error(err_message.__str__())
        #     logger.exception(sys.exc_info())


        label_tips1_filename = Label(
            self.master, text="完成...                     ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

        return 0


if __name__ == "__main__":

    set_logging()

    main_window = Tk()
    main_window.title("临时考勤数据处理工具 v.20200615")

    # 设定窗口的大小(长 * 宽)，显示窗体居中，winfo_xxx获取系统屏幕分辨率。
    sw = main_window.winfo_screenwidth()
    sh = main_window.winfo_screenheight()
    ww = 1024
    wh = 800
    x = (sw - ww) / 2
    y = (sh - wh) / 2
    main_window.geometry("%dx%d+%d+%d" % (ww, wh, x, y))  # 这里的乘是小x
    display = App(main_window)
    main_window.mainloop()
