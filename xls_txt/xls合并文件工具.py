'''
版本: V1.01
功能：
1．  遍历文件夹中的所有excel文件（xls和xlsx）
2．  将各表中的数据汇总到一个文件中

配置文件.ini
[文件合并工具]
#pendingdir : 待合并excel文件所在目录,可以是绝对路径，如：D:\待合并，只含目录名表示在程序所在目录下。
pendingdir = 待合并
#savefilename: 合并后保存的文件名,可以是绝对路径，如：D:\已合并文件名.xlsx，只含文件名表示在程序所在目录下。
savefilename = 合并文件名.xlsx
'''

from tkinter import Tk
from configparser import ConfigParser
from tkinter import messagebox,scrolledtext,Canvas,PhotoImage,Label,StringVar,Entry, Button,END, DISABLED, Toplevel  # 导入滚动文本框的模块
from os.path import exists as os_path_exists
import xlrd
import datetime
from openpyxl import load_workbook,Workbook
import os
import logging
from logging import getLogger
from logging.handlers import RotatingFileHandler
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font  #设置字体和边框需要的模块

#xl_border = Border(left=Side(style='thin',color='FF000000'),right=Side(style='thin',color='FF000000'),top=Side(style='thin',color='FF000000'),bottom=Side(style='thin',color='FF000000'),diagonal=Side(style='thin',color='FF000000'),diagonal_direction=0,outline=Side(style='thin',color='FF000000'),vertical=Side(style='thin',color='FF000000'),horizontal=Side(style='thin',color='FF000000'))
#font_set = Font(name='宋体', size=24, italic=True, color=colors.BLUE, bold=True, underline='doubleAccounting')
font_set = Font(name='宋体', size=10)

#设置日志文件配置参数
def set_logging():
    global logger
    logger = getLogger('balance_logger')
    handler = RotatingFileHandler('日志记录.log', maxBytes=5000000, backupCount=6)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)
    formatter = logging.Formatter('%(asctime)-12s %(filename)s %(lineno)d %(message)s')
    handler.setFormatter(formatter)

#定义类，脚本主要更能
class App():
    def __init__(self, master):

        self.master = master
        self.customer_sname = ''
        self.curr_month = ''
        self.autorun = ''
        self.filesymbol = ''
        self.pendingdir = ''
        self.savefilename = ''
        self.initWidgets(master)


    #按字符查找符合条件文件名，返回文件列表
    def find_filename(self, curr_path, curr_filename_path):
        list_files = []
        for parent, dirnames, filenames in os.walk(curr_path, followlinks=True):
            for filename in filenames:
                file_path = os.path.join(parent, filename)
                if curr_filename_path in filename:
                    print('文件名：%s' % file_path)
                    list_files.append(file_path)
        if len(list_files) > 0:
            return (list_files)
        else:
            return (None)


    #从数据库处理数据，导出对账文件excel
    def merge_xls_proc(self, work_dir, xlsfilename):
        self.scr.insert(END, "处理时间: "+datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')+ "\n")
        logger.info("处理时间: "+datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        self.scr.insert(END, "待合并文件所在文件夹: "+str(work_dir)+ "\ \n")
        file_lists = self.find_filename(work_dir,'xls')
        if file_lists == None:
            self.scr.insert(END, "待合并文件夹为空... "+ "\n")
            return ('nonefiletoproc')
        for one_file in file_lists:
            self.scr.insert(END, one_file +'\n')


        self.scr.insert(END, "======================================================= "+ "\n")
        self.master.update()


        xlsfilename = self.savefilename
        merage_workbook = Workbook()  #创建一个工作表
        #ws操作sheet页load_workbook(xlsfilename)  # 打开excel文件
        for_merage_worksheet = merage_workbook.active  # 根据Sheet1这个sheet名字来获取该sheet
        for_merage_worksheet.cell(1,1).value = "卡片序号"
        for_merage_worksheet.cell(1,1).font = font_set
        for_merage_worksheet.cell(1,2).value = "卡片号段号"
        for_merage_worksheet.cell(1,2).font = font_set
        for_merage_worksheet.cell(1,3).value = "卡片号段序号"
        for_merage_worksheet.cell(1,3).font = font_set
        for_merage_worksheet.cell(1,4).value = "排序状态"
        for_merage_worksheet.cell(1,4).font = font_set
        for_merage_worksheet.cell(1,5).value = "卡号"
        for_merage_worksheet.cell(1,5).font = font_set
        merage_sheet_last_row = 2

        for one_file in file_lists:
            #打开待合并xlsx文件
            row_proc_count = 0
            if '.xlsx' in one_file:
                xl=load_workbook(one_file)#打开excel
                #xl_sheet_names=xl.get_sheet_names()#获取所有sheet页名字
                #print(xl_sheet_names)#打印所有sheet页名称
                #xl_sheet=xl.get_sheet_by_name(xl_sheet_names[0])#定位到相应sheet页,[0]为sheet页索引
                xl_sheet=xl.worksheets[0]
                xl_sheet_max_row=xl_sheet.max_row    #获取行列数
                column=xl_sheet.max_column
                for xls_row in range(1+1,xl_sheet_max_row+1):  #五列：卡片序号	卡片号段号	卡片号段序号	排序状态	卡号
                    for xls_col in range(1, 5+1):
                        for_merage_worksheet.cell(merage_sheet_last_row,xls_col).value = xl_sheet.cell(row=xls_row,column=xls_col).value
                        for_merage_worksheet.cell(merage_sheet_last_row,xls_col).font = font_set
                        #获取单元格值
                    merage_sheet_last_row = merage_sheet_last_row +1
                    row_proc_count = row_proc_count +1
                logger.info('已处理文件：' + str(one_file) +' 含数据行数： ' + str(row_proc_count))
                self.scr.insert(END, '已处理文件：' + str(one_file) +' 含数据行数： ' + str(row_proc_count) + '\n')
            elif '.xls' in one_file:
                pendingfile = xlrd.open_workbook(one_file)
                table = pendingfile.sheets()[0]  ##通过索引顺序获取 不管 sheet 名称
                nrows = table.nrows  #获取该sheet中的有效行数
                for xls_row in range(1,nrows):  #五列：卡片序号	卡片号段号	卡片号段序号	排序状态	卡号
                    for xls_col in range(0, 5):
                        for_merage_worksheet.cell(merage_sheet_last_row,xls_col+1).value = table.cell_value(xls_row,xls_col)
                        for_merage_worksheet.cell(merage_sheet_last_row,xls_col+1).font = font_set
                        #获取单元格值
                    merage_sheet_last_row = merage_sheet_last_row +1
                    row_proc_count = row_proc_count +1
                logger.info('已处理文件：' + str(one_file) +' 含数据行数： ' + str(row_proc_count))
                self.scr.insert(END, '已处理文件：' + str(one_file) +' 含数据行数： ' + str(row_proc_count) + '\n')

            else:
                logger.info('导出 ~明细表~ 表' )
                self.scr.insert(END, "导出 ~明细表~ 表"+ "\n")
        try:
            merage_workbook.save(self.savefilename)  # 保存修改后的excel
            self.scr.insert(END, "\n数据合并文件已保存：" + str(self.savefilename) + "\n")
            logger.info("数据合并文件已保存：" + str(self.savefilename))
            self.scr.insert(END, "\n合并数据行数（不含标题）：" + str(merage_sheet_last_row -2) + "\n")
            logger.info("合并数据行数(不含标题)：" + str(merage_sheet_last_row -2))
            self.master.update()
        except Exception as err_message:
            print(err_message)
            self.scr.insert(END, err_message )
            self.scr.update()
            logger.error(err_message.__str__())
            logger.exception(sys.exc_info())

# 保存excel文件

# 程序主gui界面。
    def initWidgets(self, fm1):

        cp = ConfigParser()
        cp.read('配置文件.ini', encoding='gbk')
        try:
            self.autorun = '否'#cp.get('文件合并工具', 'autorun')
            self.filesymbol = '.xls'#cp.get('文件合并工具', 'filesymbol')
            self.pendingdir = cp.get('文件合并工具', 'pendingdir')
            self.savefilename = cp.get('文件合并工具', 'savefilename')
        except Exception as err_message:
            print(err_message)
            return_message = messagebox.showinfo(title='提示',message='无法打开配置文件.ini或配置有误!' )
            exit(2)

        #label_kehumingcheng = Label(fm1, text='客户名称：', font=('Arial', 12))
        #label_kehumingcheng.place(x=20, y=30)
        #self.svar_kehumingcheng.set(str_kehu_name)
        #entry_kehumingcheng = Entry(fm1, textvariable=self.svar_kehumingcheng, width=30, font=('Arial', 12))
        #entry_kehumingcheng.place(x=20, y=55)

        label_author = Label(fm1, text='by流程与信息化部IT. March,2020', font=('Arial', 9))
        label_author.place(x=500, y=777)

        self.scr = scrolledtext.ScrolledText(fm1, width=131, height=58)
        self.scr.place(x=10, y=10)

        btn_barcode_init = Button(fm1, text='文件合并', command=self.command_btn_run)
        btn_barcode_init.place(x=946, y=160)

        btn_barcode_init = Button(fm1, text=' 退  出 ', command=self.command_btn_exit)
        btn_barcode_init.place(x=946, y=270)

    # 退出键
    def command_btn_exit(self):
        self.master.destroy()

    # 主功能键
    def command_btn_run(self):
        self.scr.delete(1.0,END)
        self.merge_xls_proc(self.pendingdir,self.filesymbol)
        return 0

if __name__ == '__main__':

    set_logging()
    main_window = Tk()
    main_window.title('EXCEL文件合并小工具  v.20200319')

    # 设定窗口的大小(长 * 宽)，显示窗体居中，winfo_xxx获取系统屏幕分辨率。
    sw = main_window.winfo_screenwidth()
    sh = main_window.winfo_screenheight()
    ww = 1024
    wh = 800
    x = (sw - ww) / 2
    y = (sh - wh) / 2
    main_window.geometry("%dx%d+%d+%d" % (ww, wh, x, y))  # 这里的乘是小x
    display = App(main_window)
    main_window.mainloop()
