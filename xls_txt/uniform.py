#-*- coding:utf-8 -*-
#date: V2004231019
#auth: openjc
'''
[F配置]
Customers = 工行|交行|中行|浦发|建行
#|中行|建行|浦发

[工行]
TSHEETNAME = 工行
TTITLE = 1
T1 = SPACE
T2 = VAL|工商银行
T3 = 本方账号
T4 = 对方单位名称
T5 = 对方账号
T6 = VAL|国内
T7 = VAL|CNY
T8 = 贷方发生额
T9 = 交易时间
T10 = 摘要|用途|个性化信息
T11 = SPACE

[交行]
TSHEETNAME = 交行
TTITLE = 2
T1 = SPACE
T2 = VAL|交通银行
#收款账号不在表格内，在指定位置，如下格式，cell（1，2）
T3 = POS|1*2
T4 = 对方户名
T5 = 对方账号
T6 = VAL|国内
T7 = VAL|CNY
T8 = 贷方发生额
T9 = 交易时间
T10 = 摘要
T11 = SPACE

[中行]
TSHEETNAME = 中行
TTITLE = 1
T1 = SPACE
T2 = VAL|中国银行
#收款账号不在表格内，在指定位置，如下格式，cell（1，2）
T3 = 收款人账号[ Payee's Account Number ]
T4 = 付款人名称[ Payer's Name ]
T5 = 付款人账号[ Debit Account No. ]
T6 = VAL|国内
T7 = VAL|CNY
T8 = 交易金额[ Trade Amount ]
T9 = 交易日期[ Transaction Date ]
T10 = 用途[ Purpose ]
T11 = SPACE

[浦发]
TSHEETNAME = 浦发
TTITLE = 3
T1 = SPACE
T2 = VAL|浦发银行
#收款账号不在表格内，在指定位置，如下格式，cell（1，2）
T3 = POS|1*2
T4 = 对方户名
T5 = 对方账号
T6 = VAL|国内
T7 = VAL|CNY
T8 = 贷方金额
T9 = 交易日期
T10 = 摘要
T11 = SPACE

[建行]
TSHEETNAME = 建行
TTITLE = 1
T1 = SPACE
T2 = VAL|建设银行
T3 = POS|2*1
T4 = 对方户名
T5 = 对方账号
T6 = VAL|国内
T7 = VAL|CNY
T8 = 贷方发生额（收入）
T9 = 交易时间
T10 = 摘要|备注
T11 = SPACE

'''

from configparser import ConfigParser
from os.path import exists as os_path_exists
from datetime import datetime
from openpyxl import load_workbook
import os,sys
import logging
from logging.handlers import RotatingFileHandler
from openpyxl.styles import Border, Side, Alignment, PatternFill  #设置字体和边框需要的模块

from tkinter import Tk
from tkinter import MULTIPLE,Message,Listbox,messagebox,Label,StringVar,Scrollbar, Button,END, DISABLED, Toplevel,SUNKEN,LEFT,Y  # 导入滚动文本框的模块

xl_border = Border(left=Side(style='thin',color='FF000000'),right=Side(style='thin',color='FF000000'),top=Side(style='thin',color='FF000000'),bottom=Side(style='thin',color='FF000000'),diagonal=Side(style='thin',color='FF000000'),diagonal_direction=0,outline=Side(style='thin',color='FF000000'),vertical=Side(style='thin',color='FF000000'),horizontal=Side(style='thin',color='FF000000'))

dlevel = 1

#设置日志文件配置参数
def set_logging():
    global logger
    logger = logging.getLogger('balance_logger')
    handler = RotatingFileHandler('日志记录.log', maxBytes=5000000, backupCount=6)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)
    formatter = logging.Formatter('%(asctime)-12s %(filename)s %(lineno)d %(message)s')
    handler.setFormatter(formatter)

#定义类，脚本主要更能
class App():
    def __init__(self, master):

    # 脚本指定数据库名称sqlite3("db_dz.db3")
        self.master = master

        self.md5filename = 'filelist.md5'
        self.svar_tips = StringVar()
        self.svar_file_detail_tips = StringVar() 
        self.list_conf_customer_lists = []
        self.ftplocaldir = ''
        self.targetdir = ''
        self.label_tips = Label()
        self.list_message = Listbox()
        self.filesymbol = ''
        self.pendingdir = ''
        self.savefilename = ''
        self.btn_download_init = None #Button()
        self.file_md5_list = []
        self.file_detail_tips = []
        self.scr_history_have_clean = False
        self.customer_zone_list = []

        self.customer_name = ''
        self.kehu_pos_datail = []
        self.data_dir = ''
        self.file_from_cangkujxc = ''
        self.file_from_youjiqingdan = ''
        self.file_from_jichu = ''
        self.curr_month = ''
        self.initWidgets(master)

    #按字符查找符合条件文件名，返回文件列表
    def find_filename(self, curr_path, curr_filename_path):
        list_files = []
        for parent, dirnames, filenames in os.walk(curr_path, followlinks=True):
            for filename in filenames:
                file_path = os.path.join(parent, filename)
                if curr_filename_path in filename:
                    self.list_message.insert('文件名：%s' % file_path)
                    list_files.append(file_path)
        if len(list_files) > 0:
            return (list_files[0])
        else:
            return (None)

    # 从数据库导出价格（基础表），返回含价格信息列表

#从数据库处理数据，导出对账文件excel

    def excel_cell_rowcell_to_position(self,int_row,int_column):
        if int_row < 26:
            str_excel_cell_pos = chr(64+int_row)
            str_excel_cell_pos = str_excel_cell_pos + str(int_column)
        return str_excel_cell_pos
# 整合数据，导出生成excel文件
#从数据库处理数据，导出对账文件excel

    def excel_cell_rowcell_to_position(self,int_row,int_column):
        if int_row < 26:
            str_excel_cell_pos = chr(64+int_row)
            str_excel_cell_pos = str_excel_cell_pos + str(int_column)
        return str_excel_cell_pos

# 整合数据，导出生成excel文件



    def new_csvdata_list(self, customer, xlsfilename):
        #def date_re_format(soure_date):
        #    return (str(soure_date.year) + '/' + str(soure_date.month) + '/' + str(soure_date.day))

        self.list_message.insert(0,'导入文件数据： ' + xlsfilename)
        int_first_row = 3
        self.customer_zone_list  = []

        fu_kuan_ren_mingcheng_len_divion = 5    #付款人名称字数小于5个字 属于 系统事业产品
        date_format_export = '%Y-%m-%d'

        # 获取明细表数据
        xlsfilename = self.data_dir + xlsfilename
        #workbook = load_workbook(xlsfilename)  # 打开excel文件
        logger.info('导入 ~开票平台中客户及区域~ 表' )
        #worksheetj = workbook['开票平台中客户及区域']  # 根据Sheet1这个sheet名字来获取该sheet
        i = 0
        #max_rows = worksheetj.max_row
        
        workbook_source = load_workbook(xlsfilename)
        worksheet_source = workbook_source.worksheets[0]

        worksheet_source_maxrow = worksheet_source.max_row
        qu_yu_yu_fu_ze_ren_fen_pei_list = []
        for i in range(1,worksheet_source_maxrow +1):
            temp1 = worksheet_source.cell(i,1).value
            temp2 = worksheet_source.cell(i,2).value
            self.customer_zone_list.append([temp1,temp2])

        logger.info('import 开票平台中客户及区域' + str(len(self.customer_zone_list))+' 行')
        #logger.info(self.customer_zone_list )
        self.list_message.insert(0,'导入-开票平台中客户及区域-文件 （行）：' + str(len(self.customer_zone_list)))
        self.list_message.update()
        workbook_import = load_workbook('区域与负责人分配表.xlsx')  # 打开excel文件
        logger.info('导入 ~区域与负责人分配表~ 表' )
        self.list_message.insert(0,'导入 ~区域与负责人分配表~ 表' )
        self.list_message.update()
        worksheet_import = workbook_import['Sheet1']
        sheet_import_maxrow = worksheet_import.max_row
        qu_yu_yu_fu_ze_ren_fen_pei_list = []
        for i in range(1,sheet_import_maxrow +1):
            value1 = worksheet_import.cell(i,1).value
            value2 = worksheet_import.cell(i,2).value
            value3 = worksheet_import.cell(i,3).value
            qu_yu_yu_fu_ze_ren_fen_pei_list.append([value1,value2,value3])
        logger.info(qu_yu_yu_fu_ze_ren_fen_pei_list)
        self.list_message.insert(0,'导入-区域与负责人分配表-文件 （行）：' + str(len(qu_yu_yu_fu_ze_ren_fen_pei_list)))
        self.list_message.update()
        workbook_import.close()


        # 获取明细表数据
        #xlsfilename = self.data_dir + xlsfilename
        workbook_source = load_workbook('国内每日收款明细.xlsx')  # 打开excel文件
        logger.info('导入 ~国内每日收款明细~ 表' )
        self.list_message.insert(0,'导入 ~国内每日收款明细~ 表' )
        self.list_message.update()

        workbook_target = load_workbook('银行流水统一格式.xlsx')  # 打开excel文件
        logger.info('转换到 ~银行流水统一格式~ 表' )
        self.list_message.insert(0,'转换到 ~银行流水统一格式~ 表' )
        self.list_message.update()

        self.list_message.insert(0,'开始查找标题位置信息...')
        self.list_message.update()
        for idx_one_customer_conf in range(len(self.list_conf_customer_lists)):
            list_one_customer_conf = self.list_conf_customer_lists[idx_one_customer_conf]
            #fxl = form excel
            fxl_sheetname = list_one_customer_conf [0]
            fxl_title_row = int(list_one_customer_conf [1])
            self.list_message.insert(0,'开始查找标题位置信息...'+str(fxl_sheetname))
            self.list_message.update()

            try:        #检查是否有对应的 sheet
                worksheet_source = workbook_source[fxl_sheetname]
            except:
                self.list_message.insert(0,'='*40)
                self.list_message.insert(0,'注意 无法打开sheet：' + str(fxl_sheetname))
                self.list_conf_customer_lists[idx_one_customer_conf][1]='NOTFOUND'
                worksheet_target = workbook_target['其他流水']  # 根据Sheet1这个sheet名字来获取该sheet
                worksheet_target.cell(idx_one_customer_conf + 3 ,1).value = fxl_sheetname
                logger.info('注意 无法打开sheet：' + str(fxl_sheetname))
                logger.info(self.list_conf_customer_lists)
                continue

            fxl_tile_maxcol = worksheet_source.max_column

            #读 excel 表格 标题 行 内容
            list_excel_title = []
            for i in range(1,fxl_tile_maxcol+1):
                list_excel_title.append([i,worksheet_source.cell(fxl_title_row,i).value])
            logger.info(list_excel_title)

            for idx_one_cell_in_conf_line in range(2,12):               #T1 - T 12 
                one_cell_in_conf_line = list_one_customer_conf[idx_one_cell_in_conf_line]
                if one_cell_in_conf_line == 'SPACE':
                    continue
                if 'POS' in one_cell_in_conf_line:
                    continue
                if 'VAL' in one_cell_in_conf_line:
                    continue
                list_temp = one_cell_in_conf_line.split('|')
                if len(list_temp)<2:
                    one_cell_in_conf_line = list_temp[0]         #分拆多栏位字段
                    idx_temp = 9999
                    for temp_one_cell in list_excel_title:
                        if temp_one_cell[1] == None:
                            continue
                        temp_cell = temp_one_cell[1]
                        if one_cell_in_conf_line in temp_cell:
                            idx_temp = temp_one_cell[0]
                    if idx_temp == 9999:
                        logger.info('没有找到标题位置：' + fxl_sheetname+ '  '+one_cell_in_conf_line)
                        self.list_message.insert(0,'没有找到标题位置：' + fxl_sheetname+ '  '+ one_cell_in_conf_line)
                        self.list_message.update()
                    else:
                        self.list_conf_customer_lists[idx_one_customer_conf][idx_one_cell_in_conf_line] = one_cell_in_conf_line + '@' + str(idx_temp)
                else:   #多于2个字段
                    one_cell_restruck = ''
                    for one_cell_in_conf_line in list_temp:         #分拆多栏位字段
                        idx_temp = 9999
                        for temp_one_cell in list_excel_title:
                            if temp_one_cell[1] == None:
                                continue
                            temp_cell = temp_one_cell[1]
                            if one_cell_in_conf_line in temp_cell:
                                idx_temp = temp_one_cell[0]
                        if idx_temp == 9999:
                            logger.info('没有找到标题位置：' + fxl_sheetname+ '  '+one_cell_in_conf_line)
                            self.list_message.insert(0,'没有找到标题位置：' + fxl_sheetname+ '  '+ one_cell_in_conf_line)
                            self.list_message.update()
                        else:
                            if one_cell_restruck == '':
                                one_cell_restruck = one_cell_in_conf_line + '@' + str(idx_temp)
                            else:
                                one_cell_restruck = one_cell_restruck + '|' + one_cell_in_conf_line + '@' + str(idx_temp)
                    self.list_conf_customer_lists[idx_one_customer_conf][idx_one_cell_in_conf_line] = one_cell_restruck
        logger.info(self.list_conf_customer_lists)
        self.list_message.insert(0,'已确定标题位置信息。')

        int_from_conf_first_pos = 2
        int_from_conf_last_pos = 12
        curr_row_target=3
        guo_nei_first_row_target =2
        guo_nei_last_row_target = guo_nei_first_row_target
        xi_tong_shi_ye_first_row_target = 2
        xi_tong_shi_ye_last_row_target = xi_tong_shi_ye_first_row_target
        fxl_sheet_data_rows = 0
        #目标excel 表格 首行位置
        
        for idx_one_customer_conf in range(len(self.list_conf_customer_lists)):
            logger.info('begin load sheet data: ')
            logger.info(self.list_conf_customer_lists[idx_one_customer_conf])
            #当 未能打开 相应的 sheet ，跳过
            if 'NOTFOUND' == self.list_conf_customer_lists[idx_one_customer_conf][1]:
                continue
            
            last_guo_nei_last_row_target = guo_nei_last_row_target
            last_xi_tong_shi_ye_last_row_target = xi_tong_shi_ye_last_row_target
           
            list_one_customer_conf = self.list_conf_customer_lists[idx_one_customer_conf]
            #fxl = form excel
            fxl_sheetname = list_one_customer_conf [0]
            fxl_title_row = int(list_one_customer_conf [1])
            fxl_data_first_row = fxl_title_row +1
            #数据行首行
            worksheet_source = workbook_source[fxl_sheetname]
            fxl_tile_maxcol = worksheet_source.max_column

            #worksheet_source = workbook_source[sheet_name]  # 根据Sheet1这个sheet名字来获取该sheet

            sheet_source_maxrow = worksheet_source.max_row
            dict_strick_pos = {}

            self.list_message.insert(0,'数据处理： '+ fxl_sheetname)

            #从每一个 sheet 读取 数据第一行 到 行 尾数据
            logger.info('fxl_data_first_row: ' + str(fxl_data_first_row))
            logger.info('sheet_source_maxrow: ' +str(sheet_source_maxrow))

            last_fxl_sheet_data_rows = fxl_sheet_data_rows  #上次处理数据行数

            for int_row_data_from_source in range(fxl_data_first_row,sheet_source_maxrow+1):
                '''
                交行：贷方发生额>0
                工行：贷方发生额 空白的去除 & 摘要包含“利息划入”去除
                中行：交易类型筛选“来账”
                浦发：贷方发生额去掉空白
                建行：贷方发生额>0
                '''
                fxl_sheet_data_rows = fxl_sheet_data_rows +1                #累计处理数据行数

                jin_e_pos_str = list_one_customer_conf[9]   #金额
                jin_e_pos_list = jin_e_pos_str.split('@')
                jin_e_pos  = int(jin_e_pos_list[1])
                jin_e = str(worksheet_source.cell(int_row_data_from_source,jin_e_pos).value)
                jin_e = jin_e.strip()
                try:
                    jin_e_float = float(jin_e.replace(',',''))
                except:
                    jin_e_float = 0.0

                #if fxl_sheetname =='交行':     #全部均 需 贷方金额 > 0 
                if jin_e_float <= 0 :
                    continue

                if fxl_sheetname =='工行':
                    zhai_yao_pos_str = list_one_customer_conf[11]   #摘要
                    zhao_yao_pos_list = zhai_yao_pos_str.split('|')
                    zhai_yao_pos_str = zhao_yao_pos_list[0]
                    zhao_yao_pos_list = zhai_yao_pos_str.split('@')
                    zhao_yao_pos = int(zhao_yao_pos_list[1])
                    zhao_yao_str = str(worksheet_source.cell(int_row_data_from_source,zhao_yao_pos).value)
                    if '利息划入' in zhao_yao_str:              #排除 ’利息‘
                        logger.info(zhao_yao_str)
                        continue

                if fxl_sheetname =='中行':
                    #交易类型[ Transaction Type ]       特殊处理
                    jiao_yi_lei_xin_pos = 1
                    jiao_yi_lei_xin_str = str(worksheet_source.cell(int_row_data_from_source,jiao_yi_lei_xin_pos).value)
                    if not ('来账' in jiao_yi_lei_xin_str):              #
                        logger.info('中行排除项' + str(jiao_yi_lei_xin_str))
                        continue

                cheng_ban_ren_content = '' #承办人信息清空
                fu_kuan_ren_mingcheng_str = list_one_customer_conf[5]   #付款人名称
                fu_kuan_ren_pos_list = fu_kuan_ren_mingcheng_str.split('@')
                fu_kuan_ren_ren_pos  = int(fu_kuan_ren_pos_list[1])
                fu_kuan_ren_mingcheng = worksheet_source.cell(int_row_data_from_source,fu_kuan_ren_ren_pos).value

                if fu_kuan_ren_mingcheng != None:
                    fu_kuan_ren_mingcheng = fu_kuan_ren_mingcheng.strip()
                #查找付款人名称所属区域 begin
                ke_hu_suo_shu_qu_yu = '' #客户所属区域
                for kehu_quyu_index in range(0,len(self.customer_zone_list)):
                    kehu_quyu_search = self.customer_zone_list[kehu_quyu_index]
                    if fu_kuan_ren_mingcheng == kehu_quyu_search[0]:
                        ke_hu_suo_shu_qu_yu = kehu_quyu_search[1]
                        logger.info('客户所属区域, getit: ' + str(fu_kuan_ren_mingcheng) +';'+ str(ke_hu_suo_shu_qu_yu))
                        break
                #查找付款人名称所属区域 end
                #查找区域负责人 begin
                qu_yu_fu_ze_ren = ' '
                if ke_hu_suo_shu_qu_yu != '':
                    for kehu_quyu_index in range(0,len(qu_yu_yu_fu_ze_ren_fen_pei_list)):
                        kehu_quyu_search = qu_yu_yu_fu_ze_ren_fen_pei_list[kehu_quyu_index]
                        if ke_hu_suo_shu_qu_yu == kehu_quyu_search[1]:
                            qu_yu_fu_ze_ren = kehu_quyu_search[2]
                            logger.info('区域负责人, getit: ' + qu_yu_fu_ze_ren)
                            cheng_ban_ren_content = qu_yu_fu_ze_ren
                #查找区域负责人 end

                #切换 系统事业产品 / 国内卡产品 begin
                if fu_kuan_ren_mingcheng ==None:
                    sheet_name_switch = '国内卡产品'
                else:
                    fu_kuan_ren_mingcheng = fu_kuan_ren_mingcheng.strip()
                    if '药' in fu_kuan_ren_mingcheng:
                        sheet_name_switch = '系统事业产品'
                    elif len(fu_kuan_ren_mingcheng) > 0 and  len(fu_kuan_ren_mingcheng) < fu_kuan_ren_mingcheng_len_divion:
                        sheet_name_switch = '系统事业产品'
                    else:
                        sheet_name_switch = '国内卡产品'

                if sheet_name_switch == '系统事业产品':
                    worksheet_target = workbook_target['系统事业产品']  # 根据Sheet1这个sheet名字来获取该sheet
                    last_row_target = xi_tong_shi_ye_last_row_target
                    xi_tong_shi_ye_last_row_target = xi_tong_shi_ye_last_row_target +1
                    curr_row_target = xi_tong_shi_ye_last_row_target
                else:
                    worksheet_target = workbook_target['国内卡产品']  # 根据Sheet1这个sheet名字来获取该sheet
                    last_row_target = guo_nei_last_row_target
                    guo_nei_last_row_target = guo_nei_last_row_target +1
                    curr_row_target = guo_nei_last_row_target
                #切换 系统事业产品 / 国内卡产品 end
 
                #从 T1 到 T12 写入数
                logger.info(jin_e_float)
                for idx_one_cell_in_conf_line in range(int_from_conf_first_pos,int_from_conf_last_pos):               #T1 - T 12 

                    one_cell_in_conf_line = list_one_customer_conf[idx_one_cell_in_conf_line]
                    if one_cell_in_conf_line == 'SPACE':
                        continue
                    elif 'POS' in one_cell_in_conf_line:
                        list_temp = one_cell_in_conf_line.split('|')
                        cell_value = list_temp[1]
                        list_temp = cell_value.split('*')
                        temp_left = int(list_temp[0])
                        temp_right = int(list_temp[1])
                        str_key_word = fxl_sheetname + str(temp_left)+str(temp_right)
                        if str_key_word in dict_strick_pos:
                            cell_value = dict_strick_pos[str_key_word]
                        else:
                            cell_value = worksheet_source.cell(temp_left,temp_right).value
                            dict_strick_pos[str_key_word] = cell_value
                            logger.info(dict_strick_pos)

                        worksheet_target.cell(curr_row_target,idx_one_cell_in_conf_line-int_from_conf_first_pos+1).value = cell_value
                    
                    elif 'VAL' in one_cell_in_conf_line:
                        list_temp = one_cell_in_conf_line.split('|')
                        cell_value = list_temp[1]
                        worksheet_target.cell(curr_row_target,idx_one_cell_in_conf_line-int_from_conf_first_pos+1).value = cell_value
                    else:
                        list_temp = one_cell_in_conf_line.split('|')
                        #摘要项只有一个
                        if len(list_temp) < 2:
                            one_cell_str = list_temp[0]
                            list_temp2 = one_cell_str.split('@')
                            int_cell_position = int(list_temp2[1])
                            str_cell_name = list_temp2[0]

                            cell_value = worksheet_source.cell(int_row_data_from_source,int_cell_position).value
                            if str_cell_name == '交易时间':
                                try:
                                    shijian_datetime = datetime.strptime(cell_value, '%Y-%m-%d %H:%M:%S')
                                    cell_value = shijian_datetime.strftime('%Y-%m-%d')
                                except:
                                    try:
                                        shijian_datetime = datetime.strptime(cell_value, '%Y%m%d %H:%M:%S')
                                        cell_value = shijian_datetime.strftime('%Y-%m-%d')
                                    except:
                                        cell_value = '导入有误'
                            if '交易日期' in str_cell_name :
                                try:
                                    shijian_datetime = datetime.strptime(cell_value, '%Y%m%d')
                                    cell_value = shijian_datetime.strftime('%Y-%m-%d')
                                except:
                                    cell_value = '导入有误'


                            if  '发生额' in str_cell_name:
                                cell_value = jin_e_float
                            if  '金额' in str_cell_name:
                                cell_value = jin_e_float
                            #摘要（从一个字段中提取的）
                            if  '摘要' in str_cell_name:
                                cell_value = cell_value.strip()
                                cell_value = '[摘要]' + cell_value
                            if  '用途' in str_cell_name:
                                cell_value = cell_value.strip()
                                cell_value = '[用途]' + cell_value

                            worksheet_target.cell(curr_row_target,idx_one_cell_in_conf_line-int_from_conf_first_pos+1).value = cell_value
                        
                        else:  #摘要项二个以上
                            str_temp_str = ''
                            for list_temp_child in list_temp:
                                list_temp2 = list_temp_child.split('@')
                                int_cell_position = int(list_temp2[1])
                                str_cell_name = list_temp2[0]
                                cell_value = worksheet_source.cell(int_row_data_from_source,int_cell_position).value
                                if cell_value == None:
                                    cell_value = ' '
                                cell_value = cell_value.strip()
                                str_temp_str = str_temp_str + '[' +str(str_cell_name)+ ']' + str(cell_value) + ';'

                            worksheet_target.cell(curr_row_target,idx_one_cell_in_conf_line-int_from_conf_first_pos+1).value = str_temp_str
                    #承办人，特殊处理，位置列：11
                    worksheet_target.cell(curr_row_target,11).value = cheng_ban_ren_content

            self.list_message.insert(0,fxl_sheetname + '国内卡产品 ：'+str(guo_nei_last_row_target - last_guo_nei_last_row_target))
            self.list_message.insert(0,fxl_sheetname + '系统事业产品 ：'+str(xi_tong_shi_ye_last_row_target - last_xi_tong_shi_ye_last_row_target))
            self.list_message.insert(0,fxl_sheetname + '其他流水 ：'+str(fxl_sheet_data_rows - last_fxl_sheet_data_rows - (guo_nei_last_row_target - last_guo_nei_last_row_target) - (xi_tong_shi_ye_last_row_target - last_xi_tong_shi_ye_last_row_target)))
            self.list_message.insert(0,fxl_sheetname + '小计 ：'+str(fxl_sheet_data_rows -last_fxl_sheet_data_rows))
            self.list_message.insert(0,'='*40)
            self.list_message.update()

            print('guonei:' , guo_nei_last_row_target  - last_guo_nei_last_row_target)
            print('xi tong: ', xi_tong_shi_ye_last_row_target -last_xi_tong_shi_ye_last_row_target)
            print('fxl_sheet_data_rows: ', fxl_sheet_data_rows - last_fxl_sheet_data_rows)
                #curr_row_target = curr_row_target + 1
            worksheet_target = workbook_target['其他流水']  # 根据Sheet1这个sheet名字来获取该sheet
            
            worksheet_target.cell(idx_one_customer_conf + 3 ,1).value = fxl_sheetname

            worksheet_target.cell(idx_one_customer_conf + 3 ,2).value = (guo_nei_last_row_target - last_guo_nei_last_row_target) + (xi_tong_shi_ye_last_row_target - last_xi_tong_shi_ye_last_row_target)
            worksheet_target.cell(idx_one_customer_conf + 3 ,3).value = fxl_sheet_data_rows - last_fxl_sheet_data_rows - (guo_nei_last_row_target - last_guo_nei_last_row_target) - (xi_tong_shi_ye_last_row_target - last_xi_tong_shi_ye_last_row_target)
            worksheet_target.cell(idx_one_customer_conf + 3 ,4).value = fxl_sheet_data_rows - last_fxl_sheet_data_rows

        worksheet_target.cell(idx_one_customer_conf + 4 ,1).value = '合计'
        worksheet_target.cell(idx_one_customer_conf + 4 ,2).value = (guo_nei_last_row_target - guo_nei_first_row_target) + (xi_tong_shi_ye_last_row_target - xi_tong_shi_ye_first_row_target)
        worksheet_target.cell(idx_one_customer_conf + 4 ,3).value = fxl_sheet_data_rows - (guo_nei_last_row_target - guo_nei_first_row_target) - (xi_tong_shi_ye_last_row_target - xi_tong_shi_ye_first_row_target)
        worksheet_target.cell(idx_one_customer_conf + 4 ,4).value = fxl_sheet_data_rows

        temp_str = '银行流水统一格式(导出日期' + datetime.now().strftime('%Y-%m-%d %H%M')+').xlsx'
        temp_str = os.path.join(self.data_dir , temp_str)
        workbook_target.save(os.path.join(self.data_dir , temp_str))  # 保存修改后的excel
        self.list_message.insert(0,temp_str)

        self.list_message.insert(0,'国内卡产品 ：'+str(guo_nei_last_row_target - guo_nei_first_row_target))
        self.list_message.insert(0, '系统事业产品 ：'+str(xi_tong_shi_ye_last_row_target - xi_tong_shi_ye_first_row_target))
        self.list_message.insert(0, '其他流水 ：'+str(fxl_sheet_data_rows - (guo_nei_last_row_target - guo_nei_first_row_target) - (xi_tong_shi_ye_last_row_target - xi_tong_shi_ye_first_row_target)))
        self.list_message.insert(0, '合计 ：'+str(fxl_sheet_data_rows))
        self.list_message.insert(0,'处理完成。。。')


# 程序主gui界面。
    def initWidgets(self,fm1):
        try:
            cp = ConfigParser()
            cp.read('配置文件.ini', encoding='gbk')
            str_customers_name = cp.get('F配置', 'Customers')

            list_customers_name = str_customers_name.split('|')
            #TSKYH 收款银行
            conf_cells = ['TTITLE','T1','T2','T3','T4','T5','T6','T7','T8','T9','T10','T11']
            self.list_conf_customer_lists = []
            for str_customer_name in list_customers_name:
                one_list = [str_customer_name]
                for conf_cell in conf_cells:
                    conf_one = cp.get(str_customer_name, conf_cell)
                    one_list.append(conf_one)
                
                self.list_conf_customer_lists.append(one_list)
            logger.info('读取配置信息:')
            logger.info(self.list_conf_customer_lists)    
        except Exception as err_message:
            print(err_message)
            logger.error(err_message.__str__())
            logger.exception(sys.exc_info())
            return_message = messagebox.showinfo(title='提示',message='无法打开配置文件.ini或配置有误!' )
            return(2)

        label_author = Label(fm1, text='by流程与信息化部IT. April,2020', font=('Arial', 9))
        label_author.place(x=814, y=717)

        self.btn_download_init = Button(fm1, text='  运 行  ', command=self.command_refresh_btn_run)
        self.btn_download_init.place(x=929, y=170)

        #self.btn_sendfile_init = Button(fm1, text='发送文件', command=self.command_btn_run)
        #self.btn_sendfile_init.place(x=929, y=210)
        #btn_download_init.configure(state=DISABLED)

        btn_app_exit_init = Button(fm1, text='  退  出  ', command=self.command_btn_exit)
        btn_app_exit_init.place(x=929, y=270)

        self.sbar_lr = Scrollbar(fm1,width=20)
        self.list_message = Listbox(relief=SUNKEN,width =127,height=39,yscrollcommand=self.sbar_lr.set,font=('Arial', 10))
        #selectmode list多选模式multiple  selectmode = MULTIPLE,
        self.list_message.place(x=30, y=33)
        self.list_message.bind('<Double-Button-1>',self.click_left_printList) #双击 <Double-Button-1>
        
        self.sbar_lr.config(command=self.list_message.yview)                
        self.sbar_lr.pack(side=LEFT, fill=Y)                     
        self.sbar_lr.pack(padx=10,pady=40)

        str_tips = '刷新，请先点选要发送的文件       '
        str_tips = '      '
        self.label_tips = Label(textvariable=self.svar_tips, font=('Arial', 11))
        self.label_tips.place(x=30, y=7)
        self.svar_tips.set(str_tips)
        
        str_file_detail_tips = '双击, 查看文件大小和时间'
        str_file_detail_tips = ' '
        self.label_file_detail_tips = Label(textvariable=self.svar_file_detail_tips, font=('Arial', 10))
        self.label_file_detail_tips.place(x=30, y=704)
        self.svar_file_detail_tips.set(str_file_detail_tips)


        self.get_md5file()
        #读取MD5 文件

    # 退出键
    def command_btn_exit(self):
        self.master.destroy()
    def command_refresh_btn_run(self):
        self.new_csvdata_list(self.customer_name, '开票平台中客户及区域.xlsx')
    def command_refresh_md5_btn_run(self):
        pass
    def click_left_printList(self,event):
        pass
    def get_md5file(self):
        pass
    # 主功能键
    def command_btn_run(self):

        #if self.pricexls_db(self.customer_sname, self.file_from_jichu) == 'no':
         #   return (1)

        work_dir = '..\\仓库文件\\'

        try:
            self.csvdata_list(self.customer_name, '开票平台中客户及区域.csv')
        except Exception as err_message:
            print(err_message)
            logger.error(err_message.__str__())
            logger.exception(sys.exc_info())


        self.list_message.insert(0,'完成...                     ')
        return 0

if __name__ == '__main__':
    print('新OA银行统一流水格式化工具 V2004231019')

    set_logging()
    main_window = Tk()
    main_window.title('新OA银行统一流水格式化工具 V2004231019')

    # 设定窗口的大小(长 * 宽)，显示窗体居中，winfo_xxx获取系统屏幕分辨率。
    sw = main_window.winfo_screenwidth()
  
    sh = main_window.winfo_screenheight()
    ww = 1000
    wh = 740
    x = (sw - ww) / 2
    y = (sh - wh) / 2
    main_window.geometry("%dx%d+%d+%d" % (ww, wh, x, y))  # 这里的乘是小x
    logger.info('程序启动，program restart...')
    display = App(main_window)
    main_window.mainloop()
