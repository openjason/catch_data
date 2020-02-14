#!/bin/python3
'''
功能：广发制卡业务库存反馈文件自动生成工具。
'''
from tkinter import ttk
import tkinter
from configparser import ConfigParser
from tkinter import messagebox,scrolledtext,Canvas,PhotoImage,Label,StringVar,Entry, Button,END, DISABLED, Toplevel  # 导入滚动文本框的模块
from os.path import exists as os_path_exists
import xlrd
import datetime,time
import os,sys
import logging
from openpyxl import load_workbook
from logging.handlers import RotatingFileHandler
from calendar import monthrange

#设置日志文件配置参数
def set_logging():
    global logger
    logger = logging.getLogger('balance_logger')
    handler = RotatingFileHandler('日志记录.log', maxBytes=5000000, backupCount=6)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)
    formatter = logging.Formatter('%(asctime)-12s %(filename)s %(lineno)d %(message)s')
    handler.setFormatter(formatter)

#定义类，脚本主要更能
class App():
    def __init__(self, master):
        self.svar_proc_time1 = StringVar()
        self.svar_proc_time2 = StringVar()
        self.svar_wuliaoshiyong_filename = StringVar()
        self.svar_kehumingcheng = StringVar()
        self.svar_youjiqingdan_filename = StringVar()
        self.svar_nnnnnnnnnnnnnnn_filename = StringVar()
        self.svar_label_prompt = StringVar()
        self.master = master
        self.customer_sname = ''
        self.Holiday = []
        self.data_dir = ''
        self.file_from_wuliaoshiyong = '广发银行物料使用情况记录表'
        self.file_from_youjiqingdan = ''
        self.file_from_fuliaokucun = ''
        self.curr_month = ''
        self.initWidgets(master)
        self.work_dir = ''
        self.savefile_dir = ''
        #程序是修改的，有部分变量没有用上

    # 物料进销存日报表 处理
    def wuliaojxc_file_proc(self, export_xls_filename, xlsfilename):
        return_message = 'err'
        curr_proc_time_str = self.svar_proc_time1.get()
        try:
            date_p = datetime.datetime.strptime(curr_proc_time_str, '%Y%m%d').date()
            this_month_start = datetime.datetime(int(curr_proc_time_str[:4]), int(curr_proc_time_str[4:6]), 1)
            #today = datetime.datetime.today().date()
            last_month_end = this_month_start - datetime.timedelta(days=1)
            last_month_start = datetime.datetime(last_month_end.year, last_month_end.month, 1)
            last2_month_end = last_month_start - datetime.timedelta(days=1)
            last2_month_start = datetime.datetime(last2_month_end.year, last2_month_end.month, 1)
            last3_month_end = last2_month_start - datetime.timedelta(days=1)

            print(date_p, type(date_p))
        except:
            self.scr.insert(END, "无法查找到对应日期" + self.svar_proc_time1.get() + ".\n")
            self.master.update()
            return 0

        curr_proc_time_last_str = datetime.datetime.strftime(last_month_end,'%Y%m')
        curr_proc_time_last2_str = datetime.datetime.strftime(last2_month_end,'%Y%m')
        curr_proc_time_last3_str = datetime.datetime.strftime(last3_month_end, '%Y%m')
        logger.info('文件名前缀: '+self.file_from_wuliaoshiyong)
        wuliaoshiyongqingkjilubiao = self.file_from_wuliaoshiyong + curr_proc_time_str[:6] + '.xlsx'
        wuliaoshiyongqingkjilubiao_last = self.file_from_wuliaoshiyong + curr_proc_time_last_str+ '.xlsx'
        wuliaoshiyongqingkjilubiao_last2 = self.file_from_wuliaoshiyong + curr_proc_time_last2_str+ '.xlsx'
        wuliaoshiyongqingkjilubiao_last3 = self.file_from_wuliaoshiyong + curr_proc_time_last3_str+ '.xlsx'

        wuliaoshiyong_filelist = [wuliaoshiyongqingkjilubiao_last3,wuliaoshiyongqingkjilubiao_last2,wuliaoshiyongqingkjilubiao_last,wuliaoshiyongqingkjilubiao]
        self.scr.insert(END, wuliaoshiyongqingkjilubiao + ".\n")
        logger.info(wuliaoshiyongqingkjilubiao)
        self.scr.insert(END, wuliaoshiyongqingkjilubiao_last + ".\n")
        logger.info(wuliaoshiyongqingkjilubiao_last)
        self.scr.insert(END, wuliaoshiyongqingkjilubiao_last2 + ".\n")
        logger.info(wuliaoshiyongqingkjilubiao_last2)
        self.scr.insert(END, wuliaoshiyongqingkjilubiao_last3 + ".\n")
        logger.info(wuliaoshiyongqingkjilubiao_last3)
        self.master.update()

        #        to_xls_filename = os.path.join(self.savefile_dir,export_xls_filename)
        for wlsy_file in wuliaoshiyong_filelist:
            xlsfilename = wlsy_file
            if not os_path_exists(xlsfilename):
                print("文件不存在：", xlsfilename)
                logger.info("文件不存在："+ xlsfilename)
                self.scr.insert(END,"文件不存在："+ xlsfilename)
                self.scr.update()
                return (return_message)

        xlsfilename = '广发银行物料出入库明细表.xlsx'
        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            logger.info("文件不存在："+ xlsfilename + '\n')
            self.scr.insert(END,"文件不存在：" + xlsfilename)
            self.scr.update()
            return (return_message)

        workbook = xlrd.open_workbook(xlsfilename)
        sheet_curr = workbook.sheet_by_index(0)

        mxb_int_first_row = 3
        print("打开数据文件..." + xlsfilename)
        logger.info("打开数据文件..." + xlsfilename+'\n')
        self.scr.insert(END, "打开数据文件..." + xlsfilename+"\n")
        self.master.update()

        logger.info('sheet 广发')
        int_sheet_nrows = sheet_curr.nrows
        print('sheetname & lines:', sheet_curr, int_sheet_nrows)

        self.scr.insert(END,"\n")
        wuliaochurukumxb_list=[]
        #读取物料出入库明细表‘结存数量’
        for i in range(mxb_int_first_row,int_sheet_nrows):
            wuliaodaima_fromexcel = sheet_curr.cell(i,3).value
            wuliao_jiecunshulian = sheet_curr.cell(i,17).value
            self.scr.insert(END, str(wuliaodaima_fromexcel)+"\n")
            self.scr.insert(END, str(wuliao_jiecunshulian) + "\n")
            wuliaochurukumxb_list.append([wuliaodaima_fromexcel,wuliao_jiecunshulian])
        logger.info(wuliaochurukumxb_list)
        print(wuliaochurukumxb_list)

        wuliaoshiyong_grid = []
        logger.info('文件列表')
        logger.info(wuliaoshiyong_filelist)
        print(wuliaoshiyong_filelist)
        for wlsy_file in wuliaoshiyong_filelist:
            xlsfilename = wlsy_file
            workbook = xlrd.open_workbook(xlsfilename)
            sheet_curr = workbook.sheet_by_name('广发')

            #worksheetj = workbook['广发']
            int_first_row = 2
            print("打开数据文件..." + xlsfilename)
            logger.info("打开数据文件..." + xlsfilename)
            self.scr.insert(END, "打开数据文件..." + xlsfilename+"\n")
            self.master.update()

            logger.info('sheet 广发')
            int_sheet_nrows = sheet_curr.nrows
            print('sheetname & lines:', sheet_curr, int_sheet_nrows)

            shiyongqingkuang_int_first_row = 3
            for i in range(shiyongqingkuang_int_first_row,int_sheet_nrows):
                wuliaodaima_fromexcel = sheet_curr.cell(i,2).value
                if wuliaodaima_fromexcel == '':
                    break
                if len(wuliaodaima_fromexcel) < 2:
                    break
                print(wuliaodaima_fromexcel)
                data_date = sheet_curr.cell(1,7).value
                xls_date = xlrd.xldate_as_datetime(data_date, 0)
                month_range = monthrange(xls_date.year, xls_date.month)
                logger.info(str(month_range))
                for j in range(int(month_range[1])):
                    data_date = sheet_curr.cell(1,j*4+7).value
                    #print(data_date)
                    xls_date = xlrd.xldate_as_datetime(data_date,0)
                    if i == int_first_row:              #显示一行日期数据
                        print(xls_date)

                    try:
                        cell_value_temp = sheet_curr.cell(i, j * 4 + 7).value
                    except:
                        logging.info('明细数据格式错, excel位置：')
                        logging.info(i)
                        logging.info(j)
                    cell_value_ruku = cell_value_temp

                    try:
                        cell_value_temp = sheet_curr.cell(i, j * 4 + 7+1).value
                    except:
                        logging.info('明细数据格式错, excel位置：')
                        logging.info(i)
                        logging.info(j)
                    cell_value_dingdanshiyong = cell_value_temp

                    try:
                        cell_value_temp = sheet_curr.cell(i, j * 4 + 7+2).value
                    except:
                        logging.info('明细数据格式错, excel位置：')
                        logging.info(i)
                        logging.info(j)
                    cell_value_buhangengxinka = cell_value_temp

                    try:
                        cell_value_temp = sheet_curr.cell(i, j * 4 + 7+3).value
                    except:
                        logging.info('明细数据格式错, excel位置：')
                        logging.info(i)
                        logging.info(j)
                    cell_value_xiaohao = cell_value_temp
                    cell_value_comb = [cell_value_ruku,cell_value_dingdanshiyong,cell_value_buhangengxinka,cell_value_xiaohao]

                    wuliaoshiyong_grid.append([wuliaodaima_fromexcel,xls_date,cell_value_comb])
        wuliaoshiyong_grid.sort()
        logger.info(wuliaoshiyong_grid)

        xlsfilename = wuliaoshiyong_filelist[3]
        workbook_from = xlrd.open_workbook(xlsfilename)
        sheet_curr_from = workbook_from.sheet_by_name('广发')
        logger.info('当月excel情况表')
        logger.info(xlsfilename)

        xlsfilename = '配置\\模板-物料进销存日报表.xlsx'
        workbook = load_workbook(xlsfilename)  # 打开excel文件
        # 导出明细表begin
        logger.info('导出 ~明细表~ 表')
        self.scr.insert(END, "导出 ~明细表~ 表" + "\n")
        worksheetj = workbook['sheet1']  # 根据Sheet1这个sheet名字来获取该sheet
        #i = 0
        worksheetj.cell(1, 1).value = str(xls_date.year)+'年广发银行'+str(xls_date.month)+' 月物料收发进销存日报表'
        worksheetj.cell(3, 19).value = datetime.datetime.strptime(curr_proc_time_str[:6]+'01','%Y%m%d')
        int_first_row = 3

        # a)取“物料使用情况记录表”中的A列到E列填充到“物料进销存日报”的A—E列。
        for i in range(int_first_row, int_sheet_nrows):
            wuliaodaima_fromexcel = sheet_curr_from.cell(i, 2).value
            if wuliaodaima_fromexcel == '':
                break
            if len(wuliaodaima_fromexcel) < 2:
                break
            print(wuliaodaima_fromexcel)
            data_date = sheet_curr_from.cell(1, 7).value
            xls_date = xlrd.xldate_as_datetime(data_date, 0)
            month_range = monthrange(xls_date.year, xls_date.month)
            logger.info(str(month_range))

            self.scr.insert(END,'复制ABCDE. ')
            self.scr.update()

            for j in range(0,5):
                logger.info('复制ABCDE：')
                logger.info(i)
                logger.info(j)
                logger.info(sheet_curr_from.cell(i, j).value)
                worksheetj.cell(i+2,j+1).value = sheet_curr_from.cell(i, j).value

        # i)“出库数量”=同一时间日期、同一物料代码的物料使用情况记录表的“订单使用”+“订单使用（不含更新卡）”+“消耗”。
            day_curr_proc_time_str = curr_proc_time_str[6:8]
            for j in range(int(day_curr_proc_time_str)):
                data_date = sheet_curr.cell(1,j*4+7).value
                #print(data_date)
                xls_date = xlrd.xldate_as_datetime(data_date,0)
                if i == int_first_row:              #显示一行日期数据
                    print(xls_date)

                try:
                    cell_value_temp = sheet_curr.cell(i, j * 4 + 7).value
                except:
                    logging.info('明细数据格式错, excel位置：')
                    logging.info(i)
                    logging.info(j)
                cell_value_ruku = cell_value_temp

                try:
                    cell_value_temp = sheet_curr.cell(i, j * 4 + 7+1).value
                except:
                    logging.info('明细数据格式错, excel位置：')
                    logging.info(i)
                    logging.info(j)
                cell_value_dingdanshiyong = cell_value_temp

                try:
                    cell_value_temp = sheet_curr.cell(i, j * 4 + 7+2).value
                except:
                    logging.info('明细数据格式错, excel位置：')
                    logging.info(i)
                    logging.info(j)
                cell_value_buhangengxinka = cell_value_temp

                try:
                    cell_value_temp = sheet_curr.cell(i, j * 4 + 7+3).value
                except:
                    logging.info('明细数据格式错, excel位置：')
                    logging.info(i)
                    logging.info(j)
                cell_value_xiaohao = cell_value_temp

                worksheetj.cell(i+2,j * 2 +20).value = cell_value_dingdanshiyong + cell_value_buhangengxinka + cell_value_xiaohao

        # b)“上月仓库结存”=同一物料代码的使用情况记录表的“上月车间结存量”+物料出入库明细表的“结存数量”。
            wuliaodaima_fromexcel = sheet_curr_from.cell(i, 2).value
            shangyuechejianjiecun = sheet_curr_from.cell(i, 6).value
            logger.info(wuliaodaima_fromexcel)
            for k in range(len(wuliaochurukumxb_list)):
                mxb_jiecunshuliang_match = 0
                #temp_compare = wuliaochurukumxb_list[k][0]
                mxb_jiecunshuliang = wuliaochurukumxb_list[k][1]

                worksheetj.cell(i+2,6).value = shangyuechejianjiecun + mxb_jiecunshuliang
            self.scr.insert(END,'物料明细表_物料：')
            self.scr.insert(END,wuliaodaima_fromexcel)
            self.scr.insert(END,'\n')

            self.scr.update()

        # c)“上月余下订单数”、“本月订单数”、“未入库数量”无需处理，由客服填写。
        # d)“上月发出总量”=T - 1月物料进销存日报的“本月发出总数”。
        last_month_xlsfilename = '广发银行物料进销存日报'+curr_proc_time_last_str+'.xlsx'
        if not os_path_exists(last_month_xlsfilename):
            print("文件不存在：", last_month_xlsfilename)
            logger.info("文件不存在："+ last_month_xlsfilename + '\n')
            self.scr.insert(END,"文件不存在：" + last_month_xlsfilename+'\n')
            self.scr.update()
            return (return_message)
        workbook_wuliaojxc_lastmonth = xlrd.open_workbook(last_month_xlsfilename)
        sheet_wuliaojxc_lastmonth = workbook_wuliaojxc_lastmonth.sheet_by_index(0)
        int_sheet_nrows = sheet_wuliaojxc_lastmonth.nrows
        
        wuliaojxc_int_first_row = 4
        for i in range(wuliaojxc_int_first_row, int_sheet_nrows):
            wuliaodaima_xuhao_fromexcel = sheet_wuliaojxc_lastmonth.cell(i, 0).value
            if wuliaodaima_xuhao_fromexcel == '':
                break
            if isinstance(wuliaodaima_xuhao_fromexcel,float):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break
            if isinstance(wuliaodaima_xuhao_fromexcel,int):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break

            print(wuliaodaima_xuhao_fromexcel)

            wuliaodaima_fromexcel = sheet_wuliaojxc_lastmonth.cell(i, 2).value
            shangyue_benyuefachuzongshu = sheet_wuliaojxc_lastmonth.cell(i, 11).value
            logger.info(wuliaodaima_fromexcel)
            logger.info(shangyue_benyuefachuzongshu)
            for k in range(len(wuliaochurukumxb_list)):
                mxb_jiecunshuliang_match = 0
                temp_compare = wuliaochurukumxb_list[k][0]
                worksheetj.cell(k+5,13).value = shangyue_benyuefachuzongshu   #上月数据，保存到本月文件
                logger.info("写入上月发出总数: " + str(shangyue_benyuefachuzongshu))

            self.scr.insert(END,'上月发出总数...\n')
            self.scr.update()

        # e)“前1周周用量”=T日所在星期的前1个自然周（7天）的用量之和。
        # f)“前2周周用量”=T日所在星期的前2个自然周（14天）的用量之和。
        # g)“前12周周用量”=T日所在星期的前12个自然周（84天）的用量之和。
        last_1_day = datetime.datetime.strptime(curr_proc_time_str,'%Y%m%d') + datetime.timedelta(days=-1)
        logger.info(last_1_day)
        
        wuliaoshiyong_grid_buhangengxinka = []
        for i in range(len(wuliaoshiyong_grid)):
            temp_grid_oneline = wuliaoshiyong_grid[i]
            temp_grid_date = temp_grid_oneline[1]
            if temp_grid_date == last_1_day:
                sum_value = 0
                for j in range(i,i-84,-1):
                    temp2_grid_oneline = wuliaoshiyong_grid[j]
                    temp2_grid_yongliang_4_value = temp2_grid_oneline[2]
                    temp2_grid_buhangengxinka = temp2_grid_yongliang_4_value[2]
                    sum_value = sum_value + temp2_grid_buhangengxinka
                    if j > i-7:
                        temp_wuliao_7days = sum_value
                    if j > i-14:
                        temp_wuliao_14days = sum_value
                    temp_wuliao_84days = sum_value
                wuliaoshiyong_grid_buhangengxinka.append([temp_grid_oneline[0],temp_wuliao_7days,temp_wuliao_14days,temp_wuliao_84days])
        logger.info('物料使用不含更新卡7 14 84 合计')
        logger.info(wuliaoshiyong_grid_buhangengxinka)

        wuliaojxc_int_first_row = 4
        #worksheetj 此变量只用于用openpyxl打开的excel表格
        for i in range(wuliaojxc_int_first_row, wuliaojxc_int_first_row + len(wuliaoshiyong_grid_buhangengxinka)+1):
            wuliaodaima_xuhao_fromexcel = worksheetj.cell(i, 1).value
            logger.info('efg序号:'+str(wuliaodaima_xuhao_fromexcel))
            if wuliaodaima_xuhao_fromexcel == '':
                break
            if isinstance(wuliaodaima_xuhao_fromexcel,float):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break
            if isinstance(wuliaodaima_xuhao_fromexcel,int):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break

            wuliaodaima_fromexcel = worksheetj.cell(i, 3).value
            for k in range(len(wuliaoshiyong_grid_buhangengxinka)):
                mxb_jiecunshuliang_match = 0
                temp_compare = wuliaoshiyong_grid_buhangengxinka[k][0]
                if wuliaodaima_fromexcel == temp_compare:
                    mxb_jiecunshuliang_match = 1
                    worksheetj.cell(i, 15).value = wuliaoshiyong_grid_buhangengxinka[k][1]
                    worksheetj.cell(i, 16).value = wuliaoshiyong_grid_buhangengxinka[k][2]
                    worksheetj.cell(i, 17).value = wuliaoshiyong_grid_buhangengxinka[k][3]
                    break
                else:
                    mxb_jiecunshuliang = 0
            if mxb_jiecunshuliang_match == 0 :
                self.scr.insert(END,'物料明细表71484无法匹配 物料：')
                self.scr.insert(END,str(wuliaodaima_fromexcel))
                self.scr.update()
                logger.info('物料明细表71484无法匹配 物料：')
                logger.info(wuliaodaima_fromexcel)
            else:
                logger.info("写入7 14 84天数据: ")
                logger.info(wuliaoshiyong_grid_buhangengxinka[k])
                self.scr.insert(END,'写入7 14 84天数据\n')
                self.scr.update()

        # h)“入库数量”=同一时间日期、同一物料代码的物料出入库明细表的“入库数量”。
        curr_churukumingxibiao_xlsfilename = '广发银行物料出入库明细表.xlsx'
        if not os_path_exists(curr_churukumingxibiao_xlsfilename):
            print("文件不存在：", curr_churukumingxibiao_xlsfilename)
            logger.info("文件不存在："+ curr_churukumingxibiao_xlsfilename + '\n')
            self.scr.insert(END,"文件不存在：" + curr_churukumingxibiao_xlsfilename+'\n')
            self.scr.update()
            return (return_message)
        workbook_churukumingxibiao = xlrd.open_workbook(curr_churukumingxibiao_xlsfilename)
        sheet_churukumingxibiao = workbook_churukumingxibiao.sheet_by_index(0)
        int_sheet_nrows_churukumingxibiao = sheet_churukumingxibiao.nrows
        
        churukumingxibiao_int_first_row = 3
        for i in range(churukumingxibiao_int_first_row, int_sheet_nrows_churukumingxibiao):
            wuliaodaima_xuhao_fromexcel = sheet_churukumingxibiao.cell(i, 0).value
            if wuliaodaima_xuhao_fromexcel == '':
                break
            if isinstance(wuliaodaima_xuhao_fromexcel,float):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break
            if isinstance(wuliaodaima_xuhao_fromexcel,int):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break

            wuliaodaima_fromexcel = sheet_churukumingxibiao.cell(i, 3).value
            churukumingxibiao_rukushuliang = sheet_churukumingxibiao.cell(i, 17).value
            logger.info(wuliaodaima_fromexcel)
            logger.info(churukumingxibiao_rukushuliang)
            for k in range(len(wuliaochurukumxb_list)):
                mxb_jiecunshuliang_match = 0
                temp_compare = wuliaochurukumxb_list[k][0]
                if wuliaodaima_fromexcel == temp_compare:
                    mxb_jiecunshuliang_match = 1
                    mxb_jiecunshuliang = wuliaochurukumxb_list[k][1]
                    break
                else:
                    mxb_jiecunshuliang = 0
            if mxb_jiecunshuliang_match == 0 :
                self.scr.insert(END,'物料明细表h无法匹配 物料：')
                self.scr.insert(END,wuliaodaima_fromexcel)
                self.scr.update()
                logger.info('物料明细表h无法匹配 物料：')
                logger.info(wuliaodaima_fromexcel)
            else:
                worksheetj.cell(k+5,int(day_curr_proc_time_str)*2 +17).value = churukumingxibiao_rukushuliang   #来自物料出入库明细表，保存到进销存日报
                logger.info("进销存日报h入库数量: ")
                logger.info(churukumingxibiao_rukushuliang)
            self.scr.insert(END,'进销存日报h入库数量...'+str(churukumingxibiao_rukushuliang)+'\n')
            self.scr.update()

        temp_proc_time1 = self.svar_proc_time1.get()
        temp_write_filename = '广发银行物料进销存日报'+temp_proc_time1[:6]+'.xlsx'
        workbook.save(temp_write_filename)
        print('=' * 40)
        self.scr.insert(END, "文件输出..\n" )
        self.scr.insert(END, temp_write_filename + '\n' )
        self.master.update()
    #===========================jxc end


    # 物料预警表 文件处理：
    def wuliao_yujingbiao_file_proc(self, txtfilename, xlsfilename):
        return_message = 'err'
        curr_proc_time_str = self.svar_proc_time1.get()
        try:
            date_p = datetime.datetime.strptime(curr_proc_time_str, '%Y%m%d').date()
            this_month_start = datetime.datetime(int(curr_proc_time_str[:4]), int(curr_proc_time_str[4:6]), 1)
            #today = datetime.datetime.today().date()
            next_month_anyday = this_month_start + datetime.timedelta(days=31)
            next_month_start =  datetime.datetime(next_month_anyday.year, next_month_anyday.month, 1)
            this_month_end = next_month_start  - datetime.timedelta(days=1)
            last_month_end = this_month_start - datetime.timedelta(days=1)
            last_month_start = datetime.datetime(last_month_end.year, last_month_end.month, 1)
            last2_month_end = last_month_start - datetime.timedelta(days=1)
            last2_month_start = datetime.datetime(last2_month_end.year, last2_month_end.month, 1)
            last3_month_end = last2_month_start - datetime.timedelta(days=1)

            print(date_p, type(date_p))
        except:
            self.scr.insert(END, "无法查找到对应日期" + self.svar_proc_time1.get() + ".\n")
            self.master.update()
            return 0

        curr_proc_time_last_str = datetime.datetime.strftime(last_month_end,'%Y%m')
        curr_proc_time_last2_str = datetime.datetime.strftime(last2_month_end,'%Y%m')
        curr_proc_time_last3_str = datetime.datetime.strftime(last3_month_end, '%Y%m')
        logger.info('文件名前缀: '+self.file_from_wuliaoshiyong)
        wuliaoshiyongqingkjilubiao = self.file_from_wuliaoshiyong + curr_proc_time_str[:6] + '.xlsx'
        wuliaoshiyongqingkjilubiao_last = self.file_from_wuliaoshiyong + curr_proc_time_last_str+ '.xlsx'
        wuliaoshiyongqingkjilubiao_last2 = self.file_from_wuliaoshiyong + curr_proc_time_last2_str+ '.xlsx'
        wuliaoshiyongqingkjilubiao_last3 = self.file_from_wuliaoshiyong + curr_proc_time_last3_str+ '.xlsx'

        wuliaoshiyong_filelist = [wuliaoshiyongqingkjilubiao_last3,wuliaoshiyongqingkjilubiao_last2,wuliaoshiyongqingkjilubiao_last,wuliaoshiyongqingkjilubiao]
        self.scr.insert(END, wuliaoshiyongqingkjilubiao + ".\n")
        logger.info(wuliaoshiyongqingkjilubiao)
        self.scr.insert(END, wuliaoshiyongqingkjilubiao_last + ".\n")
        logger.info(wuliaoshiyongqingkjilubiao_last)
        self.scr.insert(END, wuliaoshiyongqingkjilubiao_last2 + ".\n")
        logger.info(wuliaoshiyongqingkjilubiao_last2)
        self.scr.insert(END, wuliaoshiyongqingkjilubiao_last3 + ".\n")
        logger.info(wuliaoshiyongqingkjilubiao_last3)
        self.master.update()

        #        to_xls_filename = os.path.join(self.savefile_dir,export_xls_filename)
        for wlsy_file in wuliaoshiyong_filelist:
            xlsfilename = wlsy_file
            if not os_path_exists(xlsfilename):
                print("文件不存在：", xlsfilename)
                logger.info("文件不存在："+ xlsfilename)
                self.scr.insert(END,"文件不存在："+ xlsfilename)
                self.scr.update()
                return (return_message)

        xlsfilename = '广发银行物料出入库明细表.xlsx'
        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            logger.info("文件不存在："+ xlsfilename + '\n')
            self.scr.insert(END,"文件不存在：" + xlsfilename)
            self.scr.update()
            return (return_message)

        workbook = xlrd.open_workbook(xlsfilename)
        sheet_curr = workbook.sheet_by_index(0)

        mxb_int_first_row = 3
        print("打开数据文件..." + xlsfilename)
        logger.info("打开数据文件..." + xlsfilename+'\n')
        self.scr.insert(END, "打开数据文件..." + xlsfilename+"\n")
        self.master.update()

        logger.info('sheet 广发')
        int_sheet_nrows = sheet_curr.nrows
        #int_sheet_ncols = sheet_curr.ncols
        print('sheetname & lines:', sheet_curr, int_sheet_nrows)

        self.scr.insert(END,"\n")
        wuliaochurukumxb_list=[]
        #读取物料出入库明细表‘结存数量’
        for i in range(mxb_int_first_row,int_sheet_nrows):
            wuliaodaima_fromexcel = sheet_curr.cell(i,3).value
            wuliao_jiecunshulian = sheet_curr.cell(i,17).value
            self.scr.insert(END, str(wuliaodaima_fromexcel)+"\n")
            self.scr.insert(END, str(wuliao_jiecunshulian) + "\n")
            wuliaochurukumxb_list.append([wuliaodaima_fromexcel,wuliao_jiecunshulian])
        logger.info(wuliaochurukumxb_list)
        print(wuliaochurukumxb_list)

        wuliaoshiyong_grid = []
        logger.info('文件列表')
        logger.info(wuliaoshiyong_filelist)
        print(wuliaoshiyong_filelist)
        for wlsy_file in wuliaoshiyong_filelist:
            xlsfilename = wlsy_file
            workbook = xlrd.open_workbook(xlsfilename)
            sheet_curr = workbook.sheet_by_name('广发')

            #worksheetj = workbook['广发']
            int_first_row = 2
            print("打开数据文件..." + xlsfilename)
            logger.info("打开数据文件..." + xlsfilename)
            self.scr.insert(END, "打开数据文件..." + xlsfilename+"\n")
            self.master.update()

            logger.info('sheet 广发')
            int_sheet_nrows = sheet_curr.nrows
            int_sheet_ncols = sheet_curr.ncols
            print('sheetname & lines:', sheet_curr, int_sheet_nrows)

            shiyongqingkuang_int_first_row = 3
            for i in range(shiyongqingkuang_int_first_row,int_sheet_nrows):
                wuliaodaima_fromexcel = sheet_curr.cell(i,2).value
                if wuliaodaima_fromexcel == '':
                    break
                if len(wuliaodaima_fromexcel) < 2:
                    break
                print(wuliaodaima_fromexcel)
                data_date = sheet_curr.cell(1,7).value
                xls_date = xlrd.xldate_as_datetime(data_date, 0)
                month_range = monthrange(xls_date.year, xls_date.month)
                logger.info(str(month_range))
                for j in range(int(month_range[1])):
                    data_date = sheet_curr.cell(1,j*4+7).value
                    #print(data_date)
                    xls_date = xlrd.xldate_as_datetime(data_date,0)
                    if i == int_first_row:              #显示一行日期数据
                        print(xls_date)

                    try:
                        cell_value_temp = sheet_curr.cell(i, j * 4 + 7).value
                    except:
                        logging.info('明细数据格式错, excel位置：')
                        logging.info(i)
                        logging.info(j)
                    cell_value_ruku = cell_value_temp

                    try:
                        cell_value_temp = sheet_curr.cell(i, j * 4 + 7+1).value
                    except:
                        logging.info('明细数据格式错, excel位置：')
                        logging.info(i)
                        logging.info(j)
                    cell_value_dingdanshiyong = cell_value_temp

                    try:
                        cell_value_temp = sheet_curr.cell(i, j * 4 + 7+2).value
                    except:
                        logging.info('明细数据格式错, excel位置：')
                        logging.info(i)
                        logging.info(j)
                    cell_value_buhangengxinka = cell_value_temp

                    try:
                        cell_value_temp = sheet_curr.cell(i, j * 4 + 7+3).value
                    except:
                        logging.info('明细数据格式错, excel位置：')
                        logging.info(i)
                        logging.info(j)
                    cell_value_xiaohao = cell_value_temp
                    cell_value_comb = [cell_value_ruku,cell_value_dingdanshiyong,cell_value_buhangengxinka,cell_value_xiaohao]

                    wuliaoshiyong_grid.append([wuliaodaima_fromexcel,xls_date,cell_value_comb])
        wuliaoshiyong_grid.sort()
        logger.info(wuliaoshiyong_grid)

        xlsfilename = wuliaoshiyong_filelist[3]
        workbook_from = xlrd.open_workbook(xlsfilename)
        sheet_curr_from = workbook_from.sheet_by_name('广发')
        logger.info('当月excel情况表')
        logger.info(xlsfilename)

        xlsfilename = '东信物料预警表-模板.xlsx'
        workbook = load_workbook(xlsfilename)  # 打开excel文件
        # 导出明细表begin
        logger.info('东信物料预警表-模板.xlsx')
        self.scr.insert(END, "东信物料预警表-模板.xlsx" + "\n")
        worksheetj = workbook['预警表']  # 根据Sheet1这个sheet名字来获取该sheet
        #无表头：worksheetj.cell(1, 1).value = str(xls_date.year)+'年广发银行'+str(xls_date.month)+' 月物料收发进销存日报表'
        #worksheetj.cell(3, 19).value = datetime.datetime.strptime(curr_proc_time_str[:6]+'01','%Y%m%d')
        
        int_first_row = 3

        # a)取“物料进销存日报”的A到E列填充到“物料预警表”的A - E列。
        for i in range(int_first_row, int_sheet_nrows):
            wuliaodaima_fromexcel = sheet_curr_from.cell(i, 2).value
            if wuliaodaima_fromexcel == '':
                break
            if len(wuliaodaima_fromexcel) < 2:
                break
            print(wuliaodaima_fromexcel)
            data_date = sheet_curr_from.cell(1, 7).value
            xls_date = xlrd.xldate_as_datetime(data_date, 0)
            month_range = monthrange(xls_date.year, xls_date.month)
            logger.info(str(month_range))

            self.scr.insert(END,'复制ABCDE.')
            self.scr.update()
            for j in range(0,5):
                logger.info('复制ABCDE：')
                logger.info(i)
                logger.info(j)
                logger.info(sheet_curr_from.cell(i, j).value)
                worksheetj.cell(i-1,j+1).value = sheet_curr_from.cell(i, j).value

        # b)“使用状态”=同一物料代码的“物料使用情况记录表”的“使用状态”。
        curr_month_shiyongqingkuangjlb = '广发银行物料使用情况记录表'+curr_proc_time_str[:6]+'.xlsx'
        if not os_path_exists(curr_month_shiyongqingkuangjlb):
            print("文件不存在：", curr_month_shiyongqingkuangjlb)
            logger.info("文件不存在："+ curr_month_shiyongqingkuangjlb + '\n')
            self.scr.insert(END,"文件不存在：" + curr_month_shiyongqingkuangjlb+'\n')
            self.scr.update()
            return (return_message)
        workbook_wuliaoyjb_shiyongqingkuangjlb = xlrd.open_workbook(curr_month_shiyongqingkuangjlb)
        sheet_wuliaoyjb_shiyongqingkuangjlb = workbook_wuliaoyjb_shiyongqingkuangjlb.sheet_by_name('广发')
        int_sheet_nrows = sheet_wuliaoyjb_shiyongqingkuangjlb.nrows
        
        wuliaoshiyong_int_first_row = 3
        for i in range(wuliaoshiyong_int_first_row, int_sheet_nrows):
            wuliaodaima_fromexcel = sheet_wuliaoyjb_shiyongqingkuangjlb.cell(i, 2).value
            yjb_shiyongzhuangtai = sheet_wuliaoyjb_shiyongqingkuangjlb.cell(i, 5).value
            logger.info(wuliaodaima_fromexcel)
            logger.info(yjb_shiyongzhuangtai)
            for k in range(len(wuliaochurukumxb_list)):
                mxb_jiecunshuliang_match = 0
                temp_compare = wuliaochurukumxb_list[k][0]
                if wuliaodaima_fromexcel == temp_compare:
                    mxb_jiecunshuliang_match = 1
                    mxb_jiecunshuliang = wuliaochurukumxb_list[k][1]
                    break
                else:
                    mxb_jiecunshuliang = 0
            if mxb_jiecunshuliang_match == 0 :
                self.scr.insert(END,'物料明细表yjb-b无匹配 物料：')
                self.scr.insert(END,str(wuliaodaima_fromexcel)+'\n')
                self.scr.update()
            else:
                worksheetj.cell(i-1,6).value = yjb_shiyongzhuangtai   #上月数据，保存到本月文件
                logger.info("yjb使用状态: " + str(yjb_shiyongzhuangtai))
            self.scr.insert(END,'写入使用状态...\n')
            self.scr.update()

        # e)“YYYYMM成品数（不含更新卡）”=同一物料代码T月份“物料使用情况记录表”的“订单使用（不含更新卡）总量”
        # 对同一文件文件操作，将此项目提前
        worksheetj.cell(1,9).value = wuliaoshiyongqingkjilubiao[-11:-5] + '成品数（不含更新卡）'
        this_month_int_days = int(this_month_end.day)
        wuliaoyjb_int_first_row = 3
        for i in range(wuliaoyjb_int_first_row, int_sheet_nrows):
            wuliaodaima_xuhao_fromexcel = sheet_wuliaoyjb_shiyongqingkuangjlb.cell(i, 0).value
            if wuliaodaima_xuhao_fromexcel == '':
                break
            if isinstance(wuliaodaima_xuhao_fromexcel,float):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break
            if isinstance(wuliaodaima_xuhao_fromexcel,int):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break

            print('ysb-c,xuhao:',wuliaodaima_xuhao_fromexcel)
            wuliaodaima_fromexcel = sheet_wuliaoyjb_shiyongqingkuangjlb.cell(i, 2).value
            shangyue_buhangengxinka = sheet_wuliaoyjb_shiyongqingkuangjlb.cell(i, 6+this_month_int_days*4+3).value
            logger.info(wuliaodaima_fromexcel)
            logger.info(shangyue_buhangengxinka)
            for k in range(len(wuliaochurukumxb_list)):
                mxb_jiecunshuliang_match = 0
                temp_compare = wuliaochurukumxb_list[k][0]
                if wuliaodaima_fromexcel == temp_compare:
                    mxb_jiecunshuliang_match = 1
                    mxb_jiecunshuliang = wuliaochurukumxb_list[k][1]
                    break
                else:
                    mxb_jiecunshuliang = 0
            if mxb_jiecunshuliang_match == 0 :
                self.scr.insert(END,'物料明细表无法匹配 物料：')
                self.scr.insert(END,wuliaodaima_fromexcel)
                self.scr.update()
            else:
                worksheetj.cell(i-1,9).value = shangyue_buhangengxinka   #上月数据，保存到本月文件
                logger.info("T,dingdanshiyong(buhangengxinka)zongliang: " + str(shangyue_buhangengxinka))
            self.scr.insert(END,'T订单使用上月发出总数...\n')
            self.scr.update()



        # c)“YYYYMM - 2 成品数（不含更新卡）”=同一物料代码T - 2 月份“物料使用情况记录表”的“订单使用（不含更新卡）总量”
        #wuliaoshiyongqingkjilubiao_last2
        worksheetj.cell(1,7).value = wuliaoshiyongqingkjilubiao_last2[-11:-5] + '成品数（不含更新卡）'
        last2_month_int_days = int(last2_month_end.day)
        if not os_path_exists(wuliaoshiyongqingkjilubiao_last2):
            print("文件不存在：", wuliaoshiyongqingkjilubiao_last2)
            logger.info("文件不存在："+ wuliaoshiyongqingkjilubiao_last2 + '\n')
            self.scr.insert(END,"文件不存在：" + wuliaoshiyongqingkjilubiao_last2+'\n')
            self.scr.update()
            return (return_message)
        workbook_wuliaoyjb_lastmonth2 = xlrd.open_workbook(wuliaoshiyongqingkjilubiao_last2)
        sheet_wuliaoyjb_lastmonth = workbook_wuliaoyjb_lastmonth2.sheet_by_name('广发')
        int_sheet_nrows = sheet_wuliaoyjb_lastmonth.nrows
        
        wuliaoyjb_int_first_row = 3
        for i in range(wuliaoyjb_int_first_row, int_sheet_nrows):
            wuliaodaima_xuhao_fromexcel = sheet_wuliaoyjb_lastmonth.cell(i, 0).value
            if wuliaodaima_xuhao_fromexcel == '':
                break
            if isinstance(wuliaodaima_xuhao_fromexcel,float):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break
            if isinstance(wuliaodaima_xuhao_fromexcel,int):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break

            print('ysb-c,xuhao:',wuliaodaima_xuhao_fromexcel)

            wuliaodaima_fromexcel = sheet_wuliaoyjb_lastmonth.cell(i, 2).value
            shangyue_buhangengxinka = sheet_wuliaoyjb_lastmonth.cell(i, 6+last2_month_int_days*4+3).value
            logger.info(wuliaodaima_fromexcel)
            logger.info(shangyue_buhangengxinka)
            for k in range(len(wuliaochurukumxb_list)):
                mxb_jiecunshuliang_match = 0
                temp_compare = wuliaochurukumxb_list[k][0]
                if wuliaodaima_fromexcel == temp_compare:
                    mxb_jiecunshuliang_match = 1
                    mxb_jiecunshuliang = wuliaochurukumxb_list[k][1]
                    break
                else:
                    mxb_jiecunshuliang = 0
            if mxb_jiecunshuliang_match == 0 :
                self.scr.insert(END,'物料明细表无法匹配 物料：')
                self.scr.insert(END,wuliaodaima_fromexcel)
                self.scr.update()
            else:
                worksheetj.cell(i-1,7).value = shangyue_buhangengxinka   #上月数据，保存到本月文件
                logger.info("T-2,dingdanshiyong(buhangengxinka)zongliang: " + str(shangyue_buhangengxinka))
            self.scr.insert(END,'T-2订单使用上月发出总数...\n')
            self.scr.update()
 
        # d)“YYYYMM - 1 成品数（不含更新卡）”=同一物料代码T - 1 月份“物料使用情况记录表”的“订单使用（不含更新卡）总量”
        worksheetj.cell(1,8).value = wuliaoshiyongqingkjilubiao_last[-11:-5] + '成品数（不含更新卡）'
        last_month_int_days = int(last_month_end.day)
        if not os_path_exists(wuliaoshiyongqingkjilubiao_last):
            print("文件不存在：", wuliaoshiyongqingkjilubiao_last)
            logger.info("文件不存在："+ wuliaoshiyongqingkjilubiao_last + '\n')
            self.scr.insert(END,"文件不存在：" + wuliaoshiyongqingkjilubiao_last+'\n')
            self.scr.update()
            return (return_message)
        workbook_wuliaoyjb_lastmonth = xlrd.open_workbook(wuliaoshiyongqingkjilubiao_last)
        sheet_wuliaoyjb_lastmonth = workbook_wuliaoyjb_lastmonth.sheet_by_name('广发')
        int_sheet_nrows = sheet_wuliaoyjb_lastmonth.nrows
        
        wuliaoyjb_int_first_row = 3
        for i in range(wuliaoyjb_int_first_row, int_sheet_nrows):
            wuliaodaima_xuhao_fromexcel = sheet_wuliaoyjb_lastmonth.cell(i, 0).value
            if wuliaodaima_xuhao_fromexcel == '':
                break
            if isinstance(wuliaodaima_xuhao_fromexcel,float):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break
            if isinstance(wuliaodaima_xuhao_fromexcel,int):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break

            print('ysb-c,xuhao:',wuliaodaima_xuhao_fromexcel)

            wuliaodaima_fromexcel = sheet_wuliaoyjb_lastmonth.cell(i, 2).value
            shangyue_buhangengxinka = sheet_wuliaoyjb_lastmonth.cell(i, 6+last2_month_int_days*4+3).value
            logger.info(wuliaodaima_fromexcel)
            logger.info(shangyue_buhangengxinka)
            for k in range(len(wuliaochurukumxb_list)):
                mxb_jiecunshuliang_match = 0
                temp_compare = wuliaochurukumxb_list[k][0]
                if wuliaodaima_fromexcel == temp_compare:
                    mxb_jiecunshuliang_match = 1
                    mxb_jiecunshuliang = wuliaochurukumxb_list[k][1]
                    break
                else:
                    mxb_jiecunshuliang = 0
            if mxb_jiecunshuliang_match == 0 :
                self.scr.insert(END,'物料明细表无法匹配 物料：')
                self.scr.insert(END,wuliaodaima_fromexcel)
                self.scr.update()
            else:
                worksheetj.cell(i-1,8).value = shangyue_buhangengxinka   #上月数据，保存到本月文件
                logger.info("T-1,dingdanshiyong(buhangengxinka)zongliang: " + str(shangyue_buhangengxinka))
            self.scr.insert(END,'T-1订单使用上月发出总数...\n')
            self.scr.update()

        # f)“14 天日均用量”=同一物料代码“物料使用情况记录表”T日往前推算14天的“订单使用（不含更新卡）”之和除以14。
        # g)“7 天日均用量”=同一物料代码“物料使用情况记录表”T日往前推算7天的“订单使用（不含更新卡）”之和除以7。
        last_1_day = datetime.datetime.strptime(curr_proc_time_str,'%Y%m%d') + datetime.timedelta(days=-1)
        logger.info(last_1_day)
        
        wuliaoshiyong_grid_buhangengxinka = []
        for i in range(len(wuliaoshiyong_grid)):
            temp_grid_oneline = wuliaoshiyong_grid[i]
            temp_grid_date = temp_grid_oneline[1]
            if temp_grid_date == last_1_day:
                sum_value = 0
                for j in range(i,i-84,-1):
                    temp2_grid_oneline = wuliaoshiyong_grid[j]
                    temp2_grid_yongliang_4_value = temp2_grid_oneline[2]
                    temp2_grid_buhangengxinka = temp2_grid_yongliang_4_value[2]
                    sum_value = sum_value + temp2_grid_buhangengxinka
                    if j > i-7:
                        temp_wuliao_7days = sum_value
                    if j > i-14:
                        temp_wuliao_14days = sum_value
                    temp_wuliao_84days = sum_value
                wuliaoshiyong_grid_buhangengxinka.append([temp_grid_oneline[0],temp_wuliao_7days,temp_wuliao_14days,temp_wuliao_84days])
        logger.info('物料使用不含更新卡7 14 84 合计')
        logger.info(wuliaoshiyong_grid_buhangengxinka)

        wuliaojxc_int_first_row = 2
        #worksheetj 此变量只用于用openpyxl打开的excel表格
        for i in range(wuliaojxc_int_first_row, wuliaojxc_int_first_row + len(wuliaoshiyong_grid_buhangengxinka)+1):
            wuliaodaima_xuhao_fromexcel = worksheetj.cell(i, 1).value
            logger.info('efg序号:'+str(wuliaodaima_xuhao_fromexcel))
            if wuliaodaima_xuhao_fromexcel == '':
                break
            if isinstance(wuliaodaima_xuhao_fromexcel,float):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break
            if isinstance(wuliaodaima_xuhao_fromexcel,int):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break

            wuliaodaima_fromexcel = worksheetj.cell(i, 3).value
            for k in range(len(wuliaoshiyong_grid_buhangengxinka)):
                mxb_jiecunshuliang_match = 0
                temp_compare = wuliaoshiyong_grid_buhangengxinka[k][0]
                if wuliaodaima_fromexcel == temp_compare:
                    mxb_jiecunshuliang_match = 1
                    worksheetj.cell(i, 12).value = wuliaoshiyong_grid_buhangengxinka[k][1]
                    worksheetj.cell(i, 11).value = wuliaoshiyong_grid_buhangengxinka[k][2]
                    break
                else:
                    mxb_jiecunshuliang = 0
            if mxb_jiecunshuliang_match == 0 :
                self.scr.insert(END,'物料明细表714无法匹配 物料：')
                self.scr.insert(END,str(wuliaodaima_fromexcel))
                self.scr.update()
                logger.info('物料明细表714无法匹配 物料：')
                logger.info(wuliaodaima_fromexcel)
            else:
                logger.info("写入7 14 天数据: ")
                logger.info(wuliaoshiyong_grid_buhangengxinka[k])
                self.scr.insert(END,'写入7 14 天数据\n')
                self.scr.update()

        # h)“截止T月MM日库存量”=同一物料代码“物料进销存日报”的“库存总数”。
        worksheetj.cell(1,14).value = '截止'+curr_proc_time_str[4:6]+'月'+curr_proc_time_str[6:8]+'日库存量'
        curr_wuliaojxcribao_xlsfilename = '广发银行物料进销存日报' + curr_proc_time_str[:6] + '.xlsx'
        if not os_path_exists(curr_wuliaojxcribao_xlsfilename):
            print("文件不存在：", curr_wuliaojxcribao_xlsfilename)
            logger.info("文件不存在："+ curr_wuliaojxcribao_xlsfilename + '\n')
            self.scr.insert(END,"文件不存在：" + curr_wuliaojxcribao_xlsfilename+'\n')
            self.scr.update()
            return (return_message)
        workbook_wuliaojxcribao = xlrd.open_workbook(curr_wuliaojxcribao_xlsfilename)
        sheet_wuliaojxcribao = workbook_wuliaojxcribao.sheet_by_index(0)
        int_sheet_nrows_wuliaojxcribao = sheet_wuliaojxcribao.nrows
        
        wuliaojxcribao_int_first_row = 4
        for i in range(wuliaojxcribao_int_first_row, int_sheet_nrows_wuliaojxcribao):
            wuliaodaima_xuhao_fromexcel = sheet_wuliaojxcribao.cell(i, 0).value
            if wuliaodaima_xuhao_fromexcel == '':
                break
            if isinstance(wuliaodaima_xuhao_fromexcel,float):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break
            if isinstance(wuliaodaima_xuhao_fromexcel,int):
                if len(str(wuliaodaima_xuhao_fromexcel)) < 1:
                    break

            print(wuliaodaima_xuhao_fromexcel)

            wuliaodaima_fromexcel = sheet_wuliaojxcribao.cell(i, 2).value
            churukumingxibiao_rukushuliang = sheet_wuliaojxcribao.cell(i, 10).value
            logger.info(wuliaodaima_fromexcel)
            logger.info(churukumingxibiao_rukushuliang)
            for k in range(len(wuliaochurukumxb_list)):
                mxb_jiecunshuliang_match = 0
                temp_compare = wuliaochurukumxb_list[k][0]
                if wuliaodaima_fromexcel == temp_compare:
                    mxb_jiecunshuliang_match = 1
                    mxb_jiecunshuliang = wuliaochurukumxb_list[k][1]
                    break
                else:
                    mxb_jiecunshuliang = 0
            if mxb_jiecunshuliang_match == 0 :
                logger.info('进销存日报库存总数')
                logger.info(churukumingxibiao_rukushuliang)
                self.scr.insert(END,'进销存日报库存总数\n')
                self.scr.update()
            else:
                logger.info('进销存日报库存总数')
                logger.info(churukumingxibiao_rukushuliang)
                worksheetj.cell(k+2,14).value = churukumingxibiao_rukushuliang   #来自物料出入库明细表，保存到进销存日报
                logger.info("进销存日报入库数量: " + str(churukumingxibiao_rukushuliang))
                self.scr.insert(END,'进销存日报库存总数\n')

        temp_proc_time1 = self.svar_proc_time1.get()
        temp_write_filename = '东信物料预警表' + temp_proc_time1 +'.xlsx'
        workbook.save(temp_write_filename)
        print('=' * 40)
        self.scr.insert(END, "文件输出..\n" )
        self.scr.insert(END, temp_write_filename + '\n' )
        self.master.update()
    #===========================yujingbiao end

    #按字符查找符合条件文件名，返回文件列表
    def find_filename(self, curr_path, curr_filename_path):
        list_files = []
        curr_filename_path = curr_filename_path.replace('*','')
        FileNames = os.listdir(curr_path)
        for file_name in FileNames:
            if file_name[0] == '~':
                continue
            fullfilename = os.path.join(os.path.abspath(os.path.dirname(curr_path)),file_name)
            if curr_filename_path in fullfilename:
                print('文件名：%s' % fullfilename)
                list_files.append(fullfilename)

        if len(list_files) > 0:
            return (list_files[0])
        else:
            return (None)

    # 从数据库导出价格（基础表），返回含价格信息列表
    def excel_cell_rowcell_to_position(self,int_row,int_column):
        if int_row < 26:
            str_excel_cell_pos = chr(64+int_row)
            str_excel_cell_pos = str_excel_cell_pos + str(int_column)
        return str_excel_cell_pos

# 整合数据，导出生成excel文件

# 程序主gui界面。
    def initWidgets(self, fm1):

        cp = ConfigParser()
        cp.read('配置\\配置文件.ini', encoding='gbk')
        str_kehu_name = cp.get('客户', '客户名称')

        try:
            self.customer_sname = cp.get('客户', 'sname')
            self.file_from_fuliaokucun = cp.get(str_kehu_name, '物料预警表')
            #self.Holiday = cp.get(str_kehu_name, '节假日')
            self.file_from_wuliaoshiyong = cp.get(str_kehu_name, '物料使用情况')
            self.work_dir = ".//"#cp.get(str_kehu_name, '工作目录')
            #self.file_from_fuliaokucun = cp.get(str_kehu_name, '辅料库存表')
        except Exception as err_message:
            print(err_message)
            logger.info('无法打开配置文件.ini或配置有误!' )
            exit(2)

        print('host: ', str_kehu_name)
        print(self.file_from_youjiqingdan)

        label_kehumingcheng = Label(fm1, text='客户名称：', font=('Arial', 12))
        label_kehumingcheng.place(x=20, y=30)
        self.svar_kehumingcheng.set(str_kehu_name)
        entry_kehumingcheng = Entry(fm1, textvariable=self.svar_kehumingcheng, width=20, font=('Arial', 12))
        entry_kehumingcheng.place(x=20, y=55)

        label_proc_time = Label(fm1, text='数据处理时间：', font=('Arial', 12))
        label_proc_time.place(x=240, y=30)

        temp_last_datetime = datetime.date.today() - datetime.timedelta(days=10)

        str_temp_last_datetime = time.strftime('%Y%m%d', time.localtime(time.time()))
        self.svar_proc_time1.set(str_temp_last_datetime)
        entry_proc_time1 = Entry(fm1, textvariable=self.svar_proc_time1, width=20, font=('Arial', 12))
        entry_proc_time1.place(x=240, y=55)

        label_dingdangenzong_filename = Label(fm1, text='订单跟踪表：', font=('Arial', 12))
        label_dingdangenzong_filename.place(x=540, y=30)

        str_temp_find_filename = self.find_filename(self.work_dir,self.file_from_wuliaoshiyong)
        if str_temp_find_filename == None:
            self.svar_wuliaoshiyong_filename.set('没有找到文件'+self.file_from_wuliaoshiyong)
        else:
            self.svar_wuliaoshiyong_filename.set(str_temp_find_filename)

        entry_dingdangenzong_filename = Entry(fm1, textvariable=self.svar_wuliaoshiyong_filename, width=50, font=('Arial', 12))
        entry_dingdangenzong_filename.place(x=540, y=55)


        label_wuliaojxc_filename = Label(fm1, text='辅料库存表：', font=('Arial', 12))
        label_wuliaojxc_filename.place(x=540, y=100)

        str_temp_find_filename = self.find_filename(self.work_dir,self.file_from_fuliaokucun)
        if str_temp_find_filename == None:
            self.svar_nnnnnnnnnnnnnnn_filename.set('没有找到文件'+self.file_from_wuliaoshiyong)
        else:
            self.svar_nnnnnnnnnnnnnnn_filename.set(str_temp_find_filename)

        entry_wuliaojxc_filename = Entry(fm1, textvariable=self.svar_nnnnnnnnnnnnnnn_filename, width=50, font=('Arial', 12))
        entry_wuliaojxc_filename.place(x=540, y=125)

        svar_label_prompt = StringVar()
        svar_label_prompt.set('客户名称：')

        label_author = Label(fm1, text='by流程与信息化部IT. Dec,2019', font=('Arial', 9))
        label_author.place(x=820, y=740)

        self.scr = scrolledtext.ScrolledText(fm1, width=70, height=48)
        self.scr.place(x=20, y=100)

        btn_barcode_init = Button(fm1, text='广发银行物料进销存日报', command=self.command_btn_run_wuliaojxc)
        btn_barcode_init.place(x=540, y=240)

        btn_barcode_init = Button(fm1, text='广发银行物料预警表', command=self.command_btn_run_wuliao_yujingbiao)
        btn_barcode_init.place(x=540, y=300)


        btn_barcode_init = Button(fm1, text=' 退  出 ', command=self.command_btn_exit)
        btn_barcode_init.place(x=540, y=420)

    # 退出键
    def command_btn_exit(self):
        self.master.destroy()

    # 主功能键
    def command_btn_run_wuliaojxc(self):

        self.scr.delete(1.0,END)

        label_tips1_filename = Label(self.master, text='读取订单跟踪表... ', font=('Arial', 12))
        label_tips1_filename.place(x=540, y=530)

        #self.file_from_wuliaoshiyong = self.svar_wuliaoshiyong_filename.get()
        #self.file_from_fuliaokucun = self.svar_nnnnnnnnnnnnnnn_filename.get()

        #str_timestamp = self.svar_proc_time1.get()

        str_temp_last_datetime = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        str_wuliaojxc_filename = '广发银行物料进销存日报'+'.xlsx'

        # self.wuliaojxc_file_proc(str_wuliaojxc_filename, self.file_from_wuliaoshiyong)
        try:
            if self.wuliaojxc_file_proc(str_wuliaojxc_filename, self.file_from_wuliaoshiyong) == 'no':
                return (1)
        except Exception as err_message:
            print(err_message)
            self.scr.insert(END, err_message )
            self.scr.update()
            logger.error(err_message.__str__())
            logger.exception(sys.exc_info())

        label_tips1_filename = Label(self.master, text='完成...                     ', font=('Arial', 12))
        label_tips1_filename.place(x=540, y=530)

        return 0

    def command_btn_run_wuliao_yujingbiao(self):

        self.scr.delete(1.0,END)
        label_tips1_filename = Label(self.master, text='读取辅料库存表... ', font=('Arial', 12))
        label_tips1_filename.place(x=540, y=530)
        #self.file_from_wuliaoshiyong = self.svar_wuliaoshiyong_filename.get()
        self.file_from_fuliaokucun = self.svar_nnnnnnnnnnnnnnn_filename.get()
        #str_timestamp = self.svar_proc_time1.get()
        #str_temp_last_datetime = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        str_chu_filename = 'temp'

        #self.wuliao_yujingbiao_file_proc(str_chu_filename, self.file_from_fuliaokucun)
        try:
            if self.wuliao_yujingbiao_file_proc(str_chu_filename, self.file_from_fuliaokucun) == 'no':
                return (1)
        except Exception as err_message:
            print(err_message)
            self.scr.insert(END, err_message)
            self.scr.update()
            logger.error(err_message.__str__())
            logger.exception(sys.exc_info())

        label_tips1_filename = Label(self.master, text='完成...                     ', font=('Arial', 12))
        label_tips1_filename.place(x=540, y=530)
        return 0

if __name__ == '__main__':

    set_logging()

    main_window = tkinter.Tk()

    main_window.title('光大制卡业务库存反馈文件生成工具 v.2001031701')

    # 设定窗口的大小(长 * 宽)，显示窗体居中，winfo_xxx获取系统屏幕分辨率。
    sw = main_window.winfo_screenwidth()
    sh = main_window.winfo_screenheight()
    ww = 1024
    wh = 770
    x = (sw - ww) / 2
    y = (sh - wh) / 2
    main_window.geometry("%dx%d+%d+%d" % (ww, wh, x, y))  # 这里的乘是小x
    display = App(main_window)
    main_window.mainloop()
