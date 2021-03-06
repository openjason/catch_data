"""
功能：考勤数据转换脚本。
"""
from tkinter import Tk
from tkinter import (
    messagebox,
    scrolledtext,
    Canvas,
    PhotoImage,
    Label,
    StringVar,
    Entry,
    Button,
    END,
    DISABLED,
    Toplevel,
)  # 导入滚动文本框的模块
from os.path import exists as os_path_exists
from openpyxl import load_workbook
import datetime, time
import xlrd
import os
import sys
import logging
from logging.handlers import RotatingFileHandler
import chardet

# 设置日志文件配置参数
def set_logging():
    global logger
    logger = logging.getLogger("balance_logger")
    handler = RotatingFileHandler("日志记录.log", maxBytes=5000000, backupCount=9)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)
    formatter = logging.Formatter("%(asctime)-12s %(filename)s %(lineno)d %(message)s")
    handler.setFormatter(formatter)


# 定义类，脚本主要更能
class App:
    def __init__(self, master):

        self.svar_proc_time1 = StringVar()
        self.svar_cangku_filename = StringVar()
        self.svar_kehumingcheng = StringVar()
        self.svar_kehumingcheng2 = StringVar()
        
        self.svar_proc_time2 = StringVar()
        self.svar_youjiqingdan_filename = StringVar()
        self.svar_jichu_filename = StringVar()
        self.svar_label_prompt = StringVar()
        self.master = master
        self.customer_sname = ""
        self.userid_list = []
        self.data_dir = ""
        self.file_from_cangkujxc = ""
        self.file_from_youjiqingdan = ""
        self.file_from_jichu = ""
        self.curr_month = ""
        self.curr_year_month = "2020-09-"
        self.next_year_month = "2020-10-"
        self.initWidgets(master)
        self.is_workday=[0,32]
        #是否工作日，列表内日子计算工时时按星期几的相反操作，即是否扣 120 分钟工作时

    #处理字符串存在非标准内容
    def fix_string_to_standard(self, massstring):
        txtfile_line = massstring
        txtfile_line = txtfile_line.replace("£º", ":")
        txtfile_line = txtfile_line.replace("：", ":")
        txtfile_line = txtfile_line.replace("##", "#")
        txtfile_line = txtfile_line.replace("时间", "")
        txtfile_line = txtfile_line.replace("上班", "")
        txtfile_line = txtfile_line.replace("下班", "")
        txtfile_line = txtfile_line.replace("（", "")
        txtfile_line = txtfile_line.replace("）", "")
        txtfile_line = txtfile_line.replace("(", "")
        txtfile_line = txtfile_line.replace(")", "")
        txtfile_line = txtfile_line.replace("忘记", "")
        txtfile_line = txtfile_line.replace("打卡", "")
        txtfile_line = txtfile_line.replace("迟到", "")
        txtfile_line = txtfile_line.replace("缺卡", "")
        txtfile_line = txtfile_line.replace("班车", "")
        txtfile_line = txtfile_line.replace("补卡", "")
        if massstring != txtfile_line:
            print(massstring,' --> ',txtfile_line)
            logger.info(massstring + ' --> ' + txtfile_line)
        return (txtfile_line)


    # 20200925更新，用补签卡汇总表excel 导入数据
    def fix_recorder_proc(self, work_dir):
        if len(self.userid_list) < 1:
            self.scr.insert(END, "error请先运行更新员工信息 ...\n")
            self.master.update()
            return 2
        file_proced_count = 0
        buqian_txt_file = open("buqian.txt", "w",encoding='gbk')
        for parent, dirnames, filenames in os.walk(work_dir, followlinks=True):
            for filename in filenames:
                file_path = os.path.join(parent, filename)
                if ('~$' in filename) or (not ("补签卡" in filename)):
                    logger.info("不处理文件名: " + file_path)
                    print("文件名: " + filename)
                    self.scr.insert(END, "不处理文件名: " + file_path + "...\n")
                else:
                    xlsfilename = file_path
                    print('OpenXlsFile: ',xlsfilename)
                    workbook = load_workbook(xlsfilename)
                    sheet_curr = workbook.worksheets[0]

                    int_sheet_nrows = sheet_curr.max_row
                    int_sheet_ncols = sheet_curr.max_column

                    int_first_row = 1
                    int_first_col = 6

                    userid_count_attend = 0
                    userid_count = 0
                    userid_attend = 0

                    file_proced_count = file_proced_count + 1
                    logger.info("文件名: " + file_path)
                    print("文件名: " + filename)

                    self.scr.insert(END, "处理： " + filename + "...\n")
                    self.master.update()

                    curr_proc_row = int_first_row
                    check_space_row_count = 0
                    while(1):
                        # 检查 工号 存在 begin
                        curr_proc_row = curr_proc_row +1
                        # 下一行
                        from_excel_uid = sheet_curr.cell(curr_proc_row,4).value
                        if from_excel_uid == '' or from_excel_uid ==None:
                            check_space_row_count = check_space_row_count +1
                            if check_space_row_count > 5 :
                                break
                            else:
                                continue
                        check_space_row_count = 0

                        
                        print('from_excel_uid',from_excel_uid)
                        testid = "00"
                        for temp_userid in self.userid_list:
                            temp_id = temp_userid[0]
                            if from_excel_uid == temp_id:
                                testid = from_excel_uid
                                break
                        if testid == "00":
                            self.scr.insert(END, "无匹配工号，姓名： " + str(username) + "\n")
                            self.master.update()

                        from_excel_username = sheet_curr.cell(curr_proc_row,3).value
                        from_excel_attend_date = sheet_curr.cell(curr_proc_row,5).value
                        from_excel_attend_time = sheet_curr.cell(curr_proc_row,6).value

                        # 检查 工号 存在 end
                        #txtfile_line = self.fix_string_to_standard(txtfile_line)

                        # 打卡文件 行 处理
                        #logger.info(txtfile_line)
                        #print(txtfile_line)
                        #line_proc_list = txtfile_line.split("#")

                        # 行格式错，‘#’号分隔符多于2个
                        # 检查日期格式正确 begin

                        try:
                            daka_date = from_excel_attend_date
                            if type(daka_date) == str:
                                daka_date = daka_date.strip()
                                # if daka_date.count("-") == 2:
                                #     daka_date_format = datetime.datetime.strptime(
                                #         daka_date, "%Y-%m-%d"
                                #     ).date()
                                # elif daka_date.count(".") == 2:
                                #     daka_date_format = datetime.datetime.strptime(
                                #         daka_date, "%Y.%m.%d"
                                #     ).date()
                                # else:
                                #     self.scr.insert(
                                #         END, "error日期格式错： " + str(daka_date) + "\n"
                                #     )
                                #     self.scr.update()
                                # daka_date = daka_date_format.strftime("%Y-%m-%d")
                            elif type(daka_date) != datetime.datetime:
                                self.scr.insert(END, "error日期type类型有误：  行" +str(curr_proc_row) + '  ' + str(daka_date) + str(from_excel_username) +"\n")
                                continue
                            temp_datetime_str = self.curr_year_month + "01 00:00"
                            if daka_date < datetime.datetime.strptime(temp_datetime_str, "%Y-%m-%d %H:%M") or \
                                daka_date > datetime.datetime.strptime(self.next_year_month + "01 00:00", "%Y-%m-%d %H:%M"):
                                    self.scr.insert(END, "跳过记录，日期超出范围：  行" +str(curr_proc_row) + '  ' + str(daka_date)+ str(from_excel_username) + "\n")
                                    self.scr.update()
                                    continue
                            print('daka_date ',daka_date)
                        except Exception as err_message:
                            print(err_message)
                            self.scr.insert(END, err_message)
                            self.scr.update()
                            logger.error(err_message.__str__())
                            logger.exception(sys.exc_info())
                            continue
                        # 检查日期格式正确 end

                        # 有两个分隔符‘#’，执行上午打卡，之后处理下午打卡，如果
                        daka_morning = from_excel_attend_time
                        
                        #print('daka_morning ',daka_morning)
                        #print('type(daka_morning) ',type(daka_morning))
                        try:
                            if type(daka_morning ) == str:
                                # daka_date_format = datetime.datetime.strptime("2020-01-01 " + daka_morning,"%Y-%m-%d %H:%M:%S")
                                # daka_date_format = datetime.datetime.strptime(
                                #     "2020-01-01 " + daka_morning, "%Y-%m-%d %H:%M")
                                
                                # daka_morning = (daka_date_format.strftime("%H:%M") + ":00")

                                # self.scr.insert(END, "error日期格式错: " + daka_morning)
                                # self.scr.update()
                                if type(daka_morning) != datetime.datetime:
                                    self.scr.insert(END, "error日期type类型有误： 行" +str(curr_proc_row) + '  '+ str(daka_morning) +str(from_excel_username) + "\n")
                                    continue
                        except Exception as err_message:
                            print(err_message)
                            self.scr.insert(END, err_message)
                            self.scr.update()
                            logger.error(err_message.__str__())
                            logger.exception(sys.exc_info())
                            continue
                        # 检查 上班 时间格式正确 end

                        # 检查 下班 时间格式正确 begin
                        # 检查 下班 时间格式正确 end

                        temp_userid_to_file = str(from_excel_uid)
                        temp_date_to_file = datetime.datetime.strftime(from_excel_attend_date,'%Y-%m-%d')
                        temp_time_to_file = from_excel_attend_time.strftime('%H:%M')
                        #datetime.datetime.strptime(from_excel_attend_time,'%H:%M') 

                        buqian_txt_file.write(temp_userid_to_file + " " + temp_date_to_file + " " + temp_time_to_file)
                        # buqian_txt_file.write(userid+' '+daka_date+' '+daka_morning+':00')
                        buqian_txt_file.write("\n")

                    # excel 下一行，循环

        buqian_txt_file.close()
        self.scr.insert(
            END,
            "文件总数：" + str(len(filenames)) + "处理文件数量： " + str(file_proced_count) + "记录数量： " +str(curr_proc_row) +"\n",
        )
        self.scr.update()

    # 研发中心补签卡
    def proc_yf_buka_proc(self, work_dir, xlsfilename):

        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message = messagebox.askquestion(
                title="提示", message="无找到文件" + xlsfilename + "，继续？"
            )  # return yes/no
            return return_message

        print(xlsfilename)
        workbook = load_workbook(xlsfilename)
        sheet_curr = workbook.worksheets[0]

        int_sheet_nrows = sheet_curr.max_row
        int_sheet_ncols = sheet_curr.max_column

        int_first_row = 4
        int_first_col = 6

        userid_count_attend = 0
        userid_count = 0
        userid_attend = 0

        file_proced_count = 0
        buqian_txt_file = open("buqian.txt", "w",encoding='gbk')
        for parent, dirnames, filenames in os.walk(work_dir, followlinks=True):
            for filename in filenames:
                file_path = os.path.join(parent, filename)
                if not ("#" in filename):
                    logger.info("不处理文件名: " + file_path)
                    print("文件名: " + filename)
                    self.scr.insert(END, "不处理文件名: " + file_path + "...\n")
                else:
                    file_proced_count = file_proced_count + 1
                    logger.info("文件名: " + file_path)
                    # print('文件名: '+ filename)
                    temp = filename.split("#")
                    username = temp[0]
                    userid = temp[1]
                    if userid[-4:] == ".txt":
                        userid = userid[:-4]
                        if not userid.isdigit():
                            logger.info("error工号格式错： " + filename)
                            self.scr.insert(END, "error工号格式错： " + filename + "...\n")
                            self.master.update()
                    else:
                        logger.info("error文件名格式错： " + filename)
                        self.scr.insert(END, "error文件名格式错： " + filename + "...\n")
                        self.master.update()

                    # 检查 工号 存在 begin

                    # 日数据开始位置
                    xls_catch_line = 0
                    for i in range(int_first_row, int_sheet_nrows + 1):
                        cell_curr_value = sheet_curr.cell(i, 1).value
                        xls_username = sheet_curr.cell(i, 1).value
                        logger.info(username)
                        xls_userid = sheet_curr.cell(i, 4).value
                        # print('username id',username,userid)

                        if xls_userid == userid:
                            xls_catch_line = i
                            print("i: ", i)
                            break

                    if xls_catch_line == 0:
                        continue
                    # 在excel表格内找不到对应的工号，不往下执行
                    print(userid, xls_userid, xls_catch_line)

                    self.scr.insert(END, "处理： " + filename + "...\n")
                    self.master.update()

                    # 检查 工号 存在 end

                    # 自动检测文件编码格式（utf-8, ansi,...）
                    with open(file_path, mode="rb") as det_file:
                        data_ = det_file.read(200000)
                        file_encode_detect_dict = chardet.detect(data_)
                        file_encode_detect = file_encode_detect_dict["encoding"]
                    # with open(file_path, 'r', encoding=file_encode_detect,errors='ignore') as f:
                    with open(file_path, "r", encoding=file_encode_detect) as f:
                        for txtfile_line in f.readlines():
                            txtfile_line = txtfile_line.strip()

                            if len(txtfile_line) < 5:
                                print("space line.")
                                continue
                            txtfile_line = txtfile_line.replace("£º", ":")
                            txtfile_line = txtfile_line.replace("：", ":")
                            txtfile_line = txtfile_line.replace("时间", "")
                            txtfile_line = txtfile_line.replace("上班", "")
                            txtfile_line = txtfile_line.replace("下班", "")
                            # 打卡文件 行 处理
                            logger.info(txtfile_line)
                            print(txtfile_line)
                            line_proc_list = txtfile_line.split("#")

                            # 行格式错，‘#’号分隔符多于2个
                            if len(line_proc_list) > 3:
                                logger.info("error line format error: " + txtfile_line)
                            # 检查日期格式正确 begin
                            try:
                                daka_date = line_proc_list[0]
                                daka_date = daka_date.strip()
                                if daka_date.count("-") == 2:
                                    daka_date_format = datetime.datetime.strptime(
                                        daka_date, "%Y-%m-%d"
                                    ).date()
                                elif daka_date.count(".") == 2:
                                    daka_date_format = datetime.datetime.strptime(
                                        daka_date, "%Y.%m.%d"
                                    ).date()
                                else:
                                    self.scr.insert(
                                        END, "error日期格式错： " + str(daka_date) + "\n"
                                    )
                                    self.scr.update()

                                if (
                                    daka_date_format
                                    < datetime.datetime.strptime(
                                        self.curr_year_month + "01", "%Y-%m-%d"
                                    ).date()
                                    or daka_date_format
                                    < datetime.datetime.strptime(
                                        self.next_year_month + "01", "%Y-%m-%d"
                                    ).date()
                                ):
                                    self.scr.insert(
                                        END, "error日期超出范围： " + str(daka_date) + "\n"
                                    )
                                    self.scr.update()
                                daka_date = daka_date_format.strftime("%Y-%m-%d")
                            except Exception as err_message:
                                print(err_message)
                                self.scr.insert(END, err_message)
                                self.scr.update()
                                logger.error(err_message.__str__())
                                logger.exception(sys.exc_info())
                                continue
                            # 检查日期格式正确 end

                            # 有两个分隔符‘#’，执行上午打卡，之后处理下午打卡，如果
                            if len(line_proc_list) == 3:
                                daka_morning = line_proc_list[1]
                                daka_afternoon = line_proc_list[2]
                            elif len(line_proc_list) == 2:
                                daka_morning = line_proc_list[1]
                                daka_afternoon = "afternoon_space"
                            else:
                                print("打卡数据行 ‘#’ 数量分割出错...")
                                self.scr.insert(END, "打卡数据行 ‘#’ 数量分割出错...")
                                self.scr.update()
                                logger.info("打卡数据行 ‘#’ 数量分割出错...")

                            # 检查 上班 时间格式正确 begin
                            try:
                                if daka_morning.count(":") == 2:
                                    daka_date_format = datetime.datetime.strptime(
                                        "2020-01-01 " + daka_morning,
                                        "%Y-%m-%d %H:%M:%S",
                                    )
                                    daka_morning = daka_date_format.strftime("%H:%M:%S")
                                elif daka_morning.count(":") == 1:
                                    daka_date_format = datetime.datetime.strptime(
                                        "2020-01-01 " + daka_morning, "%Y-%m-%d %H:%M"
                                    )
                                    daka_morning = (
                                        daka_date_format.strftime("%H:%M") + ":00"
                                    )
                                else:
                                    self.scr.insert(END, "error日期格式错: " + daka_morning)
                                    self.scr.update()
                            except Exception as err_message:
                                print(err_message)
                                self.scr.insert(END, err_message)
                                self.scr.update()
                                logger.error(err_message.__str__())
                                logger.exception(sys.exc_info())
                                continue
                            # 检查 上班 时间格式正确 end

                            # 检查 下班 时间格式正确 begin
                            # 如果 下午打卡时间为空，？？？？
                            if "space" not in daka_afternoon:
                                try:
                                    if daka_afternoon.count(":") == 2:
                                        daka_date_format = datetime.datetime.strptime(
                                            "2020-01-01 " + daka_afternoon,
                                            "%Y-%m-%d %H:%M:%S",
                                        )
                                        daka_afternoon = daka_date_format.strftime(
                                            "%H:%M:%S"
                                        )
                                    elif daka_afternoon.count(":") == 1:
                                        daka_date_format = datetime.datetime.strptime(
                                            "2020-01-01 " + daka_afternoon,
                                            "%Y-%m-%d %H:%M",
                                        )
                                        daka_afternoon = (
                                            daka_date_format.strftime("%H:%M") + ":00"
                                        )
                                    else:
                                        self.scr.insert(
                                            END, "error日期格式错: " + daka_morning
                                        )
                                        self.scr.update()
                                except Exception as err_message:
                                    print(err_message)
                                    self.scr.insert(END, err_message)
                                    self.scr.update()
                                    logger.error(err_message.__str__())
                                    logger.exception(sys.exc_info())
                                    continue
                            # 检查 下班 时间格式正确 end

                            buqian_txt_file.write(
                                userid + " " + daka_date + " " + daka_morning
                            )
                            # buqian_txt_file.write(userid+' '+daka_date+' '+daka_morning+':00')
                            buqian_txt_file.write("\n")
                            if len(line_proc_list) > 2:
                                # buqian_txt_file.write(userid+' '+daka_date+' '+daka_afternoon+':00')
                                buqian_txt_file.write(
                                    userid + " " + daka_date + " " + daka_afternoon
                                )
                                buqian_txt_file.write("\n")

                            excel_col = int(daka_date[-2:]) + int_first_col
                            temp_cell_value = sheet_curr.cell(
                                xls_catch_line, excel_col
                            ).value

                            if temp_cell_value == None:
                                sheet_curr.cell(
                                    xls_catch_line, excel_col
                                ).value = daka_morning[:5]
                            else:
                                sheet_curr.cell(xls_catch_line, excel_col).value = (
                                    temp_cell_value + "\n" + daka_morning[:5]
                                )

                            if len(line_proc_list) > 2:
                                if temp_cell_value == None:
                                    sheet_curr.cell(xls_catch_line, excel_col).value = (
                                        daka_morning[:5] + "\n" + daka_afternoon[:5]
                                    )
                                else:
                                    sheet_curr.cell(xls_catch_line, excel_col).value = (
                                        temp_cell_value
                                        + "\n"
                                        + daka_morning[:5]
                                        + "\n"
                                        + daka_afternoon[:5]
                                    )

        buqian_txt_file.close()
        workbook.save(xlsfilename[:-5] + "-buka.xlsx")
        self.scr.insert(END, "保存文件：" + xlsfilename[:-5] + "-buka.xlsx" + "\n")
        self.scr.insert(
            END,
            "文件总数：" + str(len(filenames)) + "处理文件数量： " + str(file_proced_count) + "\n",
        )
        self.scr.update()

    def dingding_data_ech(self, xlsfilename):
        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message = messagebox.askquestion(
                title="提示", message="无找到文件" + xlsfilename + "，继续？"
            )  # return yes/no
            return return_message
        except_attend_number = ["0022"]

        int_first_row = 3
        # 日数据开始位置
        print(xlsfilename)
        self.scr.insert(END, "钉钉数据" + xlsfilename + "...\n")

        # self.scr.insert(
        #     END, "Getatime " + str(time.ctime(os.path.getatime(xlsfilename))) + "...\n"
        # )
        # self.scr.insert(
        #     END, "Getmtime " + str(time.ctime(os.path.getmtime(xlsfilename))) + "...\n"
        # )
        # self.scr.insert(
        #     END, "Getctime " + str(time.ctime(os.path.getctime(xlsfilename))) + "...\n"
        # )
        # self.master.update()

        str_kaoqin_date_begin = self.svar_kehumingcheng2.get()
        str_kaoqin_date_end = self.svar_kehumingcheng.get()

        self.scr.insert(END,'处理考勤时间段 开始:'+str_kaoqin_date_begin+' 结束:（钉钉考勤表最新日期）')

        workbook = xlrd.open_workbook(xlsfilename)
        sheetsname = workbook.sheet_names()  # 获取excel里的工作表sheet名称数组

        str_curr_sheet_name = sheetsname[0]

        sheet_curr = workbook.sheet_by_name(str_curr_sheet_name)
        int_sheet_nrows = sheet_curr.nrows
        int_sheet_ncols = sheet_curr.ncols

        logger.info("sheet size:")
        logger.info(int_sheet_nrows)
        logger.info(int_sheet_ncols)

        print("sheetname & lines:", str_curr_sheet_name, int_sheet_nrows)

        int_first_col = 6
        file_txt_kaoqin = open("nc.txt", "w",encoding='gbk')

        match_list = []
        self.curr_year_month = str_kaoqin_date_begin[:8]
        # self.scr.insert(END, "\n\n\n注意： " + str(self.curr_year_month) + "...\n\n\n")
        # self.master.update()

        userid_count_attend = 0
        userid_count = 0
        userid_attend = 0

        self.scr.insert(
            END, "\nExcel File Title: " + sheet_curr.cell(0, 0).value + "...\n"
        )

        for i in range(int_first_row, int_sheet_nrows):
            cell_curr_value = sheet_curr.cell(i, 0).value
            # print('i: ',i)
            username = sheet_curr.cell(i, 0).value
            logger.info(username)
            userid = sheet_curr.cell(i, 3).value

            # 匹配工号

            if userid:
                userid_count = userid_count + 1
                userid_attend = 0
                daka_line = ""
                for j in range(int_first_col, int_sheet_ncols):

                    # 插入数据
                    cell_value = sheet_curr.cell(i, j).value
                    # print(cell_value)
                    cell_value = cell_value.replace("\n", "*")
                    cell_value_cut = cell_value.replace("\n", "*")
                    cell_value = cell_value_cut.replace(" ", "")
                    # logger.info(cell_value)
                    if len(cell_value) > 0:
                        daka_line = daka_line + str(cell_value) + "@" + str(j - 5) + "#"
                        click_one_times = cell_value.split("*")
                        for click_one_time in click_one_times:
                            if "外勤" in click_one_time:
                                continue
                            if (j - 5) > 9:
                                curr_day = str(j - 5)
                            else:
                                curr_day = "0" + str(j - 5)
                            # file_txt_kaoqin.write(username+ userid+" "+self.curr_year_month+curr_day+' '+click_one_time+':00')
                            if type(userid) != str:
                                self.scr.insert(
                                    END, "ERROR: 工号格式不正确，需为字符串   userid type error: 错误工号 " + str(userid) + username +"\n"
                                )
                            click_one_time_hour_str = click_one_time[:2]
                            #凌晨时段数据处理，如凌晨6点前，设为次日
                            #打卡记录为不是第一个记录的才设置，即不是当日首次打卡记录
                            str_checkin = self.curr_year_month + curr_day + " " + click_one_time + ":00"
                            datetime_checkin = datetime.datetime.strptime(str_checkin,'%Y-%m-%d %H:%M:%S')

                            if click_one_times.index(click_one_time) >=1 :
                                if click_one_time_hour_str < '06':
                                    datetime_checkin = datetime_checkin + datetime.timedelta(days=1)

                            str_checkin = datetime.datetime.strftime(datetime_checkin,'%Y-%m-%d %H:%M:%S')
                            #logger.info(userid)
                            if str_checkin >= str_kaoqin_date_begin:                            
                                file_txt_kaoqin.write(userid + " " + str_checkin)
                                file_txt_kaoqin.write("\n")
                                userid_attend = 1
                if userid_attend == 1:
                    userid_count_attend = userid_count_attend + 1

                logger.info(daka_line)

        file_txt_kaoqin.close()
        print("=" * 40)
        print("共导入了 ", i - int_first_row + 1, "行数据.")
        self.scr.insert(END, "共导入了 .." + str(i - int_first_row + 1) + "行数据..\n")
        self.scr.insert(END, "有工号 .." + str(userid_count) + "行数据..\n")
        self.scr.insert(END, "有工号正常打卡 .." + str(userid_count_attend) + "行数据..\n")

        self.master.update()

    # 钉钉上通信录 员工 工号 补全
    def fix_txl_id_run_proc(self, xlsfilename):
        if len(self.userid_list) < 1:
            self.scr.insert(END, "error请先运行更新员工信息 ...\n")
            self.master.update()
            return 2

        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message = messagebox.askquestion(
                title="提示", message="无找到文件" + xlsfilename + "，继续？"
            )  # return yes/no
            return return_message
        int_first_row = 4
        # 日数据开始位置
        print(xlsfilename)
        self.scr.insert(END, "钉钉数据" + xlsfilename + "...\n")
        self.master.update()

        workbook = load_workbook(xlsfilename)  # 打开excel文件
        worksheetj = workbook["员工信息表1"]  # 根据Sheet1这个sheet名字来获取该sheet
        # worksheetj.cell(1, 1).value = str(xls_date.year)+'年广发银行'+str(xls_date.month)+' 月物料收发进销存日报表'
        # worksheetj.cell(3, 19).value = datetime.datetime.strptime(curr_proc_time_str[:6]+'01','%Y%m%d')

        # workbook = openpyxl.open_workbook(xlsfilename)
        # sheetsname = workbook.sheet_names()  # 获取excel里的工作表sheet名称数组
        # str_curr_sheet_name = sheetsname[0]

        match_list = []
        i = int_first_row
        while True:

            if not worksheetj.cell(i, 1).value:
                break

            dingdinguserid = worksheetj.cell(i, 1).value
            username = worksheetj.cell(i, 2).value

            logger.info(dingdinguserid)
            if len(dingdinguserid) < 1:
                break
            # 匹配工号
            userid = "00"
            for temp_userid in self.userid_list:
                temp_j = temp_userid[1]
                if username == temp_j:
                    userid = temp_userid[0]
                    break
            if userid == "00":
                self.scr.insert(END, "无匹配工号，姓名： " + str(username) + "\n")
                self.master.update()

            else:
                if username not in match_list:
                    match_list.append(username)
                else:
                    self.scr.insert(END, "重名.." + str(username) + "\n")
                    self.master.update()

                dingding_id_line = worksheetj.cell(i, 6).value
                if len(dingding_id_line) < 2:
                    dingding_id_line = userid
                    worksheetj.cell(i, 6).value = dingding_id_line
                    logger.info(dingding_id_line)
            i = i + 1

        print("=" * 40)
        print("共导入了 ", i - int_first_row + 1, "行数据.")
        self.scr.insert(END, "共导入了 .." + str(i - int_first_row + 1) + "行数据..\n")
        self.master.update()

        workbook.save("kaoqinwenj.xlsx")

    # 下载txl上通信录 员工 工号 补全
    def download_txl_id_run_proc(self, xlsfilename):
        if len(self.userid_list) < 1:
            self.scr.insert(END, "error请先运行更新员工信息 ...\n")
            self.master.update()
            return 2

        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message = messagebox.askquestion(
                title="提示", message="无找到文件" + xlsfilename + "，继续？"
            )  # return yes/no
            return return_message
        int_first_row = 4
        # 日数据开始位置
        print(xlsfilename)
        self.scr.insert(END, "钉钉数据" + xlsfilename + "...\n")
        self.master.update()

        workbook = load_workbook(xlsfilename)  # 打开excel文件
        worksheetj = workbook["员工信息表1"]  # 根据Sheet1这个sheet名字来获取该sheet
        # worksheetj.cell(1, 1).value = str(xls_date.year)+'年广发银行'+str(xls_date.month)+' 月物料收发进销存日报表'
        # worksheetj.cell(3, 19).value = datetime.datetime.strptime(curr_proc_time_str[:6]+'01','%Y%m%d')

        # workbook = openpyxl.open_workbook(xlsfilename)
        # sheetsname = workbook.sheet_names()  # 获取excel里的工作表sheet名称数组
        # str_curr_sheet_name = sheetsname[0]

        match_list = []
        i = int_first_row
        while True:

            if not worksheetj.cell(i, 1).value:
                break

            dingdinguserid = worksheetj.cell(i, 1).value
            username = worksheetj.cell(i, 2).value

            logger.info(dingdinguserid)
            if len(dingdinguserid) < 1:
                break
            # 匹配工号
            userid = "00"
            for temp_userid in self.userid_list:
                temp_j = temp_userid[1]
                if username == temp_j:
                    userid = temp_userid[0]
                    break
            if userid == "00":
                self.scr.insert(END, "无匹配工号，姓名： " + str(username) + "\n")
                self.master.update()

            else:
                if username not in match_list:
                    match_list.append(username)
                else:
                    self.scr.insert(END, "重名.." + str(username) + "\n")
                    self.master.update()

                dingding_id_line = worksheetj.cell(i, 6).value
                if len(dingding_id_line) < 2:
                    dingding_id_line = userid
                    worksheetj.cell(i, 6).value = dingding_id_line
                    logger.info(dingding_id_line)
            i = i + 1

        print("=" * 40)
        print("共导入了 ", i - int_first_row + 1, "行数据.")
        self.scr.insert(END, "共导入了 .." + str(i - int_first_row + 1) + "行数据..\n")
        self.master.update()

        workbook.save("kaoqinwenj.xlsx")

    def bool_need_deduction_2hour(self,checkin_time):
        #是否需要扣减2小时
        
        int_weekday_click = checkin_time.weekday()
        check_work_day = checkin_time.day

        if int_weekday_click < 5:
            b_need_deduction = True
        else:
            b_need_deduction = False
        if check_work_day in self.is_workday:
                b_need_deduction = not b_need_deduction



        return(b_need_deduction)

        
    # 研发中心特殊处理
    def proc_yf_proc(self, xlsfilename):
        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message = messagebox.askquestion(
                title="提示", message="无找到文件：" + xlsfilename + "，继续？"
            )  # return yes/no
            return return_message

        int_first_row = 4
        # 日数据开始位置
        print(xlsfilename)
        self.scr.insert(END, "钉钉数据" + xlsfilename + "...\n")

        self.scr.insert(
            END, "Getatime " + str(time.ctime(os.path.getatime(xlsfilename))) + "...\n"
        )
        self.scr.insert(
            END, "Getmtime " + str(time.ctime(os.path.getmtime(xlsfilename))) + "...\n"
        )
        self.scr.insert(
            END, "Getctime " + str(time.ctime(os.path.getctime(xlsfilename))) + "...\n"
        )
        self.master.update()

        workbook = load_workbook(xlsfilename)
        # sheetsname = workbook[1]  # 获取excel里的工作表sheet名称数组

        # str_curr_sheet_name = sheetsname[0]
        sheet_curr = workbook.worksheets[0]

        int_sheet_nrows = sheet_curr.max_row
        int_sheet_ncols = sheet_curr.max_column

        logger.info("sheet size:")
        logger.info(int_sheet_nrows)
        logger.info(int_sheet_ncols)

        # print('sheetname & lines:', str_curr_sheet_name, int_sheet_nrows)

        int_first_col = 6
        file_txt_kaoqin = open("nc.txt", "w",encoding='gbk')
        match_list = []

        self.scr.insert(END, "\n\n\n注意： " + str(self.curr_year_month) + "...\n\n\n")
        self.master.update()

        userid_count_attend = 0
        userid_count = 0
        userid_attend = 0

        self.scr.insert(
            END, "\nExcel File Title: " + sheet_curr.cell(1, 1).value + "...\n"
        )

        for i in range(int_first_row, int_sheet_nrows + 1):
            cell_curr_value = sheet_curr.cell(i, 1).value
            print("i: ", i)
            username = sheet_curr.cell(i, 1).value
            logger.info(username)
            userid = sheet_curr.cell(i, 4).value
            print("username id", username, userid)

            # 匹配工号
            work_time_minute_int = 0
            work_time_str = ""

            if userid:
                userid_count = userid_count + 1
                userid_attend = 0
                daka_line = ""
                for j in range(int_first_col, int_sheet_ncols + 1):

                    # 插入数据
                    cell_value = str(sheet_curr.cell(i, j).value)
                    # print(cell_value)
                    if cell_value == None:
                        continue
                    cell_value = cell_value.replace("\n", "*")
                    cell_value_cut = cell_value.replace("\n", "*")
                    cell_value = cell_value_cut.replace(" ", "")
                    # logger.info(cell_value)
                    if len(cell_value) > 0:
                        daka_line = daka_line + str(cell_value) + "@" + str(j - 5) + "#"
                        click_one_times = cell_value.split("*")
                        #click_one_times.sort() #20200615 不能排序，凌晨下班打卡是次日，排序则变成当天
                        print(click_one_times)
                        # 打卡时间列表排序
                        if len(click_one_times) < 2:
                            print(
                                "一天打卡小于2次，工号： ",
                                userid,
                                "col: ",
                                j,
                                "记录：",
                                click_one_times,
                            )
                            continue
                        for click_one_time in click_one_times:
                            if "外勤" in click_one_time:
                                print("含外勤打卡： ", click_one_times)
                                continue
                        if (j - int_first_col) > 9:
                            curr_day = str(j - int_first_col)
                        else:
                            curr_day = "0" + str(j - int_first_col)

                        first_click = click_one_times[0]
                        last_click = click_one_times[-1]
                        print(
                            "click_day: ",
                            int(self.curr_year_month[:4]),
                            int(self.curr_year_month[5:7]),
                            int(curr_day),
                        )
                        first_time = datetime.datetime(
                            int(self.curr_year_month[:4]),
                            int(self.curr_year_month[5:7]),
                            int(curr_day),
                            int(first_click[:2]),
                            int(first_click[-2:]),
                            0,
                            0,
                        )

                        last_time = datetime.datetime(
                            int(self.curr_year_month[:4]),
                            int(self.curr_year_month[5:7]),
                            int(curr_day),
                            int(last_click[:2]),
                            int(last_click[-2:]),
                            0,
                            0,
                        )
                        #如第二次打开时间为4点以前，算第二天
                        if int(last_click[:2]) < 4:
                            last_time = last_time + datetime.timedelta(days=1)
                            
                        weekday_click = first_time.weekday()
                        check_work_day = first_time.day
                        #if check_work_day == 28:
                        #    print('day:28')
                        temp_min_time = (last_time - first_time).seconds / 60

                        if temp_min_time > 16*60:
                            print('工作时间超16小时，请核查。。。')
                            print('UserID:',userid,'  DAY:',curr_day)
                            self.scr.insert(END, "工作时间超16小时，请核查... UserID:" + str(userid) + '  DAY:' + str(curr_day) +".\n")

                        if self.bool_need_deduction_2hour(first_time):
                            #判断打卡时间是否为中午时间
                            #上班打卡
                            deduction_minutes = 120
                            str_noon_checkin = datetime.datetime.strftime(first_time,'%H:%M')
                            #print('str_noon_check:',str_noon_checkin)
                            if str_noon_checkin > '11:30':
                                if str_noon_checkin < '13:30':
                                    deduction_minutes = 13*60 + 30 - int(str_noon_checkin[:2]) * 60 - int(str_noon_checkin[-2:])
                            #temp_min_time = temp_min_time - deduction_minutes
                            #下班打卡
                            str_noon_checkout = datetime.datetime.strftime(last_time,'%H:%M')
                            if str_noon_checkout > '11:30':
                                if str_noon_checkout < '13:30':
                                    deduction_minutes = int(str_noon_checkout[:2]) * 60 + int(str_noon_checkout[-2:]) - 11*60 - 30
                            if str_noon_checkin > '13:30':
                                print(str_noon_checkin,'upper than 13:30')
                                deduction_minutes = 0
                            print('deduction_minutes:',deduction_minutes)
                            temp_min_time = temp_min_time - deduction_minutes


                        work_time_minute_int = work_time_minute_int + temp_min_time
                        work_time_str = (
                            work_time_str + "+" + str(round(temp_min_time / 60, 2))
                        )

                        # file_txt_kaoqin.write(username+ userid+" "+self.curr_year_month+curr_day+' '+click_one_time+':00')
                        file_txt_kaoqin.write(
                            userid
                            + " "
                            + str(curr_day)
                            + " "
                            + click_one_time
                            + ":00"
                            + "@ "
                            + str(temp_min_time)
                        )
                        file_txt_kaoqin.write("\n")
                        userid_attend = 1

                print("工作时长 ", str(round(work_time_minute_int / 60, 2)))

                sheet_curr.cell(i, 38).value = round(work_time_minute_int / 60, 2)
                sheet_curr.cell(i, 40).value = work_time_str

                if userid_attend == 1:
                    userid_count_attend = userid_count_attend + 1

                logger.info(daka_line)

        file_txt_kaoqin.close()
        workbook.save("东信和平科技股份有限公司_打卡时间表-研发-工时.xlsx")
        print("=" * 40)
        print("共导入了 ", i - int_first_row + 1, "行数据.")
        self.scr.insert(END, "共导入了 .." + str(i - int_first_row + 1) + "行数据..\n")
        self.scr.insert(END, "有工号 .." + str(userid_count) + "行数据..\n")
        self.scr.insert(END, "有工号正常打卡 .." + str(userid_count_attend) + "行数据..\n")

        self.master.update()

    # 从数据库导入价格（基础表），返回含价格信息列表
    def user_id_import_list(self, xlsfilename):

        if not os_path_exists(xlsfilename):
            print("文件不存在：", xlsfilename)
            return_message = messagebox.askquestion(
                title="提示", message="无找到文件" + xlsfilename + "，继续？"
            )  # return yes/no
            return return_message

        int_first_row = 1
        # day_column_start = 7  # 日数据开始位置
        self.userid_list = []

        print("清空原有staff 数据...")
        print(xlsfilename)
        self.scr.insert(END, "清空列表（staff）数据...\n")
        self.master.update()

        workbook = xlrd.open_workbook(xlsfilename)
        sheetsname = workbook.sheet_names()  # 获取excel里的工作表sheet名称数组

        str_curr_sheet_name = sheetsname[0]

        sheet_curr = workbook.sheet_by_name(str_curr_sheet_name)
        int_sheet_nrows = sheet_curr.nrows
        print("sheetname & lines:", str_curr_sheet_name, int_sheet_nrows)

        for i in range(int_first_row, int_sheet_nrows):
            cell_curr_value = sheet_curr.cell(i, 0).value
            # print('i: ',i)
            if True:  # not isinstance(cell_curr_value,str):         #判断数据是否最后一行
                userid = sheet_curr.cell(i, 0).value
                username = sheet_curr.cell(i, 1).value

                if int(len(userid)) > 0:  # testing
                    # 插入数据
                    self.userid_list.append([userid, username])
                else:
                    self.scr.insert(
                        END, "ERROR基础数据表（price）数据导入: " + str(username) + ".\n"
                    )
                    self.master.update()
        #logger.info(self.userid_list)
        print("=" * 40)
        logger.info("共导入了 " + str(i - int_first_row + 1) + "行数据.")
        print("共导入了 ", i - int_first_row + 1, "行数据.")
        self.scr.insert(END, "共导入了.." + str(i - int_first_row + 1) + "行数据..\n")
        self.master.update()

        match_list = []
        self.scr.insert(END, "检查重名：\n")

        for temp_userid in self.userid_list:
            temp_j = temp_userid[1]
            if temp_j not in match_list:
                match_list.append(temp_j)
            else:
                self.scr.insert(END, "重名.." + str(temp_j) + "\n")
                self.master.update()

    # 程序主gui界面。
    def initWidgets(self, fm1):


        self.customer_sname = "ep"
        kehu_conf_jxc = "仓库进销存"
        self.Holiday = "节假日"
        self.file_from_cangkujxc = "仓库进销存"
        self.file_from_youjiqingdan = "邮寄清单"
        self.file_from_jichu = "基础数据文件"

        # print('host: ', str_kehu_name)
        # print(self.file_from_youjiqingdan)

        temp_curr_datetime = datetime.datetime.now()
        if temp_curr_datetime.day < 5:
            temp_curr_datetime = temp_curr_datetime - datetime.timedelta(days=6)
        str_kehu_name = temp_curr_datetime.strftime('%Y-%m-%d')
        str_kehu_name2 = (temp_curr_datetime- datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        str_kehu_name = '钉钉考勤表最新日期'

        # label_kehumingcheng = Label(fm1, text='客户名称：', font=('Arial', 12))
        # label_kehumingcheng.place(x=20, y=30)
        self.svar_kehumingcheng.set(str_kehu_name)
        #entry_kehumingcheng = Entry(fm1, textvariable=self.svar_kehumingcheng, width=20, font=('Arial', 12))
        #entry_kehumingcheng.place(x=620, y=125)

        self.svar_kehumingcheng2.set(str_kehu_name2)
        entry_kehumingcheng2 = Entry(fm1, textvariable=self.svar_kehumingcheng2, width=20, font=('Arial', 12))
        entry_kehumingcheng2.place(x=620, y=100)

        label_proc_time = Label(fm1, text='请输入处理的考勤开始日期，例子：2020-08-09', font=('Arial', 12))
        label_proc_time.place(x=620, y=70)

        temp_last_datetime = datetime.date.today() - datetime.timedelta(days=10)

        label_author = Label(fm1, text="by流程与信息化部IT. May,2020", font=("Arial", 9))
        label_author.place(x=820, y=770)

        self.scr = scrolledtext.ScrolledText(fm1, width=80, height=54)
        self.scr.place(x=20, y=30)

        btn_id_import_init = Button(
            fm1, text="更新员工信息", command=self.command_id_import_run
        )
        btn_id_import_init.place(x=620, y=200)

        btn_dingding_exchage_run = Button(fm1, text="钉钉数据转换", command=self.command_dingding_ech_run)
        btn_dingding_exchage_run.place(x=620, y=270)

        btn_fix_rec_run = Button(
            fm1, text="补 签 卡", command=self.command_fix_recorder_run
        )
        btn_fix_rec_run.place(x=620, y=340)

        # btn_fix_userid_run = Button(fm1, text='匹配通信录工号', command=self.command_fix_txl_id_run)
        # btn_fix_userid_run.place(x=620, y=410)

        btn_fix_userid_run = Button(fm1, text="研发中心算工时", command=self.command_fix_yf_run)
        btn_fix_userid_run.place(x=620, y=410)

        btn_fix_userid_run = Button(
            fm1, text="研发补卡>xls", command=self.command_fix_yf_buka_run
        )
        btn_fix_userid_run.place(x=620, y=510)

        btn_barcode_init = Button(fm1, text=" 退  出 ", command=self.command_btn_exit)
        btn_barcode_init.place(x=620, y=690)

    # 退出键
    def command_btn_exit(self):
        self.master.destroy()

    # 导入员工工号
    def command_id_import_run(self):
        # 功能停用
        # return(0)

        label_tips1_filename = Label(
            self.master, text="正在导入员工工号数据... ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

        userid_filename = "在职人员信息表.xls"
        self.user_id_import_list(userid_filename)

        label_tips1_filename = Label(
            self.master, text="完成...                     ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

    # 补签卡
    def command_fix_recorder_run(self):
        label_tips1_filename = Label(
            self.master, text="正在导入补签卡数据... ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

        work_dir = "补签卡\\"
        self.fix_recorder_proc(work_dir)

        label_tips1_filename = Label(
            self.master, text="完成...                     ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

    # 匹配员工工号
    def command_fix_yf_run(self):

        dingding_userid_filename = "东信和平科技股份有限公司_打卡时间表-yf-buka.xlsx"

        self.proc_yf_proc(dingding_userid_filename)

        label_tips1_filename = Label(
            self.master, text="完成...                     ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

    # 研发中心 补卡
    def command_fix_yf_buka_run(self):

        dingding_userid_filename = "东信和平科技股份有限公司_打卡时间表-yf.xlsx"

        work_dir = "补签卡\\"

        self.proc_yf_buka_proc(work_dir, dingding_userid_filename)

        label_tips1_filename = Label(
            self.master, text="完成...                     ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

    # 匹配员工工号
    def command_fix_txl_id_run(self):
        label_tips1_filename = Label(
            self.master, text="正在匹配员工工号数据... ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=530)

        dingding_userid_filename = "东信和平科技股份有限公司-通讯录.xlsx"

        self.download_txl_id_run_proc(dingding_userid_filename)

        label_tips1_filename = Label(
            self.master, text="完成...                     ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

    # 钉钉数据转NC 文本文件
    def command_dingding_ech_run(self):

        label_tips1_filename = Label(
            self.master, text="正在处理钉钉数据（只导出含工号的数据）... ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

        file_from_dingding = "东信和平科技股份有限公司_打卡时间表.xlsx"

        try:
            self.dingding_data_ech(file_from_dingding)
        except Exception as err_message:
            print(err_message)
            self.scr.insert(END, err_message)
            self.scr.update()
            logger.error(err_message.__str__())
            logger.exception(sys.exc_info())


        label_tips1_filename = Label(
            self.master, text="完成...                     ", font=("Arial", 12)
        )
        label_tips1_filename.place(x=620, y=590)

        return 0


if __name__ == "__main__":

    set_logging()

    main_window = Tk()
    main_window.title("临时考勤数据处理工具 v.20200615")

    # 设定窗口的大小(长 * 宽)，显示窗体居中，winfo_xxx获取系统屏幕分辨率。
    sw = main_window.winfo_screenwidth()
    sh = main_window.winfo_screenheight()
    ww = 1024
    wh = 800
    x = (sw - ww) / 2
    y = (sh - wh) / 2
    main_window.geometry("%dx%d+%d+%d" % (ww, wh, x, y))  # 这里的乘是小x
    display = App(main_window)
    main_window.mainloop()
