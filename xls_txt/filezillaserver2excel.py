#！usr/bin/python3
# -*- coding: utf-8 -*-
#==========================
import xml.etree.ElementTree as ET
from openpyxl import Workbook

from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border,Side
from openpyxl.styles import Alignment

# 合并拆分单元格
# 再次编辑中，这次是在使用删除列的时候发现，合并单元格会出现漏删除情况，才想到用拆分单元格，没想到unmerge_cells（），worksheet.merged_cells返回的合并单元格对象居然不能迭代，
# 函数参数也变了，居然可以直接上参数 ；；openpyxl = Version: 2.5.9；
# 列 ： worksheet.delete_cols(2, 1)
# 表示第二列开始，删除一列  ，行
# worksheet.delete_rows(2, 1)
# worksheet.unmerge_cells(start_row=1, start_column=7, end_row=2, end_column=7)
# 表示第一行开始，第二行结束， 低7列开始第七列结束， 就是把G1: G2合并的单元格给拆分了，下面的是合并单元格就不多说了
# worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
# worksheet.merged_cells获取已经合并单元格的信息；再使用worksheet.unmerge_cells（）拆分单元格；
# ws.merge_cells(‘A2:A10’)
# 合并A2到A10之间的单元格
align = Alignment(horizontal='left',vertical='center',wrap_text=False)
# ws.[‘D1’].alignment = align
# horizontal代表水平方向，可以左对齐left，还有居中center和右对齐right，分散对齐distributed，跨列居中centerContinuous，两端对齐justify，填充fill，常规general
# vertical代表垂直方向，可以居中center，还可以靠上top，靠下bottom，两端对齐justify，分散对齐distributed
# 另外还有自动换行：wrap_text，这是个布尔类型的参数，这个参数还可以写作wrapText

fill = PatternFill(fill_type = None,start_color='FFFFFF',end_color='000000')
#ws.[‘B1’].fill = fill
# 这里官方文档给出了很多种填充类型（类似于我们操作excel表格 fill_type; start_color代表前景色，end_color是背景色）
# ‘none’、‘solid’、‘darkDown’、‘darkGray’、‘darkGrid’、‘darkHorizontal’、‘darkTrellis’、‘darkUp’、‘darkVertical’、‘gray0625’、‘gray125’、‘lightDown’、‘lightGray’、‘lightGrid’、‘lightHorizontal’、‘lightTrellis’、‘lightUp’、‘lightVertical’、‘mediumGray’

font = Font(u'宋体',size = 11,bold=True,italic=True,strike=True,color='000000')
#ws.[‘A1’].font = font

border = Border(left=Side(border_style='thin',color='000000'),
right=Side(border_style='thin',color='000000'),
top=Side(border_style='thin',color='000000'),
bottom=Side(border_style='thin',color='000000'))
#ws.[‘C1’].border = border
# 注意这里需要导入Border和Side两个函数
# 边框的样式有很多，官方给出的样式如下：
# ‘dashDot’,‘dashDotDot’,‘dashed’,‘dotted’,‘double’,‘hair’,‘medium’,‘mediumDashDot’,‘mediumDashDotDot’,‘mediumDashed’,‘slantDashDot’,‘thick’,‘thin’
# 注意，如果没有定义边框的样式，那么后面的参数将没有作用
# 另外，边框不只left，right，top，bottom，官方还给出了几个参数，我这里不详细讲了，目测对角线什么的也不一定能用到
# ‘diagonal’,‘diagonal_direction’,‘vertical’,‘horizontal’



option_dict = {'FileRead': 'FR', 'FileWrite': 'W', 'FileDelete': 'D', 'FileAppend': 'A', \
               'DirCreate': 'DC', 'DirDelete': 'D', 'DirList': 'L', 'DirSubdirs': 'S', \
               'IsHome': 'IH', 'AutoCreate': 'AC'}
value_dic = {'1': '+', '0':'-'}
def Optionformat(optionname,value):
    rt = option_dict[optionname] + value_dic[value]
    return (rt)
wb = Workbook()
ws = wb.create_sheet("User Sheet",0)
xuhao = 1
ws.cell(row=1, column=1).value = '序号'
ws.column_dimensions['A'].width=5
ws.cell(row=1, column=2).value ='户名'
ws.column_dimensions['B'].width=10
ws.cell(row=1, column=3).value = '所属组名称'
ws.column_dimensions['C'].width=30
ws.cell(row=1, column=4).value = '文件夹(组成员请查看对应组权限)'
ws.column_dimensions['D'].width=30
ws.cell(row=1, column=5).value = '权限：FileRead,FileWrite，FileDelete，FileAppend，DirCreate，DirDelete，DirList，DirSubdirs,IsHome,AutoCreate'
ws.column_dimensions['E'].width=35


#ws = wb.active
cell_row=1
last_rownum = cell_row
tree = ET.parse('FileZillaServer.xml')
root = tree.getroot()
print('root-tag:',root.tag,',root-attrib:',root.attrib,',root-text:',root.text)
for child in root:
    print('child-tag是：',child.tag,',child.attrib：',child.attrib,',child.text：',child.text)
    if child.tag != 'Users':
        continue
    for sub in child:
        print('sub-tag是：',sub.tag,',sub.attrib：',sub.attrib,',sub.text：',sub.text)
        if sub.tag == 'User':
            cell_row = cell_row + 1
            if cell_row > last_rownum + 1 :
                ws.merge_cells(start_row=last_rownum, start_column=2, end_row=cell_row - 1, end_column=2)
                ws.merge_cells(start_row=last_rownum, start_column=1, end_row=cell_row - 1, end_column=1)
                ws.merge_cells(start_row=last_rownum, start_column=3, end_row=cell_row - 1, end_column=3)
            last_rownum = cell_row
            ws.cell(row=cell_row, column=2).value = str(sub.attrib['Name'])
            ws.cell(row=cell_row, column=1).value = str(xuhao)
            xuhao = xuhao + 1
        for sub2 in sub:
            print('sub-tag2是：',sub2.tag,',sub2.attrib：',sub2.attrib,',sub2.text：',sub2.text)
            #ws.cell(row=cell_row, column=3).value = str(sub2.text)
            try:
                #if sub2.attrib['Name'] = {}:
                if sub2.attrib['Name'] == 'Group':
                    ws.cell(row=cell_row, column=3).value = str(sub2.text)

                    if str(sub2.text) != 'None':
                        pass
                    else:
                        ws.cell(row=cell_row, column=3).value =''
            except:
                pass
            row_next = 0
            for sub3 in sub2:
                print('sub-tag3是：',sub3.tag,',sub3.attrib：',sub3.attrib,',sub3.text：',sub3.text)
                row_next = 0
                if sub3.tag == 'Permission':
                    try:
                        ws.cell(row=cell_row, column=4).value = str(sub3.attrib['Dir'])
                        ws.cell(row=cell_row, column=4).border = border
                        row_next = 1
                    except:
                        pass
                for sub4 in sub3:
                    print('sub-tag4是：',sub4.tag,',sub4.attrib：',sub4.attrib,',sub4.text：',sub4.text)
                    if sub4.tag == 'Option':
                        try:
                            #ws.cell(row=cell_row, column=5).value = str(ws.cell(row=cell_row, column=5).value) + str(sub4.attrib['Name']) + str(sub4.text)
                            if ws.cell(row=cell_row, column=5).value == None:
                                ws.cell(row=cell_row, column=5).value = Optionformat(sub4.attrib['Name'] , str(sub4.text))
                                ws.cell(row=cell_row, column=5).border = border
                            else:
                                ws.cell(row=cell_row, column=5).value = str(ws.cell(row=cell_row, column=5).value) + Optionformat(sub4.attrib['Name'] , str(sub4.text))
                                ws.cell(row=cell_row, column=5).border = border
                        except:
                            pass
                if row_next == 1:
                    cell_row = cell_row +1
            if row_next == 1:
                cell_row = cell_row -1

cell_row = cell_row + 1
if cell_row > last_rownum + 1 :
    ws.merge_cells(start_row=last_rownum, start_column=2, end_row=cell_row - 1, end_column=2)
    ws.merge_cells(start_row=last_rownum, start_column=1, end_row=cell_row - 1, end_column=1)
    ws.merge_cells(start_row=last_rownum, start_column=3, end_row=cell_row - 1, end_column=3)


for i in range(1,cell_row):
    for j in range(1,6):
        ws.cell(row=i, column=j).border = border
        ws.cell(row=i, column=j).alignment = align


ws = wb.create_sheet("Group sheet",0)
cell_row=1
last_rownum = cell_row
tree = ET.parse('FileZillaServer.xml')
root = tree.getroot()
print('root-tag:',root.tag,',root-attrib:',root.attrib,',root-text:',root.text)
xuhao = 1
ws.cell(row=1, column=1).value = '序号'
ws.column_dimensions['A'].width=5
ws.cell(row=1, column=2).value ='组名称'
ws.column_dimensions['B'].width=15
ws.cell(row=1, column=3).value = '文件夹'
ws.column_dimensions['C'].width=50
ws.cell(row=1, column=4).value = '权限：FileRead,FileWrite，FileDelete，FileAppend，DirCreate，DirDelete，DirList，DirSubdirs,IsHome,AutoCreate'
ws.column_dimensions['D'].width=35
for child in root:
    print('child-tag是：',child.tag,',child.attrib：',child.attrib,',child.text：',child.text)
    if child.tag != 'Groups':
        continue
    for sub in child:
        print('sub-tag是：',sub.tag,',sub.attrib：',sub.attrib,',sub.text：',sub.text)
        if sub.tag == 'Group':
            cell_row = cell_row + 1

            if cell_row > last_rownum + 1 :
                ws.merge_cells(start_row=last_rownum, start_column=2, end_row=cell_row -1, end_column=2)
                ws.merge_cells(start_row=last_rownum, start_column=1, end_row=cell_row -1, end_column=1)
            last_rownum = cell_row

            ws.cell(row=cell_row, column=2).value = str(sub.attrib['Name'])
            ws.cell(row=cell_row, column=2).border = border
            ws.cell(row=cell_row, column=1).value = str(xuhao)
            ws.cell(row=cell_row, column=1).border = border
            xuhao = xuhao +1
        for sub2 in sub:
            print('sub-tag2是：',sub2.tag,',sub2.attrib：',sub2.attrib,',sub2.text：',sub2.text)
            #ws.cell(row=cell_row, column=3).value = str(sub2.text)
            try:
                #if sub2.attrib['Name'] = {}:
                if sub2.attrib['Name'] == 'Group':
                    ws.cell(row=cell_row, column=3).value = str(sub2.text)
                    ws.cell(row=cell_row, column=3).border = border
            except:
                pass

            for sub3 in sub2:
                print('sub-tag3是：',sub3.tag,',sub3.attrib：',sub3.attrib,',sub3.text：',sub3.text)
                row_next = 0
                if sub3.tag == 'Permission':
                    try:
                        ws.cell(row=cell_row, column=3).value = str(sub3.attrib['Dir'])
                        ws.cell(row=cell_row, column=3).border = border
                        row_next = 1
                    except:
                        pass

                for sub4 in sub3:
                    print('sub-tag4是：',sub4.tag,',sub4.attrib：',sub4.attrib,',sub4.text：',sub4.text)
                    if sub4.tag == 'Option':
                        try:
                            #ws.cell(row=cell_row, column=5).value = str(ws.cell(row=cell_row, column=5).value) + str(sub4.attrib['Name']) + str(sub4.text)
                            if ws.cell(row=cell_row, column=4).value == None:
                                ws.cell(row=cell_row, column=4).value = Optionformat(sub4.attrib['Name'] , str(sub4.text))
                                ws.cell(row=cell_row, column=4).border = border
                            else:
                                ws.cell(row=cell_row, column=4).value = str(ws.cell(row=cell_row, column=4).value) + Optionformat(sub4.attrib['Name'] , str(sub4.text))
                                ws.cell(row=cell_row, column=4).border = border
                        except:
                            pass
                if row_next == 1:
                    cell_row = cell_row +1
            if row_next == 1:
                cell_row = cell_row -1

cell_row = cell_row + 1
if cell_row > last_rownum + 1 :
    ws.merge_cells(start_row=last_rownum, start_column=2, end_row=cell_row - 1, end_column=2)
    ws.merge_cells(start_row=last_rownum, start_column=1, end_row=cell_row - 1, end_column=1)

for i in range(1,cell_row):
    for j in range(1,5):
        ws.cell(row=i, column=j).border = border
        ws.cell(row=i, column=j).alignment = align

wb.save('FTP用户表.xlsx')
