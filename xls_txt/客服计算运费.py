'''
版本: V0.01
功能：
1．  对excel表格遍历（xls和xlsx），日期列，数量列，同一日期合计数量，数量小于10的，在运费列加一个10.72
2．  保存。

'''

from tkinter import Tk
from tkinter import filedialog
from configparser import ConfigParser
from tkinter import messagebox,scrolledtext,Canvas,PhotoImage,Label,StringVar,Entry, Button,END, DISABLED, Toplevel  # 导入滚动文本框的模块
from os.path import exists as os_path_exists
import xlrd
import datetime
from openpyxl import load_workbook,Workbook
import os,sys 
import logging
from logging import getLogger
from logging.handlers import RotatingFileHandler
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font  #设置字体和边框需要的模块

#xl_border = Border(left=Side(style='thin',color='FF000000'),right=Side(style='thin',color='FF000000'),top=Side(style='thin',color='FF000000'),bottom=Side(style='thin',color='FF000000'),diagonal=Side(style='thin',color='FF000000'),diagonal_direction=0,outline=Side(style='thin',color='FF000000'),vertical=Side(style='thin',color='FF000000'),horizontal=Side(style='thin',color='FF000000'))
#font_set = Font(name='宋体', size=24, italic=True, color=colors.BLUE, bold=True, underline='doubleAccounting')
font_set = Font(name='宋体', size=10)

#设置日志文件配置参数
def set_logging():
    global logger
    logger = getLogger('balance_logger')
    handler = RotatingFileHandler('日志记录.log', maxBytes=2000000, backupCount=3)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)
    formatter = logging.Formatter('%(asctime)-12s %(filename)s %(lineno)d %(message)s')
    handler.setFormatter(formatter)

#定义类，脚本主要更能
class App():
    def __init__(self, master):

        self.master = master
        self.customer_sname = ''
        self.curr_month = ''
        self.autorun = ''
        self.filesymbol = ''
        self.pendingdir = ''
        self.savefilename = ''
        self.initWidgets(master)


    #按字符查找符合条件文件名，返回文件列表
    def find_filename(self, curr_path, curr_filename_path):
        list_files = []
        for parent, dirnames, filenames in os.walk(curr_path, followlinks=True):
            for filename in filenames:
                file_path = os.path.join(parent, filename)
                if curr_filename_path in filename:
                    print('文件名：%s' % file_path)
                    list_files.append(file_path)
        if len(list_files) > 0:
            return (list_files)
        else:
            return (None)


    #从数据库处理数据，导出对账文件excel
    def count_express_fee(self, work_dir, xlsfilename):
        self.scr.insert(1.0, "处理时间: "+datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')+ "\n")
        logger.info("处理时间: "+datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        self.scr.insert(1.0, "待合并文件所在文件夹: "+str(work_dir)+ "\ \n")
        file_lists = self.find_filename(work_dir,'xls')

        self.scr.insert(1.0, "======================================================= "+ "\n")
        self.master.update()


        #xlsfilename = self.savefilename

        #打开待合并xlsx文件
        row_proc_count = 0
        xl=load_workbook(xlsfilename)#打开excel
            #xl_sheet_names=xl.get_sheet_names()#获取所有sheet页名字
            #print(xl_sheet_names)#打印所有sheet页名称
            #xl_sheet=xl.get_sheet_by_name(xl_sheet_names[0])#定位到相应sheet页,[0]为sheet页索引
        xl_sheet=xl.worksheets[0]
        xl_sheet_max_row=xl_sheet.max_row    #获取行列数
        #column=xl_sheet.max_column
        date_counter_list = []
        print(xl_sheet_max_row)
        for xls_row in range(1+1,xl_sheet_max_row+1):  #五列：卡片序号	卡片号段号	卡片号段序号	排序状态	卡号
            if xl_sheet.cell(xls_row,1).value==None:
                break
            if len(xl_sheet.cell(xls_row,1).value) < 3:
                break
            bank_branch_name = xl_sheet.cell(xls_row,1).value
            bank_branch_name_str = str(bank_branch_name)
            bank_branch_name_str = bank_branch_name_str.strip()
            
            one_row_date = xl_sheet.cell(xls_row,5).value
            one_row_date_str = str(one_row_date)
            one_row_date_str = one_row_date_str[:10]

            one_row_count = xl_sheet.cell(xls_row,6).value

            try:
                int_temp = int(one_row_count)
            except Exception as err_message:
                print(err_message)
                self.scr.insert(1.0, str(err_message) +'\n')
                xls_row
                self.scr.insert(1.0, '出错, 行 '+ str(xls_row) + ' | ' + str(bank_branch_name_str)+'\n')
                self.scr.insert(1.0, '注意, 请处理后从新运行。。。\n')
                self.scr.update()
                logger.error(err_message.__str__())
                logger.exception(sys.exc_info())
                return 2
            
            bank_branch_name_n_date_str = bank_branch_name_str + one_row_date_str

            logger.info(bank_branch_name_n_date_str + ' | ' + str(one_row_count))

            for i in range(0,len(date_counter_list)):
                bank_one_record = date_counter_list[i]
                if bank_branch_name_n_date_str == bank_one_record[0]:
                    bank_one_record_count = bank_one_record[1]
                    bank_one_record_count = bank_one_record_count + one_row_count
                    date_counter_list[i]=[bank_branch_name_n_date_str,bank_one_record_count]

            date_counter_list.append ([bank_branch_name_n_date_str,one_row_count])

            #merage_sheet_last_row = merage_sheet_last_row +1
            row_proc_count = row_proc_count +1
        #print(date_counter_list)
        print('date_counter_list length: ',len(date_counter_list))
        logger.info(date_counter_list)
        print('date_counter_list length: ' + str(len(date_counter_list)))
        for i in range(0,len(date_counter_list)):
            bank_one_record = date_counter_list[i]
            list_bank_name = bank_one_record[0]
            list_bank_count = bank_one_record[1]
            if list_bank_count >= 10: #数量大于10 的 不计费
                logger.info('数量大于>=10 的 不计费')
                logger.info(bank_one_record)
            else:
                logger.info('数量小于10 计费')
                logger.info(bank_one_record)
                for xls_row in range(1+1,xl_sheet_max_row+1):       #搜索excel表格，填入运费
                    if xl_sheet.cell(xls_row,1).value==None:
                        break
                    if len(xl_sheet.cell(xls_row,1).value) < 3:
                        break
                    bank_branch_name = xl_sheet.cell(xls_row,1).value
                    bank_branch_name_str = str(bank_branch_name)
                    bank_branch_name_str = bank_branch_name_str.strip()
                    
                    one_row_date = xl_sheet.cell(xls_row,5).value
                    one_row_date_str = str(one_row_date)
                    one_row_date_str = one_row_date_str[:10]

                    one_row_count = xl_sheet.cell(xls_row,6).value
                    if one_row_count == None:
                        one_row_count = 0
                    
                    bank_branch_name_n_date_str = bank_branch_name_str + one_row_date_str
                    if bank_branch_name_n_date_str == list_bank_name:
                        xl_sheet.cell(xls_row,8).value = 10.72
                        logger.info('update: '+ bank_branch_name_n_date_str + ' | ' + str(list_bank_count))
                        self.scr.insert(1.0,str(i)+' 更新数据:  '+ bank_branch_name_n_date_str + ' | ' + str(list_bank_count) + '\n')
                        self.scr.update()
                        #print(str(i)+' 更新数据:  '+ bank_branch_name_n_date_str + ' | ' + str(list_bank_count) )
                        break


        #print('已处理文件：' + str('one_file') +' 含数据行数： ' + str(xls_row))
        logger.info('已处理数据行： ' + str('ok...'))
        self.scr.insert(1.0, '完成' + '\n')

        logger.info('导出 ~明细表~ 表' )

        filename_extension = os.path.splitext(xlsfilename)
        len_ext = 0- len(filename_extension[1])
        xlsfilename = xlsfilename[:len_ext]
        xlsfilename = xlsfilename +'-含运费.xlsx'

        xl.save(xlsfilename)

        self.scr.insert(1.0, "保存文件： "+str(xlsfilename)+ "\n")

        self.scr.insert(1.0, "完成时间: "+datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')+ "\n")
        logger.info("完成时间: "+datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))


# 保存excel文件

# 程序主gui界面。
    def initWidgets(self, fm1):

        cp = ConfigParser()
        cp.read('配置文件.ini', encoding='gbk')
        try:
            self.autorun = '否'#cp.get('文件合并工具', 'autorun')
            self.filesymbol = '.xls'#cp.get('文件合并工具', 'filesymbol')
            self.pendingdir = cp.get('文件合并工具', 'pendingdir')
            self.savefilename = cp.get('文件合并工具', 'savefilename')
        except Exception as err_message:
            print(err_message)
            return_message = messagebox.showinfo(title='提示',message='无法打开配置文件.ini或配置有误!' )
            exit(2)

        #label_kehumingcheng = Label(fm1, text='客户名称：', font=('Arial', 12))
        #label_kehumingcheng.place(x=20, y=30)
        #self.svar_kehumingcheng.set(str_kehu_name)
        #entry_kehumingcheng = Entry(fm1, textvariable=self.svar_kehumingcheng, width=30, font=('Arial', 12))
        #entry_kehumingcheng.place(x=20, y=55)

        label_author = Label(fm1, text='by流程与信息化部IT. March,2020', font=('Arial', 9))
        label_author.place(x=500, y=777)

        self.scr = scrolledtext.ScrolledText(fm1, width=131, height=58)
        self.scr.place(x=10, y=10)

        btn_barcode_init = Button(fm1, text='打开文件', command=self.command_btn_run)
        btn_barcode_init.place(x=946, y=160)

        btn_barcode_init = Button(fm1, text=' 退  出 ', command=self.command_btn_exit)
        btn_barcode_init.place(x=946, y=270)

    # 退出键
    def command_btn_exit(self):
        self.master.destroy()

    # 主功能键
    def command_btn_run(self):
        self.scr.delete(1.0,END)
        
        cgnsfname = filedialog.askopenfilename(title='平安银行对账表文件',filetypes=[('Excel文件','*.xls?')])

        if cgnsfname == None or cgnsfname =='':
            return 1
        
        try:
            self.count_express_fee(self.pendingdir,cgnsfname)
        except Exception as err_message:
            print(err_message)
            self.scr.insert(1.0, str(err_message) +'\n')
            xls_row
            self.scr.insert(1.0, '出错, 行 '+ str(xls_row) + ' | ' + str(bank_branch_name_str)+'\n')
            self.scr.insert(1.0, '注意, 请处理后从新运行。。。\n')
            self.scr.update()
            logger.error(err_message.__str__())
            logger.exception(sys.exc_info())
            return 1



if __name__ == '__main__':

    set_logging()
    main_window = Tk()
    main_window.title('EXCEL文件 计算运费小工具  v.20200331')

    # 设定窗口的大小(长 * 宽)，显示窗体居中，winfo_xxx获取系统屏幕分辨率。
    sw = main_window.winfo_screenwidth()
    sh = main_window.winfo_screenheight()
    ww = 1024
    wh = 800
    x = (sw - ww) / 2
    y = (sh - wh) / 2
    main_window.geometry("%dx%d+%d+%d" % (ww, wh, x, y))  # 这里的乘是小x
    display = App(main_window)
    main_window.mainloop()
